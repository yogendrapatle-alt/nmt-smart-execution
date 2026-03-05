

import subprocess
from subprocess import PIPE
from flask import Flask, request, jsonify, g
from flask_socketio import SocketIO, emit
import re
import requests
import os
import json
import logging
import uuid
import time
import traceback
import signal
import atexit
import sys


# ---------------- Logging Setup ----------------
# Use relative path for cross-platform compatibility
log_file_path = os.path.join(os.path.dirname(__file__), 'nmt_ui.log')
logging.basicConfig(
    filename=log_file_path,
    level=logging.INFO,   # <-- capture ALL levels
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)

# Optional: also print INFO+ logs to console
console = logging.StreamHandler()
console.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
console.setFormatter(formatter)
logging.getLogger("").addHandler(console)
# ------------------------------------------------

# Email schedule functionality moved to email_routes.py blueprint
import random
from datetime import datetime, timedelta
# from run_backend.run_backen import get_hostport_and_ip

# Database imports
# from database import SessionLocal, init_db, save_alerts_to_db
# from models.alert import Alert

# Register Blueprints
from routes.alerts import alerts_bp
from routes.email_routes import email_routes
from routes.environment_routes import environment_routes


# Initialize Scheduler Service
from services.scheduler_service import get_scheduler
host_port = None
# Get scheduler instance (will be initialized when needed)
scheduler_service = None  # Will be set after app starts
try:
    from dateutil.parser import parse as parse_datetime
except ImportError:
    def parse_datetime(dt_str):
        # Fallback: try strptime for basic ISO8601 (no timezone)
        return datetime.strptime(dt_str.split('Z')[0], "%Y-%m-%dT%H:%M:%S")
from flask_cors import CORS
from copy_ncm_utils import KubeRemoteClient




app = Flask(__name__)
# Initialize SocketIO for real-time logs
socketio = SocketIO(app, cors_allowed_origins=["http://localhost:3000", "http://localhost:5173"], async_mode='threading')

# Register socketio with smart_execution_service
from services import smart_execution_service
smart_execution_service.set_socketio(socketio)

# Configure CORS to allow DELETE method
CORS(app, resources={
    r"/api/*": {
        "origins": ["http://localhost:3000", "http://localhost:5173"],
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type"],
        "supports_credentials": True
    }
})

# --- Config API Endpoints ---
from database import save_config_to_db, fetch_latest_config_for_pc_ip, fetch_latest_config_for_testbed, init_db, SessionLocal, save_workload_to_db, save_testbed_to_db, fetch_workloads_by_pc_ip, fetch_testbeds_by_pc_ip, fetch_workload_by_uuid, fetch_testbed_by_uuid, fetch_all_testbed_labels, fetch_latest_testbed_by_label, fetch_all_workload_labels, fetch_latest_workload_by_label, fetch_workloads_by_testbed_label, fetch_latest_workload_by_testbed_label, save_env_run_to_db, fetch_env_run_by_uuid, save_dynamic_workload_to_db, fetch_dynamic_workloads_by_uuid,fetch_latest_workload_by_unique_testbed_id, update_testbed_deployment_info, fetch_testbed_by_unique_id, recover_orphaned_executions, get_pool_status, log_pool_status
from models.config import Config
from models.testbed import Testbed

@app.route('/api/save-config', methods=['POST'])
def save_config():
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    unique_testbed_id = data.get("unique_testbed_id")
    pc_ip_from_request = data.get('pc_ip')
    config_json = data.get('config')
    
    # Extract pc_ip from config if not in request
    pc_ip = pc_ip_from_request or config_json.get('Config', {}).get('pc_ip')
    
    logging.info(f"Printing unique testbed id: {unique_testbed_id}")
    logging.info(f"PC IP: {pc_ip}")

    # Require either unique_testbed_id OR pc_ip
    if not unique_testbed_id and not pc_ip:
        return jsonify({'error': 'Missing unique_testbed_id or pc_ip - at least one is required'}), 400
    

    try:    
        # If unique_testbed_id is provided, check if the testbed exists
        if unique_testbed_id:
            testbed = fetch_testbed_by_unique_id(g.db, unique_testbed_id)
            if not testbed:
                logging.error(f"Testbed with unique_testbed_id {unique_testbed_id} not found in database")
                return jsonify({'error': f'Testbed with ID {unique_testbed_id} not found. Please save the testbed first.'}), 404
            # Set pc_ip to None for testbed-based configs (legacy behavior)
            pc_ip_for_db = None
        else:
            # Direct onboarding scenario - use pc_ip from config
            logging.info(f"Direct onboarding mode - using pc_ip: {pc_ip}")
            pc_ip_for_db = pc_ip
        
        unique_rule_id = str(uuid.uuid4())
        save_config_to_db(g.db, unique_rule_id, unique_testbed_id, pc_ip_for_db, config_json)
        
        return jsonify({'success': True, 'unique_rule_id': unique_rule_id})
            
    except Exception as db_error:
        logging.error(f"Failed to save rule-config to database: {str(db_error)}")
        return jsonify({'error': f'Failed to save config: {str(db_error)}'}), 500

@app.route('/api/fetch-config', methods=['GET'])
def fetch_config():
    pc_ip = request.args.get('pc_ip')
    if not pc_ip:
        return jsonify({'error': 'Missing pc_ip'}), 400
    config = fetch_latest_config_for_pc_ip(g.db, pc_ip)
    if not config:
        return jsonify({'error': 'No config found for this PC-IP'}), 404
    return jsonify({'success': True, 'config': config.config_json})

# Register Blueprints
app.register_blueprint(alerts_bp)
app.register_blueprint(email_routes)
app.register_blueprint(environment_routes)

# Register test routes for debugging
from routes.test_routes import test_routes
app.register_blueprint(test_routes)

# Register AI Smart Execution routes
from routes.smart_execution_ai_routes import smart_execution_ai_bp
app.register_blueprint(smart_execution_ai_bp)

# Phase 3: Register Scheduled Execution routes
from routes.scheduled_execution_routes import scheduled_execution_bp
app.register_blueprint(scheduled_execution_bp)

# Phase 3: Register Multi-Testbed Orchestration routes
from routes.multi_testbed_routes import multi_testbed_bp
app.register_blueprint(multi_testbed_bp)

# Phase 3 Stage 5: Register Cost Tracking routes
from routes.cost_routes import cost_bp
app.register_blueprint(cost_bp)

# Phase 3 Stage 6: Register Analytics routes
from routes.analytics_routes import analytics_bp
app.register_blueprint(analytics_bp)

# Register Cleanup routes (for removing fake data)
from routes.cleanup_routes import cleanup_bp
app.register_blueprint(cleanup_bp)

# Register Migration routes (for running database migrations)
from routes.migration_routes import migration_bp
app.register_blueprint(migration_bp)

## Removed APScheduler setup and scheduler import (not used)

# Initialize the database (create tables if not exist)
init_db()

# Recover any orphaned executions from previous backend restart/crash
logging.info("🔍 Checking for orphaned executions from previous backend restart...")
recovered_count = recover_orphaned_executions()
if recovered_count > 0:
    logging.warning(f"⚠️  Recovered {recovered_count} orphaned execution(s) - marked as FAILED")
else:
    logging.info("✅ No orphaned executions found - all clear")

# Log initial connection pool status
log_pool_status()

# Provide a session per request
@app.before_request
def create_session():
    g.db = SessionLocal()

@app.teardown_request
def remove_session(exception=None):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def is_prometheus_exposed(pc_ip):
    try:
        url = f"http://{pc_ip}:9090"
        response = requests.get(url, timeout=2)
        return response.status_code == 200
    except requests.RequestException:
        return False

@app.route('/api/expose-prometheus', methods=['POST'])
def expose_prometheus():
    data = request.json
    pc_ip = data.get('pcIp')
    username = data.get('username')
    password = data.get('password')


    #ip_port, ip = get_hostport_and_ip()
    #print(f"IP and Port number is {ip_port} and {ip}")
    if not pc_ip or not username or not password:
        return jsonify({'error': 'Missing pc_ip, username, or password'}), 400

    # Always run the NCM script when continue is pressed
    try:
        logging.info(f"=== NCM SETUP START for {pc_ip} ===")

        # Configuration for NCM setup - using environment variable with fallback
        kubeconfig_path = os.environ.get('KUBECONFIG_PATH', '/home/nutanix/ncm.cfg')
        prom_namespace = "ntnx-system"
        svc_name = "prometheus-k8s"
        new_svc_name = "prometheus-automation"

        # Initialize the KubeRemoteClient
        kube_client = KubeRemoteClient(pc_ip, username, password)

        # Print the host (VM) IP using run_command
        try:
            host_ip = kube_client.run_command("hostname -I | awk '{print $1}'")
            logging.info(f"[NCM] Remote host (VM) IP: {host_ip}")
        except Exception as host_ip_error:
            logging.error(f"[NCM] Failed to get remote host IP: {host_ip_error}")

        # Get NCM IP and node information
        logging.info("[NCM] Getting NCM IP and node information...")
        # ncm_ip, ncm_node = kube_client.get_ncm_ip_and_node(kubeconfig_path)
        ncm_ip, ncm_node = kube_client.get_ncm_ip_and_node()
        logging.info(f"[NCM] Found NCM IP: {ncm_ip}, Node: {ncm_node}")

        # Expose the Prometheus service (if not already exposed)
        logging.info("[NCM] Exposing Prometheus service...")
        try:
            kube_client.expose_service( prom_namespace, svc_name, new_svc_name)
            logging.info("[NCM] Service exposed successfully")
        except Exception as expose_error:
            # Service might already exist, that's okay
            logging.warning(f"[NCM] Service already exists or exposure failed: {str(expose_error)}")

        # Get the exposed port
        logging.info("[NCM] Getting exposed port...")
        node_port = kube_client.get_port(new_svc_name, prom_namespace)
        logging.info(f"[NCM] Prometheus port: {node_port}")

        # Get PC UUID
        logging.info("[NCM] Getting PC UUID...")
        pc_uuid = kube_client.get_pc_uuid()
        logging.info(f"[NCM] PC UUID: {pc_uuid}")

        # Clean up SSH connection
        kube_client.close()

        # Determine the final endpoint
        if node_port and node_port.strip():
            endpoint = f"http://{ncm_ip}:{node_port.strip()}"
            port_info = node_port.strip()
        else:
            # Fallback to standard port if NodePort not available
            endpoint = f"http://{ncm_ip}:9090"
            port_info = "9090"
            logging.info("[NCM] Falling back to standard Prometheus port 9090")

        logging.info(f"=== NCM SETUP COMPLETE: http://{ncm_ip}:{node_port.strip()} ===")

        return jsonify({
            'endpoint': endpoint,
            'ncm_ip': ncm_ip,
            'ncm_node': ncm_node,
            'node_port': port_info,
            'pc_uuid': pc_uuid.strip() if pc_uuid else None,
            'message': 'NCM setup completed successfully'
        })

    except Exception as e:
        logging.error(f"[ERROR] NCM setup failed: {str(e)}")
        return jsonify({'error': f'Failed to run NCM setup: {str(e)}'}), 500







# New direct way of deploying config copy
@app.route('/api/deploy-config-immediate', methods=['POST'])
def deploy_config_immediate():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        unique_testbed_id = data.get("unique_testbed_id")
        unique_rule_id = data.get("unique_rule_id")
        config = data.get('config')
        timestamp = data.get('timestamp')
        wait_for_jita = data.get('wait_for_jita', False)  # New parameter to control behavior
        pc_ip_from_request = data.get('pc_ip')
        
        # Extract pc_ip from config if not in request
        pc_ip = pc_ip_from_request or config.get('Config', {}).get('pc_ip')
        
        if not config or not timestamp:
            return jsonify({'error': 'Missing config or timestamp'}), 400
        
        # Require either unique_testbed_id OR pc_ip
        if not unique_testbed_id and not pc_ip:
            return jsonify({'error': 'Missing unique_testbed_id or pc_ip - at least one is required'}), 400
        
        # Save JSON to temporary file in the backend directory
        filename = f"nmt_config_{timestamp.replace(':', '-').replace('.', '-')}.json"
        
        # Ensure configs directory exists
        configs_dir = os.path.join(os.path.dirname(__file__), 'configs')
        os.makedirs(configs_dir, exist_ok=True)
        
        filepath = os.path.join(configs_dir, filename)
        
        # Save config to file
        logging.info(f"Saving config to temporary file: {filepath}")
        with open(filepath, 'w') as f:
            json.dump(config, f, indent=2)
        
        # Determine env status file path based on unique_testbed_id or pc_ip
        if unique_testbed_id:
            env_status_file_path = f"/mnt/data/nmt_backend/jita_env_status_{unique_testbed_id}.json"
        else:
            # Direct onboarding - no env status file needed
            env_status_file_path = ""
            logging.info(f"Direct onboarding mode with pc_ip: {pc_ip} - no env status file needed")
        
        if wait_for_jita:
            # Scenario 2: Wait for JITA job completion before deploying
            if not unique_testbed_id:
                return jsonify({'error': 'wait_for_jita requires unique_testbed_id'}), 400
            return _deploy_config_after_jita_completion(filepath, env_status_file_path, unique_testbed_id, unique_rule_id)
        else:
            # Scenario 1: Deploy immediately without waiting for JITA
            return _deploy_config_immediately(filepath, env_status_file_path, unique_rule_id)
        
    except Exception as e:
        logging.error(f"Failed to deploy config: {str(e)}")
        return jsonify({'error': f'Failed to deploy config: {str(e)}'}), 500


def _deploy_config_immediately(filepath, env_status_file_path="", unique_rule_id=None):
    """Deploy config immediately without waiting for JITA job completion"""
    try:
        run_backend_script = "/mnt/data/nmt_backend/main.py"  # Use main.py instead of run_backend.py
        
        # Generate a unique_rule_id if not provided (since save-config happens after deployment)
        if not unique_rule_id:
            unique_rule_id = str(uuid.uuid4())
            logging.info(f"Generated unique_rule_id for immediate deployment: {unique_rule_id}")
        
        # Check if running in development mode (main.py doesn't exist locally)
        if not os.path.exists(run_backend_script):
            logging.warning(f"⚠️  Development mode: {run_backend_script} not found. Skipping actual Prometheus deployment.")
            logging.info(f"📋 Rule config saved to: {filepath}")
            logging.info(f"✅ In production, this would deploy to Prometheus")
            
            # Return success for development mode
            return jsonify({
                'success': True,
                'message': 'Config saved successfully (Development mode - Prometheus deployment skipped)',
                'deployment_type': 'development',
                'config_file': filepath,
                'note': 'Rules saved to database. In production, they would be deployed to Prometheus.'
            }), 200
        
        logging.info(f"Executing immediate deployment: python3 {run_backend_script} {filepath} {env_status_file_path} {unique_rule_id}")
        logging.info(f"File exists check - main.py: {os.path.exists(run_backend_script)}")
        logging.info(f"File exists check - config file: {os.path.exists(filepath)}")
        
        # Execute main.py for immediate deployment (Scenario 1) with all three parameters
        result = subprocess.run(
            ['python3', run_backend_script, filepath, env_status_file_path, unique_rule_id],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            universal_newlines=True,
            cwd='/mnt/data/nmt_backend'
        )
        
        # Print the output and any errors
        logging.info("STDOUT:\n%s", result.stdout)
        if result.stderr:
            logging.error("STDERR:\n%s", result.stderr)
        
        # Parse output for prometheus host and port
        host_port = None
        host_ip = None

        for line in result.stdout.splitlines():
            # Extract host_port from: "Host port for prometheus_10_36_199_44: 9117"
            m_port = re.match(r"Host port for [^:]+: (\d+)", line)
            if m_port:
                host_port = int(m_port.group(1))
                continue

            # Extract host_ip from: "Host IP: 10.53.61.226"
            m_ip = re.match(r"Host IP: ([\d\.]+)", line)
            if m_ip:
                host_ip = m_ip.group(1)
                continue

        # Save prometheus info for later retrieval
        if host_port and host_ip:
            prometheus_port_file = os.path.join(os.path.dirname(__file__), 'prometheus_port.json')
            with open(prometheus_port_file, 'w') as f:
                json.dump({'host_port': host_port, 'host_ip': host_ip}, f)
                logging.info(f"[DEPLOY] Host Port: {host_port}, Host IP: {host_ip}")

        if result.returncode == 0:
            # Clean up the temporary file
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({
                'success': True, 
                'message': 'Config deployed immediately',
                'deployment_type': 'immediate',
                'stdout': result.stdout,
                'stderr': result.stderr,
                'host_port': host_port,
                'host_ip': host_ip
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Deployment failed with return code {result.returncode}',
                'deployment_type': 'immediate',
                'stdout': result.stdout,
                'stderr': result.stderr
            }), 500
            
    except Exception as e:
        logging.error(f"Error in immediate deployment: {str(e)}")
        return jsonify({'error': f'Failed to deploy immediately: {str(e)}'}), 500


def _deploy_config_after_jita_completion(filepath, env_status_file_path, unique_testbed_id, unique_rule_id=None):
    """Deploy config after JITA job completion by starting a background monitoring process"""
    try:
        run_backend_script = "/mnt/data/nmt_backend/main.py"
        
        # Generate a unique_rule_id if not provided (since save-config happens after deployment)
        if not unique_rule_id:
            unique_rule_id = str(uuid.uuid4())
            logging.info(f"Generated unique_rule_id for wait-for-jita deployment: {unique_rule_id}")
        
        # Check if running in development mode (main.py doesn't exist locally)
        if not os.path.exists(run_backend_script):
            logging.warning(f"⚠️  Development mode: {run_backend_script} not found. Skipping JITA-based Prometheus deployment.")
            logging.info(f"📋 Rule config saved to: {filepath}")
            
            # Return success for development mode
            return jsonify({
                'success': True,
                'message': 'Config saved successfully (Development mode - JITA deployment skipped)',
                'deployment_type': 'development',
                'unique_testbed_id': unique_testbed_id,
                'note': 'Rules saved to database. In production, they would be deployed after JITA job completion.'
            })
        
        logging.info(f"Starting background deployment monitoring for testbed: {env_status_file_path}")
        logging.info(f"Will monitor: {env_status_file_path}")
        logging.info(f"Will execute: python3 {run_backend_script} {filepath} {env_status_file_path} {unique_rule_id}")
        
        # Start background process to monitor JITA status and deploy when complete
        subprocess.Popen(
            ['python3', run_backend_script, filepath, env_status_file_path, unique_rule_id],
            cwd='/mnt/data/nmt_backend',
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        
        return jsonify({
            'success': True,
            'message': 'Prometheus deployment queued - will deploy after JITA job completion',
            'deployment_type': 'wait_for_jita',
            'unique_testbed_id': unique_testbed_id,
            'monitoring_file': env_status_file_path
        })
        
    except Exception as e:
        logging.error(f"Error starting background deployment: {str(e)}")
        return jsonify({'error': f'Failed to start background deployment: {str(e)}'}), 500

@app.route('/api/deployment-config-status', methods=['GET'])
def get_deployment_config_status():
    """Get JITA job status and trigger rule-config deployment when JITA job succeeds"""
    try:
        unique_testbed_id = request.args.get("unique_testbed_id")
        unique_rule_id = request.args.get("unique_rule_id")
        if not unique_testbed_id:
            return jsonify({
                'success': False,
                'error': 'Missing unique_testbed_id'
            }), 400

        # Check JITA job status file
        jita_status_file = f'/mnt/data/nmt_backend/jita_env_status_{unique_testbed_id}.json'
        
        jita_status = {
            'task_id': None,
            'status': 'not_started',
            'job_profile_id': None,
            'job_profile_name': None,
            'timestamp': None,
            'pc_uuid': None,
            'rule_config_triggered': False,
            'rule_config_status': None
        }
        
        # Read JITA job status from file
        if os.path.exists(jita_status_file):
            logging.info(f"Reading JITA status from file: {jita_status_file}")
            with open(jita_status_file, 'r') as f:
                jita_data = json.load(f)
                logging.info(f"JITA status data: {jita_data}")
                jita_status['task_id'] = jita_data.get('task_id')
                jita_status['status'] = jita_data.get('status', 'unknown')
                jita_status['job_profile_id'] = jita_data.get('job_profile_id')
                jita_status['timestamp'] = jita_data.get('timestamp')
                jita_status['pc_uuid'] = jita_data.get('pc_uuid')
                
                # Extract job profile name from job_profile_id if available
                if jita_data.get('job_profile_id'):
                    jita_status['job_profile_name'] = f"Job Profile {jita_data.get('job_profile_id')}"
        else:
            logging.info(f"JITA status file not found: {jita_status_file}")
        
        # Check if JITA job has succeeded and trigger rule-config deployment
        if jita_status['status'].lower() in ['completed', 'succeeded']:
            _trigger_rule_config_deployment(unique_testbed_id, jita_status, unique_rule_id)
        
        # Check if rule-config deployment has been triggered
        rule_config_triggered_file = f'/mnt/data/nmt_backend/rule_config_triggered_{unique_testbed_id}.json'
        if os.path.exists(rule_config_triggered_file):
            with open(rule_config_triggered_file, 'r') as f:
                rule_config_data = json.load(f)
                jita_status['rule_config_triggered'] = True
                jita_status['rule_config_status'] = 'completed' if rule_config_data.get('deployment_result') else 'failed'
        
        return jsonify({
            'success': True,
            'jita_job_status': jita_status
        })
        
    except Exception as e:
        logging.error(f"Error getting JITA job status: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Failed to get JITA job status: {str(e)}'
        }), 500


def _trigger_rule_config_deployment(unique_testbed_id, jita_status, unique_rule_id=None):
    """Trigger rule-config deployment when JITA job succeeds"""
    try:
        # Check if rule-config deployment has already been triggered
        rule_config_triggered_file = f'/mnt/data/nmt_backend/rule_config_triggered_{unique_testbed_id}.json'
        
        if os.path.exists(rule_config_triggered_file):
            logging.info(f"Rule-config deployment already triggered for testbed {unique_testbed_id}")
            return
        
        # Always fetch config from database for this testbed
        db_session = SessionLocal()
        try:
            # Get the latest config for this testbed
            config = fetch_latest_config_for_testbed(db_session, unique_testbed_id)
            if not config:
                logging.warning(f"No config found for testbed {unique_testbed_id}")
                return
            
            # Use the passed unique_rule_id or get from config if not provided
            if not unique_rule_id:
                # Get the unique_rule_id from the config
                unique_rule_id = config.unique_rule_id
                if not unique_rule_id:
                    logging.warning(f"No unique_rule_id found in config for testbed {unique_testbed_id}")
                    return
            else:
                logging.info(f"Using provided unique_rule_id: {unique_rule_id}")
                
        finally:
            db_session.close()
        
        # Trigger rule-config deployment
        logging.info(f"Triggering rule-config deployment for testbed {unique_testbed_id}")
        
        # Create a timestamp for the deployment
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"nmt_config_{timestamp}.json"
        
        # Ensure configs directory exists
        configs_dir = os.path.join(os.path.dirname(__file__), 'configs')
        os.makedirs(configs_dir, exist_ok=True)
        
        filepath = os.path.join(configs_dir, filename)
        
        # Save config to file
        with open(filepath, 'w') as f:
            json.dump(config.config_json, f, indent=2)
        
        # Execute rule-config deployment
        run_backend_script = "/mnt/data/nmt_backend/main.py"
        
        # The environment status file that jita_main.py generates
        env_status_file = f'/mnt/data/nmt_backend/jita_env_status_{unique_testbed_id}.json'
        
        logging.info(f"Executing rule-config deployment: python3 {run_backend_script} {filepath} {env_status_file} {unique_rule_id}")
        
        # Execute main.py for rule-config deployment with the environment status file and unique_rule_id
        result = subprocess.run(
            ['python3', run_backend_script, filepath, env_status_file, unique_rule_id],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            universal_newlines=True,
            cwd='/mnt/data/nmt_backend'
        )
        
        # Log the output
        logging.info("Rule-config deployment STDOUT:\n%s", result.stdout)
        if result.stderr:
            logging.error("Rule-config deployment STDERR:\n%s", result.stderr)
        
        # Mark as triggered
        with open(rule_config_triggered_file, 'w') as f:
            json.dump({
                'triggered_at': datetime.now().isoformat(),
                'jita_task_id': jita_status.get('task_id'),
                'jita_status': jita_status.get('status'),
                'deployment_result': result.returncode == 0
            }, f, indent=2)
        
        if result.returncode == 0:
            logging.info(f"Rule-config deployment completed successfully for testbed {unique_testbed_id}")
        else:
            logging.error(f"Rule-config deployment failed for testbed {unique_testbed_id}")
            
    except Exception as e:
        logging.error(f"Error triggering rule-config deployment: {str(e)}")


def monitor_jita_jobs_background():
    """Background monitoring function that checks all active JITA jobs and triggers rule-config deployment"""
    try:
        # Check if running in production environment
        nmt_backend_dir = '/mnt/data/nmt_backend'
        
        if not os.path.exists(nmt_backend_dir):
            # Development/localhost mode - skip JITA monitoring
            logging.debug("⏭️  Skipping JITA monitoring (development mode - no /mnt/data/nmt_backend)")
            return
        
        logging.info("Starting background JITA job monitoring...")
        
        # Get all unique testbed IDs that have JITA status files
        jita_status_files = []
        
        if os.path.exists(nmt_backend_dir):
            for filename in os.listdir(nmt_backend_dir):
                if filename.startswith('jita_env_status_') and filename.endswith('.json'):
                    unique_testbed_id = filename.replace('jita_env_status_', '').replace('.json', '')
                    jita_status_files.append(unique_testbed_id)
        
        logging.info(f"Found {len(jita_status_files)} JITA status files to monitor")
        
        for unique_testbed_id in jita_status_files:
            try:
                # Check JITA job status
                jita_status_file = f'/mnt/data/nmt_backend/jita_env_status_{unique_testbed_id}.json'
                
                if os.path.exists(jita_status_file):
                    with open(jita_status_file, 'r') as f:
                        jita_data = json.load(f)
                    
                    jita_status = jita_data.get('status', 'unknown')
                    task_id = jita_data.get('task_id')
                    
                    logging.info(f"Monitoring testbed {unique_testbed_id}: JITA status = {jita_status}, task_id = {task_id}")
                    
                    # Check if JITA job has succeeded and trigger rule-config deployment
                    if jita_status.lower() in ['completed', 'succeeded']:
                        logging.info(f"JITA job completed for testbed {unique_testbed_id}, triggering rule-config deployment")
                        # Generate a unique_rule_id for background monitoring
                        unique_rule_id = str(uuid.uuid4())
                        _trigger_rule_config_deployment(unique_testbed_id, jita_data, unique_rule_id)
                    elif jita_status.lower() in ['failed', 'error']:
                        logging.warning(f"JITA job failed for testbed {unique_testbed_id}")
                    else:
                        logging.info(f"JITA job still running for testbed {unique_testbed_id}: {jita_status}")
                        
            except Exception as e:
                logging.error(f"Error monitoring testbed {unique_testbed_id}: {str(e)}")
                
    except Exception as e:
        logging.error(f"Error in background JITA job monitoring: {str(e)}")


def collect_continuous_testbed_metrics():
    """Background job to continuously collect metrics for all testbeds"""
    try:
        from database import save_metrics_history
        from services.metrics_collector import MetricsCollector
        
        logging.info("📊 Starting continuous metrics collection...")
        
        # Get all testbeds
        testbeds = []
        session = SessionLocal()
        try:
            from sqlalchemy import text
            query = text("SELECT unique_testbed_id, testbed_label, ncm_ip, testbed_json FROM testbeds")
            result = session.execute(query)
            testbeds = [dict(row._mapping) for row in result.fetchall()]
        except Exception as e:
            logging.error(f"Error fetching testbeds for metrics: {e}")
            return
        finally:
            session.close()
        
        if not testbeds:
            logging.debug("No testbeds found for metrics collection")
            return
        
        logging.info(f"📊 Collecting metrics for {len(testbeds)} testbeds")
        
        # Collect metrics for each testbed
        for testbed in testbeds:
            try:
                testbed_id = testbed['unique_testbed_id']
                testbed_label = testbed.get('testbed_label', 'Unknown')
                ncm_ip = testbed.get('ncm_ip')
                
                # Get Prometheus URL from testbed_json
                testbed_json = testbed.get('testbed_json', {})
                if isinstance(testbed_json, str):
                    import json
                    testbed_json = json.loads(testbed_json)
                
                prometheus_endpoint = testbed_json.get('prometheus_endpoint')
                
                # If prometheus_endpoint not stored, construct it from ncm_ip
                if not prometheus_endpoint and ncm_ip:
                    prometheus_port = testbed_json.get('prometheus_port', 31943)
                    prometheus_endpoint = f"https://{ncm_ip}:{prometheus_port}"
                    logging.debug(f"Constructed Prometheus endpoint for {testbed_label}: {prometheus_endpoint}")
                
                if not prometheus_endpoint:
                    logging.debug(f"No Prometheus endpoint for testbed {testbed_label}, skipping")
                    continue
                
                # Initialize metrics collector
                collector = MetricsCollector(prometheus_endpoint, ncm_ip)
                
                # Get instant snapshot
                snapshot = collector.get_instant_metrics_snapshot()
                
                # Get pod metrics
                pod_metrics_data = collector.collect_pod_metrics()
                
                # Extract scalar values
                cpu_percent = snapshot.get('cpu', {}).get('percent')
                memory_percent = snapshot.get('memory', {}).get('percent')
                disk_percent = snapshot.get('disk', {}).get('percent')
                network_rx = snapshot.get('network', {}).get('rx_mbps')
                network_tx = snapshot.get('network', {}).get('tx_mbps')
                
                # Save to metrics_history
                save_metrics_history(
                    testbed_id=testbed_id,
                    cpu_percent=cpu_percent,
                    memory_percent=memory_percent,
                    disk_percent=disk_percent,
                    network_rx_mbps=network_rx,
                    network_tx_mbps=network_tx,
                    pod_metrics=pod_metrics_data.get('pods'),
                    active_alerts=0,  # Will be populated when we query alerts
                    alert_details=None,
                    full_metrics=snapshot
                )
                
                logging.debug(f"✅ Metrics collected for {testbed_label}: CPU={cpu_percent:.1f}%, Memory={memory_percent:.1f}%")
                
            except Exception as e:
                logging.warning(f"Error collecting metrics for testbed {testbed.get('testbed_label')}: {e}")
                continue
        
        logging.info(f"📊 Metrics collection cycle complete")
        
    except Exception as e:
        logging.error(f"Error in continuous metrics collection: {e}")


def monitor_prometheus_jobs_background():
    """Background monitoring function that checks prometheus status and triggers dynamic workload"""
    try:
        # Check if running in production environment
        nmt_backend_dir = '/mnt/data/nmt_backend'
        
        if not os.path.exists(nmt_backend_dir):
            # Development/localhost mode - skip Prometheus monitoring
            logging.debug("⏭️  Skipping Prometheus monitoring (development mode - no /mnt/data/nmt_backend)")
            return
        
        logging.info("Starting background prometheus job monitoring...")
        
        # Get all prometheus status files
        prometheus_files = []
        dynamic_workload_triggered_dir = '/mnt/data/nmt_backend/dynamic_workload_triggered'
        
        # Ensure directories exist (only in production)
        try:
            os.makedirs(dynamic_workload_triggered_dir, exist_ok=True)
        except OSError as e:
            logging.warning(f"⚠️  Could not create directory {dynamic_workload_triggered_dir}: {e}")
            return
        
        if os.path.exists(nmt_backend_dir):
            for filename in os.listdir(nmt_backend_dir):
                if filename.startswith('prometheus_') and filename.endswith('_status.json'):
                    # Extract unique_rule_id from filename
                    unique_rule_id = filename.replace('prometheus_', '').replace('_status.json', '')
                    prometheus_files.append(unique_rule_id)
        
        logging.info(f"Found {len(prometheus_files)} prometheus status files to monitor")
        
        for unique_rule_id in prometheus_files:
            try:
                # Check if dynamic workload already triggered
                triggered_file = os.path.join(dynamic_workload_triggered_dir, f'dynamic_workload_triggered_{unique_rule_id}.json')
                if os.path.exists(triggered_file):
                    logging.debug(f"Dynamic workload already triggered for rule ID {unique_rule_id}")
                    continue
                
                # Read prometheus status file
                prometheus_status_file = os.path.join(nmt_backend_dir, f'prometheus_{unique_rule_id}_status.json')
                
                if os.path.exists(prometheus_status_file):
                    with open(prometheus_status_file, 'r') as f:
                        prometheus_data = json.load(f)
                    
                    prometheus_status = prometheus_data.get('status', 'unknown')
                    container_name = prometheus_data.get('container_name', 'unknown')
                    
                    logging.info(f"Monitoring rule ID {unique_rule_id}: Prometheus status = {prometheus_status}, container = {container_name}")
                    
                    # Check if prometheus is successful
                    if prometheus_status.lower() == 'success':
                        logging.info(f"Prometheus successful for rule ID {unique_rule_id}, triggering dynamic workload")
                        _trigger_dynamic_workload_deployment(unique_rule_id, prometheus_data)
                    elif prometheus_status.lower() == 'failure':
                        logging.warning(f"Prometheus failed for rule ID {unique_rule_id}")
                        # Mark as triggered to avoid retries
                        _mark_dynamic_workload_triggered(unique_rule_id, prometheus_data, False)
                    else:
                        logging.debug(f"Prometheus still running for rule ID {unique_rule_id}: {prometheus_status}")
                        
            except Exception as e:
                logging.error(f"Error monitoring prometheus rule ID {unique_rule_id}: {str(e)}")
                
    except Exception as e:
        logging.error(f"Error in background prometheus job monitoring: {str(e)}")


def _trigger_dynamic_workload_deployment(unique_rule_id, prometheus_data):
    """Trigger dynamic workload deployment when prometheus succeeds"""
    try:
        # Check if already triggered
        dynamic_workload_triggered_dir = '/mnt/data/nmt_backend/dynamic_workload_triggered'
        triggered_file = os.path.join(dynamic_workload_triggered_dir, f'dynamic_workload_triggered_{unique_rule_id}.json')
        
        if os.path.exists(triggered_file):
            logging.info(f"Dynamic workload already triggered for rule ID {unique_rule_id}")
            return
        
        # Get testbed ID from database using unique_rule_id
        testbed_id = _get_testbed_id_from_rule_id(unique_rule_id)
        
        if not testbed_id:
            logging.warning(f"Could not find testbed ID for rule ID {unique_rule_id}")
            return
        
        logging.info(f"Triggering dynamic workload for rule ID {unique_rule_id}, testbed {testbed_id}")
        
        # Generate unique workload ID
        unique_workload_id = str(uuid.uuid4())
        
        # Execute dynamic workload generation
        dynamic_script = "/mnt/data/nmt_backend/dynamic_workload_generation.py"
        
        cmd = [
            'python3', dynamic_script,
            '--job_profile_id', '68d3c4afd24d822c3c6952c1',  # Dynamic workload profile ID
            '--profile_name', 'dynamic_workload',
            '--username', 'nutanix',
            '--password', 'nutanix/4u',
            '--unique_workload_id', unique_workload_id,
            '--testbed_id', testbed_id
        ]
        
        logging.info(f"Executing dynamic workload: {' '.join(cmd)}")
        
        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            universal_newlines=True,
            cwd='/mnt/data/nmt_backend'
        )
        
        # Log the output
        if result.stdout:
            logging.info(f"Dynamic workload STDOUT:\n{result.stdout}")
        if result.stderr:
            logging.error(f"Dynamic workload STDERR:\n{result.stderr}")
        
        success = result.returncode == 0
        
        # Mark as triggered
        _mark_dynamic_workload_triggered(unique_rule_id, prometheus_data, success)
        
        if success:
            logging.info(f"Dynamic workload triggered successfully for rule ID {unique_rule_id}")
        else:
            logging.error(f"Dynamic workload failed for rule ID {unique_rule_id}")
            
    except Exception as e:
        logging.error(f"Error triggering dynamic workload for rule ID {unique_rule_id}: {e}")


def _get_testbed_id_from_rule_id(unique_rule_id):
    """Get testbed ID associated with this rule ID"""
    try:
        # Create a new database session
        db_session = SessionLocal()
        try:
            # Query the database for config with this unique_rule_id
            config = db_session.query(Config).filter(
                Config.unique_rule_id == unique_rule_id
            ).first()
            
            if config:
                return config.unique_testbed_id
            
            # Fallback: look in config files
            configs_dir = os.path.join(os.path.dirname(__file__), 'configs')
            if os.path.exists(configs_dir):
                for config_file in os.listdir(configs_dir):
                    if config_file.startswith('nmt_config_') and config_file.endswith('.json'):
                        try:
                            config_path = os.path.join(configs_dir, config_file)
                            with open(config_path, 'r') as f:
                                config_data = json.load(f)
                            
                            if config_data.get('unique_rule_id') == unique_rule_id:
                                return config_data.get('unique_testbed_id')
                        except Exception:
                            continue
            
            return None
            
        finally:
            db_session.close()
            
    except Exception as e:
        logging.error(f"Error getting testbed ID for rule ID {unique_rule_id}: {e}")
        return None


def _mark_dynamic_workload_triggered(unique_rule_id, prometheus_data, success):
    """Mark that dynamic workload has been triggered for this rule ID"""
    try:
        dynamic_workload_triggered_dir = '/mnt/data/nmt_backend/dynamic_workload_triggered'
        os.makedirs(dynamic_workload_triggered_dir, exist_ok=True)
        
        triggered_file = os.path.join(dynamic_workload_triggered_dir, f'dynamic_workload_triggered_{unique_rule_id}.json')
        
        trigger_data = {
            'triggered_at': datetime.now().isoformat(),
            'unique_rule_id': unique_rule_id,
            'prometheus_status': prometheus_data.get('status'),
            'prometheus_container': prometheus_data.get('container_name'),
            'prometheus_host_port': prometheus_data.get('host_port'),
            'dynamic_workload_result': success,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        with open(triggered_file, 'w') as f:
            json.dump(trigger_data, f, indent=2)
        
        logging.info(f"Marked dynamic workload as triggered for rule ID {unique_rule_id}")
        
    except Exception as e:
        logging.error(f"Error marking dynamic workload as triggered for rule ID {unique_rule_id}: {e}")


@app.route('/api/trigger-background-monitoring', methods=['POST'])
def trigger_background_monitoring():
    """Manually trigger background monitoring (for testing)"""
    try:
        monitor_jita_jobs_background()
        monitor_prometheus_jobs_background()
        return jsonify({
            'success': True,
            'message': 'Background monitoring completed'
        })
    except Exception as e:
        logging.error(f"Error in manual background monitoring: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Failed to run background monitoring: {str(e)}'
        }), 500


@app.route('/api/dynamic-workload-status/<unique_testbed_id>', methods=['GET'])
def get_dynamic_workload_status(unique_testbed_id):
    """Get dynamic workload status for a specific testbed (latest entry only)"""
    try:
        # Look for dynamic workload status files
        nmt_backend_dir = '/mnt/data/nmt_backend'
        dynamic_workload_status = []
        
        # Check for jita_dynamicload_status files
        for filename in os.listdir(nmt_backend_dir):
            if filename.startswith('jita_dynamicload_status_') and filename.endswith('.json'):
                try:
                    status_file = os.path.join(nmt_backend_dir, filename)
                    with open(status_file, 'r') as f:
                        status_data = json.load(f)
                    
                    # Only include if it matches the testbed ID or contains testbed info
                    testbed_match = (
                        status_data.get('unique_testbed_id') == unique_testbed_id or
                        unique_testbed_id in filename
                    )
                    
                    if testbed_match:
                        # Extract workload ID from filename
                        workload_id = filename.replace('jita_dynamicload_status_', '').replace('.json', '')
                        
                        dynamic_workload_status.append({
                            'workload_id': workload_id,
                            'task_id': status_data.get('task_id'),
                            'status': status_data.get('status'),
                            'job_profile_id': status_data.get('job_profile_id'),
                            'timestamp': status_data.get('timestamp'),
                            'file_timestamp': os.path.getmtime(status_file)  # For sorting
                        })
                except Exception as e:
                    logging.error(f"Error reading dynamic workload status file {filename}: {e}")
                    continue
        
        # Also check for dynamic_workload_triggered files
        dynamic_workload_triggered_dir = '/mnt/data/nmt_backend/dynamic_workload_triggered'
        if os.path.exists(dynamic_workload_triggered_dir):
            for filename in os.listdir(dynamic_workload_triggered_dir):
                if filename.startswith('dynamic_workload_triggered_') and filename.endswith('.json'):
                    try:
                        triggered_file = os.path.join(dynamic_workload_triggered_dir, filename)
                        with open(triggered_file, 'r') as f:
                            triggered_data = json.load(f)
                        
                        # Only include if it matches the testbed ID
                        testbed_match = (
                            triggered_data.get('unique_testbed_id') == unique_testbed_id or
                            unique_testbed_id in filename
                        )
                        
                        if testbed_match:
                            rule_id = filename.replace('dynamic_workload_triggered_', '').replace('.json', '')
                            
                            dynamic_workload_status.append({
                                'rule_id': rule_id,
                                'triggered_at': triggered_data.get('triggered_at'),
                                'prometheus_status': triggered_data.get('prometheus_status'),
                                'dynamic_workload_result': triggered_data.get('dynamic_workload_result'),
                                'timestamp': triggered_data.get('timestamp'),
                                'file_timestamp': os.path.getmtime(triggered_file)  # For sorting
                            })
                    except Exception as e:
                        logging.error(f"Error reading dynamic workload triggered file {filename}: {e}")
                        continue
        
        # Sort by file timestamp (most recent first) and return only the latest entry
        if dynamic_workload_status:
            dynamic_workload_status.sort(key=lambda x: x.get('file_timestamp', 0), reverse=True)
            # Remove the file_timestamp field before returning
            for item in dynamic_workload_status:
                item.pop('file_timestamp', None)
            # Return only the most recent entry
            dynamic_workload_status = dynamic_workload_status[:1]
        
        return jsonify({
            'success': True,
            'dynamic_workload_status': dynamic_workload_status
        })
        
    except Exception as e:
        logging.error(f"Error getting dynamic workload status for testbed {unique_testbed_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Failed to get dynamic workload status: {str(e)}'
        }), 500


@app.route('/api/testbed-pc-ips/<unique_testbed_id>', methods=['GET'])
def get_testbed_pc_ips(unique_testbed_id):
    """Get PC IPs for a testbed from database"""
    try:
        db_session = SessionLocal()
        try:
            # Query testbed table for PC IPs
            testbed = db_session.query(Testbed).filter(
                Testbed.unique_testbed_id == unique_testbed_id
            ).first()
            
            if not testbed:
                return jsonify({
                    'success': False,
                    'error': 'Testbed not found'
                }), 404
            
            # Get PC IPs from testbed record
            pc_ips = []
            if hasattr(testbed, 'pc_ip') and testbed.pc_ip:
                pc_ips.append(testbed.pc_ip)
            
            # Also check if there are multiple PC IPs in deployment details
            if hasattr(testbed, 'deployment_details') and testbed.deployment_details:
                try:
                    deployment_data = json.loads(testbed.deployment_details) if isinstance(testbed.deployment_details, str) else testbed.deployment_details
                    
                    # Look for PC IPs in various fields
                    if 'pc_ip' in deployment_data:
                        if deployment_data['pc_ip'] not in pc_ips:
                            pc_ips.append(deployment_data['pc_ip'])
                    
                    if 'pc_ips' in deployment_data:
                        for ip in deployment_data['pc_ips']:
                            if ip not in pc_ips:
                                pc_ips.append(ip)
                                
                except Exception as e:
                    logging.warning(f"Error parsing deployment details for testbed {unique_testbed_id}: {e}")
            
            return jsonify({
                'success': True,
                'testbed_id': unique_testbed_id,
                'pc_ips': pc_ips,
                'deployment_status': getattr(testbed, 'deployment_status', 'unknown')
            })
            
        finally:
            db_session.close()
            
    except Exception as e:
        logging.error(f"Error getting PC IPs for testbed {unique_testbed_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Failed to get PC IPs: {str(e)}'
        }), 500


@app.route('/api/prometheus-port', methods=['GET'])    
def get_prometheus_port():
    try:
        prometheus_port_file = os.path.join(os.path.dirname(__file__), 'prometheus_port.json')
        with open(prometheus_port_file, 'r') as f:
            data = json.load(f)
        return jsonify({'host_ip': data['host_ip'], 'host_port': data['host_port']})
    except Exception:
        # Fallback/default values if file not found
        return jsonify({'host_ip': '127.0.0.1', 'host_port': 9090})





# Email schedule routes moved to email_routes.py blueprint
# This ensures proper scheduler integration and database persistence

@app.route('/api/health', methods=['GET'])
def health_check():
    """
    Simple health check endpoint for auto-discovery
    """
    return jsonify({
        'status': 'healthy',
        'service': 'nmt-backend',
        'timestamp': datetime.now().isoformat()
    })


@app.route('/api/db-pool-status', methods=['GET'])
def db_pool_status():
    """
    Get database connection pool status for monitoring.
    
    Returns pool statistics including:
    - Total connections in pool
    - Connections currently in use
    - Connections available
    - Overflow connections
    - Pool health status
    
    Useful for:
    - Monitoring connection pool utilization
    - Debugging connection exhaustion issues
    - Capacity planning
    """
    try:
        pool_info = get_pool_status()
        return jsonify({
            'success': True,
            'pool': pool_info,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        logging.error(f"Error getting pool status: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# --- Workload Upload Endpoint ---
@app.route('/api/upload-workload', methods=['POST'])
def upload_workload():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        unique_testbed_id = data.get("unique_testbed_id")
        unique_rule_id = data.get("unique_rule_id")
        logging.info(f"Printing rule-id {unique_rule_id}")
        
        unique_workload_id = str(uuid.uuid4())

        # Ensure submitted-workloads directory exists
        workloads_dir = os.path.join(os.path.dirname(__file__), '..', 'submitted-workloads')
        os.makedirs(workloads_dir, exist_ok=True)
        
        # Use workload_label or timestamp for filename
        label = data.get('workload_label', 'workload')
        timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
        filename = f"{label.replace(' ', '_')}_{timestamp}.json"
        filepath = os.path.join(workloads_dir, filename)
        
        # Save to file
        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2)
        logging.info(f"Saved workload JSON to {filepath}")
        
        # Save to database
        try:
            # Get PC UUID, PC IP, and testbed label from frontend request
            pc_uuid = data.get('pc_uuid')  # PC UUID from onboarding context
            pc_ip = data.get('pc_ip')      # PC IP from onboarding context
            testbed_label = data.get('testbed_label')  # Testbed name from onboarding context

            
            # Save to database using the PC UUID and testbed label
            save_workload_to_db(g.db, unique_workload_id, unique_rule_id,unique_testbed_id, pc_ip, pc_uuid, label, data, testbed_label)
            logging.info(f"Saved workload to database with PC UUID: {pc_uuid} and testbed label: {testbed_label}")
            
        except Exception as db_error:
            logging.error(f"Failed to save workload to database: {str(db_error)}")
            # Don't fail the entire request if DB save fails - file save succeeded
        
        return jsonify({'success': True, 'unique_workload_id': unique_workload_id, 'filename': filename, 'filepath': filepath})
    except Exception as e:
        logging.error(f"Failed to upload workload: {str(e)}")
        return jsonify({'error': f'Failed to upload workload: {str(e)}'}), 500

# --- Testbed Upload Endpoint ---
@app.route('/api/upload-testbed', methods=['POST'])
def upload_testbed():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        unique_testbed_id = str(uuid.uuid4())
        
        # Ensure submitted-testbeds directory exists
        testbeds_dir = os.path.join(os.path.dirname(__file__), '..', 'submitted-testbeds')
        os.makedirs(testbeds_dir, exist_ok=True)
        
        # Extract testbed information from request
        # Support both old format (Testbed Name) and new format (testbed_label)
        testbed_name = data.get('testbed_label') or data.get('Testbed Name', 'testbed')
        pc_ip = data.get('pc_ip')
        ncm_ip = data.get('ncm_ip')
        testbed_uuid = data.get('uuid')
        username = data.get('username')
        password = data.get('password')
        testbed_json = data.get('testbed_json', data)  # Use testbed_json if provided, else use entire data
        
        logging.info(f"📥 Uploading testbed: {testbed_name}")
        logging.info(f"   PC IP: {pc_ip}")
        logging.info(f"   NCM IP: {ncm_ip}")
        logging.info(f"   UUID: {testbed_uuid}")
        
        # Use testbed name or timestamp for filename
        timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
        filename = f"{unique_testbed_id}_{testbed_name.replace(' ', '_')}_{timestamp}.json"
        filepath = os.path.join(testbeds_dir, filename)
        
        # Save to file
        with open(filepath, 'w') as f:
            json.dump(testbed_json, f, indent=2)
        logging.info(f"✅ Saved testbed JSON to {filepath}")
        
        # Save to database with all available information
        try:
            save_testbed_to_db(
                g.db,
                unique_testbed_id, 
                pc_ip, 
                testbed_uuid, 
                testbed_name, 
                testbed_json, 
                filepath,
                ncm_ip=ncm_ip,
                username=username,
                password=password
            )
            logging.info(f"✅ Saved testbed to database:")
            logging.info(f"   Testbed ID: {unique_testbed_id}")
            logging.info(f"   Testbed Label: {testbed_name}")
            logging.info(f"   NCM IP: {ncm_ip}")
            
        except Exception as db_error:
            logging.error(f"❌ Failed to save testbed to database: {str(db_error)}")
            logging.error(f"   Error details: {traceback.format_exc()}")
            # Don't fail the entire request if DB save fails - file save succeeded

        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': filepath,
            'unique_testbed_id': unique_testbed_id,
            'message': f'Testbed "{testbed_name}" saved successfully with NCM IP {ncm_ip}'
        })

    except Exception as e:
        logging.error(f"❌ Failed to upload testbed: {str(e)}")
        logging.error(f"   Error details: {traceback.format_exc()}")
        return jsonify({'error': str(e), 'success': False}), 500

# --- Workload Retrieval Endpoints ---
@app.route('/api/get-workloads', methods=['GET'])
def get_workloads():
    """Get all workloads, optionally filtered by PC IP"""
    try:
        pc_ip = request.args.get('pc_ip')
        if pc_ip:
            workloads = fetch_workloads_by_pc_ip(g.db, pc_ip)
        else:
            # If no pc_ip provided, get all workloads
            from models.workload import Workload
            workloads = g.db.query(Workload).order_by(Workload.timestamp.desc()).all()
        
        workload_list = [workload.to_dict() for workload in workloads]
        return jsonify({'success': True, 'workloads': workload_list})
    except Exception as e:
        logging.error(f"Failed to fetch workloads: {str(e)}")
        return jsonify({'error': f'Failed to fetch workloads: {str(e)}'}), 500

@app.route('/api/get-workload', methods=['GET'])
def get_latest_workload():
    """Get the latest workload (for compatibility with frontend)"""
    try:
        from models.workload import Workload
        latest_workload = g.db.query(Workload).order_by(Workload.timestamp.desc()).first()
        
        if latest_workload:
            return jsonify(latest_workload.workload_json)
        else:
            return jsonify({'error': 'No workloads found'}), 404
    except Exception as e:
        logging.error(f"Failed to fetch latest workload: {str(e)}")
        return jsonify({'error': f'Failed to fetch latest workload: {str(e)}'}), 500

@app.route('/api/get-workload/<uuid>', methods=['GET'])
def get_workload_by_uuid(uuid):
    """Get a specific workload by UUID"""
    try:
        workload = fetch_workload_by_uuid(g.db, uuid)
        if workload:
            return jsonify({'success': True, 'workload': workload.to_dict()})
        else:
            return jsonify({'error': 'Workload not found'}), 404
    except Exception as e:
        logging.error(f"Failed to fetch workload: {str(e)}")
        return jsonify({'error': f'Failed to fetch workload: {str(e)}'}), 500

# --- Testbed Retrieval Endpoints ---
@app.route('/api/get-testbeds', methods=['GET'])
def get_testbeds():
    """Get all testbeds, optionally filtered by PC IP"""
    try:
        pc_ip = request.args.get('pc_ip')
        if pc_ip:
            testbeds = fetch_testbeds_by_pc_ip(g.db, pc_ip)
        else:
            # If no pc_ip provided, get all testbeds
            from models.testbed import Testbed
            testbeds = g.db.query(Testbed).order_by(Testbed.timestamp.desc()).all()
        
        testbed_list = [testbed.to_dict() for testbed in testbeds]
        return jsonify({'success': True, 'testbeds': testbed_list})
    except Exception as e:
        logging.error(f"Failed to fetch testbeds: {str(e)}")
        return jsonify({'error': f'Failed to fetch testbeds: {str(e)}'}), 500

@app.route('/api/get-testbed/<uuid>', methods=['GET'])
def get_testbed_by_uuid(uuid):
    """Get a specific testbed by UUID"""
    try:
        testbed = fetch_testbed_by_uuid(g.db, uuid)
        if testbed:
            return jsonify({'success': True, 'testbed': testbed.to_dict()})
        else:
            return jsonify({'error': 'Testbed not found'}), 404
    except Exception as e:
        logging.error(f"Failed to fetch testbed: {str(e)}")
        return jsonify({'error': f'Failed to fetch testbed: {str(e)}'}), 500

@app.route('/api/get-testbed-labels', methods=['GET'])
def get_testbed_labels():
    """Get all unique testbed labels from the database"""
    try:
        labels = fetch_all_testbed_labels(g.db)
        # Format for dropdown - include testbed label and pc_ip from latest testbed with that label
        testbed_options = []
        for label in labels:
            latest_testbed = fetch_latest_testbed_by_label(g.db, label)
            if latest_testbed:
                testbed_options.append({
                    'value': label,
                    'label': label,
                    'pcIp': latest_testbed.pc_ip
                })
        
        return jsonify({'success': True, 'testbed_labels': testbed_options})
    except Exception as e:
        logging.error(f"Failed to fetch testbed labels: {str(e)}")
        return jsonify({'error': f'Failed to fetch testbed labels: {str(e)}'}), 500

@app.route('/api/get-testbed-by-label/<testbed_label>', methods=['GET'])
def get_testbed_by_label(testbed_label):
    """Get the latest testbed data for a specific label"""
    try:
        testbed = fetch_latest_testbed_by_label(g.db, testbed_label)
        if testbed:
            return jsonify({'success': True, 'testbed': testbed.to_dict()})
        else:
            return jsonify({'error': 'Testbed not found'}), 404
    except Exception as e:
        logging.error(f"Failed to fetch testbed by label: {str(e)}")
        return jsonify({'error': f'Failed to fetch testbed by label: {str(e)}'}), 500


@app.route('/api/delete-testbed/<testbed_id>', methods=['DELETE'])
def delete_testbed(testbed_id):
    """Delete a testbed by unique_testbed_id and all associated data"""
    try:
        from sqlalchemy import text
        session = SessionLocal()
        
        # Find testbed by unique_testbed_id
        testbed = session.query(Testbed).filter(
            Testbed.unique_testbed_id == testbed_id
        ).first()
        
        if not testbed:
            session.close()
            return jsonify({'success': False, 'error': 'Testbed not found'}), 404
        
        testbed_label = testbed.testbed_label
        
        try:
            # Step 1: Delete workloads associated with this testbed (they reference configs)
            delete_workloads_sql = text("DELETE FROM workloads WHERE unique_testbed_id = :testbed_id")
            workload_result = session.execute(delete_workloads_sql, {"testbed_id": testbed_id})
            workloads_deleted = workload_result.rowcount
            logging.info(f"Deleted {workloads_deleted} workload(s) for testbed {testbed_id}")
            
            # Step 2: Delete associated rules/configs
            configs = session.query(Config).filter(
                Config.unique_testbed_id == testbed_id
            ).all()
            configs_deleted = len(configs)
            for config in configs:
                session.delete(config)
                logging.info(f"Deleted config {config.unique_rule_id} for testbed {testbed_id}")
            
            # Step 3: Delete associated executions
            try:
                delete_executions_sql = text("DELETE FROM executions WHERE testbed_id = :testbed_id")
                exec_result = session.execute(delete_executions_sql, {"testbed_id": testbed_id})
                executions_deleted = exec_result.rowcount
                logging.info(f"Deleted {executions_deleted} execution(s) for testbed {testbed_id}")
            except Exception as exec_err:
                # Table might not exist yet, that's okay
                logging.warning(f"Could not delete executions (table might not exist): {exec_err}")
                executions_deleted = 0
            
            # Step 4: Delete operation metrics
            try:
                delete_op_metrics_sql = text("DELETE FROM operation_metrics WHERE testbed_id = :testbed_id")
                op_result = session.execute(delete_op_metrics_sql, {"testbed_id": testbed_id})
                op_metrics_deleted = op_result.rowcount
                logging.info(f"Deleted {op_metrics_deleted} operation metric(s) for testbed {testbed_id}")
            except Exception as op_err:
                logging.warning(f"Could not delete operation metrics: {op_err}")
                op_metrics_deleted = 0
            
            # Step 5: Delete metrics history
            try:
                delete_metrics_history_sql = text("DELETE FROM metrics_history WHERE testbed_id = :testbed_id")
                metrics_result = session.execute(delete_metrics_history_sql, {"testbed_id": testbed_id})
                metrics_deleted = metrics_result.rowcount
                logging.info(f"Deleted {metrics_deleted} metrics history record(s) for testbed {testbed_id}")
            except Exception as metrics_err:
                logging.warning(f"Could not delete metrics history: {metrics_err}")
                metrics_deleted = 0
            
            # Step 6: Delete the testbed
            session.delete(testbed)
            session.commit()
            
            logging.info(f"Testbed {testbed_id} ({testbed_label}) deleted successfully")
            
            return jsonify({
                'success': True,
                'message': f'Testbed "{testbed_label}" and all associated data deleted successfully',
                'deleted_counts': {
                    'workloads': workloads_deleted,
                    'configs': configs_deleted,
                    'executions': executions_deleted,
                    'operation_metrics': op_metrics_deleted,
                    'metrics_history': metrics_deleted
                }
            })
            
        except Exception as e:
            session.rollback()
            raise e
        finally:
            session.close()
        
    except Exception as e:
        logging.error(f"Error deleting testbed {testbed_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/get-workload-labels', methods=['GET'])
def get_workload_labels():
    """Get all unique workload labels from the database"""
    try:
        labels = fetch_all_workload_labels(g.db)
        # Format for dropdown - include workload label and pc_ip from latest workload with that label
        workload_options = []
        for label in labels:
            latest_workload = fetch_latest_workload_by_label(g.db, label)
            if latest_workload:
                workload_options.append({
                    'value': label,
                    'label': label,
                    'pcIp': latest_workload.pc_ip
                })
        
        return jsonify({'success': True, 'workload_labels': workload_options})
    except Exception as e:
        logging.error(f"Failed to fetch workload labels: {str(e)}")
        return jsonify({'error': f'Failed to fetch workload labels: {str(e)}'}), 500

@app.route('/api/get-workload-by-testbed/<testbed_label>', methods=['GET'])
def get_workload_by_testbed(testbed_label):
    """Get the latest workload data for a specific testbed label"""
    try:
        # 1. Fetch testbed info by label
        testbed = fetch_latest_testbed_by_label(g.db, testbed_label)
        if not testbed:
            return jsonify({'error': 'Testbed not found'}), 404

        unique_testbed_id = str(testbed.unique_testbed_id)

        # 2. Fetch workload by unique_testbed_id
        workload = fetch_latest_workload_by_unique_testbed_id(g.db, unique_testbed_id)
        if workload:
            workload_dict = workload.to_dict()

            # 🔧 Unwrap nested "output" if present
            w_json = workload_dict.get("workload_json", {})
            if isinstance(w_json, dict) and "output" in w_json:
                workload_dict["workload_json"] = w_json["output"]

            return jsonify({'success': True, 'workload': workload_dict})
        else:
            return jsonify({'error': 'No workload found for this testbed'}), 404

    except Exception as e:
        logging.error(f"Failed to fetch workload by testbed label: {str(e)}")
        return jsonify({'error': f'Failed to fetch workload by testbed label: {str(e)}'}), 500

#@app.route('/api/get-workload-by-testbed/<testbed_label>', methods=['GET'])
#def get_workload_by_label(workload_label):
#    """Get the latest workload data for a specific label"""
#    try:
#        # Use the request-scoped DB session
#        workload = fetch_latest_workload_by_label(g.db, workload_label)
#
#        if workload:
#            # Convert the ORM object to dictionary
#            return jsonify({
#                'success': True,
#                'workload': workload.to_dict()  # Ensure your model has a to_dict() method
#            })
#        else:
#            return jsonify({'success': False, 'error': 'Workload not found'}), 404
#
#    except Exception as e:
#        logging.error(f"Failed to fetch workload by label: {str(e)}")
#        return jsonify({'success': False, 'error': f'Failed to fetch workload by label: {str(e)}'}), 500


# --- Rule Config Upload Endpoint ---
@app.route('/api/upload-rule-config', methods=['POST'])
def upload_rule_config():
    try:
        data = request.json
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        unique_testbed_id = data.get("unique_testbed_id")
        pc_ip_from_request = data.get("pc_ip")
        config = data.get("config")
        
        # Extract pc_ip from config if not in request
        pc_ip = pc_ip_from_request or (config.get('Config', {}).get('pc_ip') if config else None)
        
        # Log for debugging
        logging.info(f"upload-rule-config: unique_testbed_id={unique_testbed_id}, pc_ip={pc_ip}")
        
        # Require either unique_testbed_id OR pc_ip
        if not unique_testbed_id and not pc_ip:
            return jsonify({'error': 'Missing unique_testbed_id or pc_ip - at least one is required'}), 400
            
        # Ensure submitted-rules directory exists
        rules_dir = os.path.join(os.path.dirname(__file__), '..', 'submitted-rules')
        os.makedirs(rules_dir, exist_ok=True)
        
        # Use timestamp for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"rule_config_{timestamp}.json"
        filepath = os.path.join(rules_dir, filename)
        
        with open(filepath, 'w') as f:
            json.dump(data, f, indent=2)
            
        return jsonify({'success': True, 'filename': filename})
        
    except Exception as e:
        return jsonify({'error': f'Failed to upload rule config: {str(e)}'}), 500

# ============================================================================
# RULE MANAGEMENT API ENDPOINTS
# ============================================================================

@app.route('/api/get-rules-by-testbed/<testbed_id>', methods=['GET'])
def get_rules_by_testbed(testbed_id):
    """
    Get all rules for a specific testbed.
    
    Response:
    {
        "success": true,
        "rules": [
            {
                "id": "rule-uuid",
                "testbed_id": "testbed-uuid",
                "rule_name": "High CPU Alert",
                "rule_type": "pod",
                "severity": "critical",
                "enabled": true,
                "config": {...}
            }
        ]
    }
    """
    try:
        # Query configs table for rules matching testbed_id
        session = SessionLocal()
        configs = session.query(Config).filter(
            Config.unique_testbed_id == testbed_id
        ).order_by(Config.timestamp.desc()).all()
        
        rules_list = []
        for config in configs:
            rule_data = {
                'id': config.unique_rule_id,
                'testbed_id': config.unique_testbed_id,
                'pc_ip': config.pc_ip,
                'timestamp': config.timestamp.isoformat() if config.timestamp else None,
                'config': config.config_json
            }
            rules_list.append(rule_data)
        
        session.close()
        
        return jsonify({
            'success': True,
            'rules': rules_list,
            'count': len(rules_list)
        })
        
    except Exception as e:
        logging.error(f"Error getting rules for testbed {testbed_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/add-rule', methods=['POST'])
def add_rule():
    """
    Add a new rule for a testbed.
    
    Request body:
    {
        "testbed_id": "testbed-uuid",
        "rule_config": {...}
    }
    """
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        testbed_id = data.get('testbed_id')
        rule_config = data.get('rule_config')
        
        if not testbed_id or not rule_config:
            return jsonify({'success': False, 'error': 'testbed_id and rule_config required'}), 400
        
        # Generate unique rule ID
        import uuid
        rule_id = str(uuid.uuid4())
        
        # Get testbed details for pc_ip
        session = SessionLocal()
        testbed = session.query(Testbed).filter(
            Testbed.unique_testbed_id == testbed_id
        ).first()
        
        pc_ip = testbed.pc_ip if testbed else "0.0.0.0"
        
        # Save to configs table
        save_config_to_db(
            session,
            unique_rule_id=rule_id,
            unique_testbed_id=testbed_id,
            pc_ip=pc_ip,
            config_json=rule_config
        )
        
        session.commit()
        session.close()
        
        logging.info(f"Rule {rule_id} added for testbed {testbed_id}")
        
        return jsonify({
            'success': True,
            'rule_id': rule_id,
            'message': 'Rule added successfully'
        })
        
    except Exception as e:
        logging.error(f"Error adding rule: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/update-rule/<rule_id>', methods=['PUT'])
def update_rule(rule_id):
    """Update an existing rule"""
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        rule_config = data.get('rule_config')
        if not rule_config:
            return jsonify({'success': False, 'error': 'rule_config required'}), 400
        
        session = SessionLocal()
        config = session.query(Config).filter(
            Config.unique_rule_id == rule_id
        ).first()
        
        if not config:
            session.close()
            return jsonify({'success': False, 'error': 'Rule not found'}), 404
        
        config.config_json = rule_config
        config.timestamp = datetime.utcnow()
        
        session.commit()
        session.close()
        
        logging.info(f"Rule {rule_id} updated")
        
        return jsonify({
            'success': True,
            'message': 'Rule updated successfully'
        })
        
    except Exception as e:
        logging.error(f"Error updating rule: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/delete-rule/<rule_id>', methods=['DELETE'])
def delete_rule(rule_id):
    """Delete a rule and all associated workloads"""
    try:
        from sqlalchemy import text
        session = SessionLocal()
        
        # First, check if rule exists
        config = session.query(Config).filter(
            Config.unique_rule_id == rule_id
        ).first()
        
        if not config:
            session.close()
            return jsonify({'success': False, 'error': 'Rule not found'}), 404
        
        try:
            # Delete associated workloads first (to avoid foreign key constraint violation)
            delete_workloads_query = text("DELETE FROM workloads WHERE unique_rule_id = :rule_id")
            workload_result = session.execute(delete_workloads_query, {'rule_id': rule_id})
            workloads_deleted = workload_result.rowcount
            
            if workloads_deleted > 0:
                logging.info(f"Deleted {workloads_deleted} workload(s) associated with rule {rule_id}")
            
            # Now delete the rule
            session.delete(config)
            session.commit()
            
            logging.info(f"Rule {rule_id} deleted successfully")
            
            return jsonify({
                'success': True,
                'message': f'Rule deleted successfully (removed {workloads_deleted} associated workload(s))'
            })
            
        except Exception as e:
            session.rollback()
            raise e
        finally:
            session.close()
        
    except Exception as e:
        logging.error(f"Error deleting rule {rule_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# --- Jita Job Runner Endpoint ---
# ============================================================================
# NEW EXECUTION API ENDPOINTS (Replacing JITA)
# ============================================================================

from adapters.nmt_execution_adapter import get_nmt_execution_adapter

@app.route('/api/start-execution', methods=['POST'])
def start_execution():
    """
    Start a new execution for a testbed.
    
    Request body:
    {
        "unique_testbed_id": "testbed-uuid",
        "workload_config": {
            "entities": [...],
            "duration": 60,
            "parallel": 5,
            "distribution": "LINEAR"
        },
        "total_operations": 100
    }
    
    Response:
    {
        "success": true,
        "execution_id": "NMT-20260127-123456-abc123",
        "message": "Execution started successfully",
        "status": "PENDING"
    }
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        unique_testbed_id = data.get('unique_testbed_id')
        if not unique_testbed_id:
            return jsonify({'success': False, 'error': 'unique_testbed_id is required'}), 400
        
        # DEBUG: Log the incoming request data
        logging.info(f"🔍 DEBUG - Received execution request data: {data}")
        logging.info(f"🔍 DEBUG - Config field: {data.get('config')}")
        logging.info(f"🔍 DEBUG - Total operations field: {data.get('total_operations')}")
        
        logging.info(f"Starting execution for testbed: {unique_testbed_id}")
        
        # Fetch testbed from database to get credentials
        testbed = fetch_testbed_by_unique_id(g.db, unique_testbed_id)
        if not testbed:
            logging.error(f"Testbed not found: {unique_testbed_id}")
            return jsonify({'success': False, 'error': f'Testbed {unique_testbed_id} not found'}), 404
        
        # Enrich execution data with testbed credentials
        data['pc_ip'] = testbed.pc_ip
        data['ncm_ip'] = testbed.ncm_ip
        data['username'] = testbed.username
        data['password'] = testbed.password
        data['testbed_label'] = testbed.testbed_label
        data['testbed_uuid'] = testbed.uuid
        
        logging.info(f"Testbed details - PC IP: {testbed.pc_ip}, NCM IP: {testbed.ncm_ip}, Label: {testbed.testbed_label}")
        
        adapter = get_nmt_execution_adapter()
        result = adapter.start_testbed_execution(data)
        
        if result['success']:
            logging.info(f"Execution started: {result['execution_id']}")
            return jsonify(result), 200
        else:
            logging.error(f"Failed to start execution: {result.get('error')}")
            return jsonify(result), 400
            
    except Exception as e:
        logging.exception("Error starting execution")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/get-executions', methods=['GET'])
def get_executions():
    """
    Get all executions from the database
    
    Response:
    {
        "success": true,
        "executions": [
            {
                "execution_id": "NMT-...",
                "testbed_id": "...",
                "status": "COMPLETED",
                "progress": 100,
                "start_time": "...",
                "end_time": "...",
                "total_operations": 10,
                "completed_operations": 10,
                "successful_operations": 9,
                "failed_operations": 1
            },
            ...
        ]
    }
    """
    try:
        from sqlalchemy import text
        session = SessionLocal()
        
        query = text("""
            SELECT execution_id, testbed_id, status, progress, 
                   total_operations, completed_operations, successful_operations, failed_operations,
                   start_time, end_time, created_at, last_error
            FROM executions
            ORDER BY created_at DESC
        """)
        
        result = session.execute(query)
        executions = []
        
        for row in result.fetchall():
            executions.append({
                'execution_id': row[0],
                'testbed_id': row[1],
                'status': row[2],
                'progress': row[3],
                'total_operations': row[4],
                'completed_operations': row[5],
                'successful_operations': row[6],
                'failed_operations': row[7],
                'start_time': row[8].isoformat() if row[8] else None,
                'end_time': row[9].isoformat() if row[9] else None,
                'created_at': row[10].isoformat() if row[10] else None,
                'last_error': row[11]
            })
        
        session.close()
        
        return jsonify({
            'success': True,
            'executions': executions
        }), 200
            
    except Exception as e:
        logging.exception("Error getting executions")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/execution-report/<execution_id>', methods=['GET'])
def get_execution_report(execution_id):
    """
    Get comprehensive execution report with system metrics.
    
    Response:
    {
        "success": true,
        "report": {
            "execution_id": "NMT-20260127-123456-abc123",
            "testbed": {...},
            "workload": {...},
            "execution": {...},
            "metrics": {
                "cpu": {"average": [...], "maximum": [...], "summary": {...}},
                "memory": {"usage_percentage": [...], "summary": {...}},
                "latency": {...},
                "disk": {...},
                "network": {...},
                "alerts": [...]
            },
            "summary": {
                "success": true,
                "operations_completed": 100,
                "avg_cpu": 45.2,
                "max_cpu": 78.5,
                "avg_memory": 62.3,
                "alerts_triggered": 0
            }
        }
    }
    """
    try:
        adapter = get_nmt_execution_adapter()
        result = adapter.get_execution_report(execution_id)
        
        if result['success']:
            return jsonify(result), 200
        else:
            return jsonify(result), 404
            
    except Exception as e:
        logging.exception("Error getting execution report")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/execution-report-detailed/<execution_id>', methods=['GET'])
def get_execution_report_detailed(execution_id):
    """
    Get detailed execution report with operation metrics for UI display
    """
    try:
        from sqlalchemy import text
        session = SessionLocal()
        
        # Get execution details
        execution_query = text("""
            SELECT execution_id, testbed_id, status, progress, 
                   total_operations, completed_operations, successful_operations, failed_operations,
                   start_time, end_time, created_at, last_error, config, metrics, prometheus_url
            FROM executions
            WHERE execution_id = :execution_id
        """)
        
        result = session.execute(execution_query, {'execution_id': execution_id})
        execution_row = result.fetchone()
        
        if not execution_row:
            return jsonify({'success': False, 'error': 'Execution not found'}), 404
        
        execution_data = {
            'execution_id': execution_row[0],
            'testbed_id': execution_row[1],
            'status': execution_row[2],
            'progress': execution_row[3],
            'total_operations': execution_row[4],
            'completed_operations': execution_row[5],
            'successful_operations': execution_row[6],
            'failed_operations': execution_row[7],
            'start_time': execution_row[8].isoformat() if execution_row[8] else None,
            'end_time': execution_row[9].isoformat() if execution_row[9] else None,
            'created_at': execution_row[10].isoformat() if execution_row[10] else None,
            'last_error': execution_row[11],
            'config': execution_row[12],
            'metrics': execution_row[13] or {},
            'prometheus_url': execution_row[14],
            'duration_minutes': ((execution_row[9] - execution_row[8]).total_seconds() / 60) if execution_row[9] and execution_row[8] else 0
        }
        
        # Get testbed details
        testbed_query = text("""
            SELECT unique_testbed_id, testbed_label, pc_ip, ncm_ip, username, testbed_json
            FROM testbeds
            WHERE unique_testbed_id = :testbed_id
        """)
        
        testbed_result = session.execute(testbed_query, {'testbed_id': execution_data['testbed_id']})
        testbed_row = testbed_result.fetchone()
        
        testbed_data = {
            'unique_testbed_id': testbed_row[0],
            'testbed_label': testbed_row[1],
            'pc_ip': testbed_row[2],
            'ncm_ip': testbed_row[3],
            'username': testbed_row[4],
            'testbed_json': testbed_row[5]
        } if testbed_row else {}
        
        # Get operation metrics
        ops_query = text("""
            SELECT id, execution_id, testbed_id, entity_type, operation_type, entity_name, entity_uuid,
                   started_at, completed_at, duration_seconds, status, error_message, 
                   metrics_snapshot, pod_cpu_percent, pod_memory_mb, pod_network_rx_mbps, pod_network_tx_mbps
            FROM operation_metrics
            WHERE execution_id = :execution_id
            ORDER BY started_at ASC
        """)
        
        ops_result = session.execute(ops_query, {'execution_id': execution_id})
        operation_metrics = []
        
        for row in ops_result.fetchall():
            operation_metrics.append({
                'id': row[0],
                'execution_id': row[1],
                'testbed_id': row[2],
                'entity_type': row[3],
                'operation_type': row[4],
                'entity_name': row[5],
                'entity_uuid': row[6],
                'started_at': row[7].isoformat() if row[7] else None,
                'completed_at': row[8].isoformat() if row[8] else None,
                'duration_seconds': float(row[9]) if row[9] else 0,
                'status': row[10],
                'error_message': row[11],
                'metrics_snapshot': row[12],
                'pod_cpu_percent': float(row[13]) if row[13] else 0,
                'pod_memory_mb': float(row[14]) if row[14] else 0,
                'pod_network_rx_mbps': float(row[15]) if row[15] else 0,
                'pod_network_tx_mbps': float(row[16]) if row[16] else 0
            })
        
        # Get alerts for this testbed during execution time
        alerts_query = text("""
            SELECT id, alert_id, testbed_id, testbed_label, alert_name, alert_type, 
                   severity, status, description, rule_id, triggered_at, resolved_at, 
                   webhook_url, slack_status, metadata
            FROM slack_alerts
            WHERE testbed_id = :testbed_id
            AND triggered_at BETWEEN :start_time AND :end_time
            ORDER BY triggered_at DESC
        """)
        
        alerts_result = session.execute(alerts_query, {
            'testbed_id': execution_data['testbed_id'],
            'start_time': execution_row[8],
            'end_time': execution_row[9] or datetime.utcnow()
        })
        
        alerts = []
        for row in alerts_result.fetchall():
            alerts.append({
                'id': row[0],
                'alert_id': row[1],
                'testbed_id': row[2],
                'testbed_label': row[3],
                'alert_name': row[4],
                'alert_type': row[5],
                'severity': row[6],
                'status': row[7],
                'description': row[8],
                'rule_id': row[9],
                'triggered_at': row[10].isoformat() if row[10] else None,
                'resolved_at': row[11].isoformat() if row[11] else None,
                'webhook_url': row[12],
                'slack_status': row[13],
                'metadata': row[14]
            })
        
        session.close()
        
        return jsonify({
            'success': True,
            'execution': execution_data,
            'testbed': testbed_data,
            'operation_metrics': operation_metrics,
            'alerts': alerts
        }), 200
            
    except Exception as e:
        logging.exception("Error getting detailed execution report")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/execution-report/<execution_id>/export/<format>', methods=['GET'])
def export_execution_report(execution_id, format):
    """
    Export execution report in various formats (for Execution Workload Manager)
    
    Formats: csv, json, excel
    """
    try:
        if format.lower() == 'excel' or format.lower() == 'xlsx':
            # Excel export - query database for pod-level data
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                from database import SessionLocal
                from sqlalchemy import text
                import json as json_lib
                from flask import make_response
                import io
                
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Pod Metrics"
                
                # Headers matching snapshot file format
                headers = [
                    "Execution ID", "Operation #", "Entity Type", "Operation", "Status",
                    "Pod Name", "Namespace", "Node", "CPU Before (%)", "CPU After (%)", 
                    "CPU Delta (%)", "Memory Before (MB)", "Memory After (MB)", 
                    "Memory Delta (MB)", "Network RX Before (Mbps)", "Network RX After (Mbps)",
                    "Network TX Before (Mbps)", "Network TX After (Mbps)", "Timestamp", "Duration (s)"
                ]
                
                # Style headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Query database for pod metrics
                session = SessionLocal()
                row = 2
                op_idx = 0
                
                try:
                    # Get pod_operation_correlation (most complete data)
                    pod_query = text("""
                        SELECT 
                            entity_type, operation_type, pod_name, namespace, node_name,
                            cpu_percent_before, cpu_percent_after, cpu_delta,
                            memory_mb_before, memory_mb_after, memory_delta,
                            network_rx_mbps_before, network_rx_mbps_after, network_rx_delta,
                            network_tx_mbps_before, network_tx_mbps_after, network_tx_delta,
                            measured_at
                        FROM pod_operation_correlation
                        WHERE execution_id = :execution_id
                        ORDER BY measured_at ASC
                    """)
                    
                    pod_result = session.execute(pod_query, {'execution_id': execution_id})
                    pod_correlations = pod_result.fetchall()
                    
                    # Process pod correlations
                    for pod_row in pod_correlations:
                        op_idx += 1
                        ws.cell(row=row, column=1, value=execution_id)
                        ws.cell(row=row, column=2, value=op_idx)
                        ws.cell(row=row, column=3, value=pod_row[0] or '')
                        ws.cell(row=row, column=4, value=pod_row[1] or '')
                        ws.cell(row=row, column=5, value='COMPLETED')
                        ws.cell(row=row, column=6, value=pod_row[2] or '')
                        ws.cell(row=row, column=7, value=pod_row[3] or '')
                        ws.cell(row=row, column=8, value=pod_row[4] or '')
                        ws.cell(row=row, column=9, value=float(pod_row[5]) if pod_row[5] is not None else 0)
                        ws.cell(row=row, column=10, value=float(pod_row[6]) if pod_row[6] is not None else 0)
                        ws.cell(row=row, column=11, value=float(pod_row[7]) if pod_row[7] is not None else 0)
                        ws.cell(row=row, column=12, value=float(pod_row[8]) if pod_row[8] is not None else 0)
                        ws.cell(row=row, column=13, value=float(pod_row[9]) if pod_row[9] is not None else 0)
                        ws.cell(row=row, column=14, value=float(pod_row[10]) if pod_row[10] is not None else 0)
                        ws.cell(row=row, column=15, value=float(pod_row[11]) if pod_row[11] is not None else 0)
                        ws.cell(row=row, column=16, value=float(pod_row[12]) if pod_row[12] is not None else 0)
                        ws.cell(row=row, column=17, value=float(pod_row[13]) if pod_row[13] is not None else 0)
                        ws.cell(row=row, column=18, value=float(pod_row[14]) if pod_row[14] is not None else 0)
                        ws.cell(row=row, column=19, value=pod_row[16].isoformat() if pod_row[16] else '')
                        ws.cell(row=row, column=20, value=0)
                        row += 1
                    
                    # Fallback: If no pod correlations, use operation_metrics
                    if row == 2:
                        query = text("""
                            SELECT 
                                entity_type, operation_type, entity_name,
                                started_at, completed_at, duration_seconds, status,
                                pod_metrics_before, pod_metrics_after
                            FROM operation_metrics
                            WHERE execution_id = :execution_id
                            ORDER BY started_at ASC
                        """)
                        
                        result = session.execute(query, {'execution_id': execution_id})
                        db_metrics = result.fetchall()
                        
                        for db_metric in db_metrics:
                            op_idx += 1
                            pods_before = {}
                            pods_after = {}
                            
                            if db_metric[7]:  # pod_metrics_before
                                try:
                                    if isinstance(db_metric[7], str):
                                        pods_before = json_lib.loads(db_metric[7])
                                    else:
                                        pods_before = db_metric[7]
                                except:
                                    pass
                            
                            if db_metric[8]:  # pod_metrics_after
                                try:
                                    if isinstance(db_metric[8], str):
                                        pods_after = json_lib.loads(db_metric[8])
                                    else:
                                        pods_after = db_metric[8]
                                except:
                                    pass
                            
                            if pods_before or pods_after:
                                all_pods = set(list(pods_before.keys()) + list(pods_after.keys()))
                                for pod_name in all_pods:
                                    pod_before = pods_before.get(pod_name, {})
                                    pod_after = pods_after.get(pod_name, {})
                                    
                                    ws.cell(row=row, column=1, value=execution_id)
                                    ws.cell(row=row, column=2, value=op_idx)
                                    ws.cell(row=row, column=3, value=db_metric[0] or '')
                                    ws.cell(row=row, column=4, value=db_metric[1] or '')
                                    ws.cell(row=row, column=5, value=db_metric[6] or '')
                                    ws.cell(row=row, column=6, value=pod_name)
                                    ws.cell(row=row, column=7, value=pod_before.get('namespace') or pod_after.get('namespace', ''))
                                    ws.cell(row=row, column=8, value=pod_before.get('node') or pod_after.get('node', ''))
                                    ws.cell(row=row, column=9, value=pod_before.get('cpu_usage', 0))
                                    ws.cell(row=row, column=10, value=pod_after.get('cpu_usage', 0))
                                    ws.cell(row=row, column=11, value=(pod_after.get('cpu_usage', 0) - pod_before.get('cpu_usage', 0)))
                                    ws.cell(row=row, column=12, value=pod_before.get('memory_mb', 0))
                                    ws.cell(row=row, column=13, value=pod_after.get('memory_mb', 0))
                                    ws.cell(row=row, column=14, value=(pod_after.get('memory_mb', 0) - pod_before.get('memory_mb', 0)))
                                    ws.cell(row=row, column=15, value=pod_before.get('network_rx_mbps', 0))
                                    ws.cell(row=row, column=16, value=pod_after.get('network_rx_mbps', 0))
                                    ws.cell(row=row, column=17, value=pod_before.get('network_tx_mbps', 0))
                                    ws.cell(row=row, column=18, value=pod_after.get('network_tx_mbps', 0))
                                    ws.cell(row=row, column=19, value=db_metric[3].isoformat() if db_metric[3] else '')
                                    ws.cell(row=row, column=20, value=float(db_metric[5]) if db_metric[5] else 0)
                                    row += 1
                
                finally:
                    session.close()
                
                # Auto-adjust column widths
                for col_idx in range(1, len(headers) + 1):
                    max_length = 0
                    column = get_column_letter(col_idx)
                    for cell in ws[column]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
                
                # Save to BytesIO
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = f'attachment; filename=snapshot_individual_pods_{execution_id[:20]}.xlsx'
                return response
                
            except ImportError:
                return jsonify({
                    'success': False,
                    'error': 'Excel export requires openpyxl library. Install with: pip install openpyxl',
                    'available_formats': ['csv', 'json']
                }), 501
            except Exception as e:
                logging.exception("Error exporting Excel report")
                return jsonify({'success': False, 'error': str(e)}), 500
        
        else:
            return jsonify({'success': False, 'error': f'Unsupported format: {format}. Supported: excel'}), 400
        
    except Exception as e:
        logging.exception("Error exporting execution report")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/execution-html-report/<execution_id>', methods=['GET'])
def get_execution_html_report(execution_id):
    """
    Generate and download comprehensive HTML report for an execution
    """
    try:
        from sqlalchemy import text
        from services.html_report_service import HTMLReportService
        from flask import make_response
        
        session = SessionLocal()
        
        # Get execution details
        execution_query = text("""
            SELECT execution_id, testbed_id, status, progress, 
                   total_operations, completed_operations, successful_operations, failed_operations,
                   start_time, end_time, created_at, last_error, config, metrics
            FROM executions
            WHERE execution_id = :execution_id
        """)
        
        result = session.execute(execution_query, {'execution_id': execution_id})
        execution_row = result.fetchone()
        
        if not execution_row:
            return jsonify({'success': False, 'error': 'Execution not found'}), 404
        
        execution_data = {
            'execution_id': execution_row[0],
            'testbed_id': execution_row[1],
            'status': execution_row[2],
            'progress': execution_row[3],
            'total_operations': execution_row[4],
            'completed_operations': execution_row[5],
            'successful_operations': execution_row[6],
            'failed_operations': execution_row[7],
            'start_time': execution_row[8].isoformat() if execution_row[8] else None,
            'end_time': execution_row[9].isoformat() if execution_row[9] else None,
            'duration_minutes': ((execution_row[9] - execution_row[8]).total_seconds() / 60) if execution_row[9] and execution_row[8] else 0,
            'config': execution_row[12],
            'metrics': execution_row[13] or {}
        }
        
        # Get testbed
        testbed_query = text("""
            SELECT unique_testbed_id, testbed_label, pc_ip, ncm_ip
            FROM testbeds
            WHERE unique_testbed_id = :testbed_id
        """)
        
        testbed_result = session.execute(testbed_query, {'testbed_id': execution_data['testbed_id']})
        testbed_row = testbed_result.fetchone()
        
        testbed_data = {
            'unique_testbed_id': testbed_row[0],
            'testbed_label': testbed_row[1],
            'pc_ip': testbed_row[2],
            'ncm_ip': testbed_row[3]
        } if testbed_row else {}
        
        # Get operation metrics
        ops_query = text("""
            SELECT entity_type, operation_type, entity_name, started_at, completed_at, 
                   duration_seconds, status, error_message, pod_cpu_percent, pod_memory_mb
            FROM operation_metrics
            WHERE execution_id = :execution_id
            ORDER BY started_at ASC
        """)
        
        ops_result = session.execute(ops_query, {'execution_id': execution_id})
        operation_metrics = []
        
        for row in ops_result.fetchall():
            operation_metrics.append({
                'entity_type': row[0],
                'operation_type': row[1],
                'entity_name': row[2],
                'started_at': row[3].isoformat() if row[3] else None,
                'completed_at': row[4].isoformat() if row[4] else None,
                'duration_seconds': float(row[5]) if row[5] else 0,
                'status': row[6],
                'error_message': row[7],
                'pod_cpu_percent': float(row[8]) if row[8] else 0,
                'pod_memory_mb': float(row[9]) if row[9] else 0
            })
        
        # Get alerts
        alerts_query = text("""
            SELECT alert_name, severity, status, description, triggered_at, resolved_at
            FROM slack_alerts
            WHERE testbed_id = :testbed_id
            AND triggered_at BETWEEN :start_time AND :end_time
            ORDER BY triggered_at DESC
        """)
        
        alerts_result = session.execute(alerts_query, {
            'testbed_id': execution_data['testbed_id'],
            'start_time': execution_row[8],
            'end_time': execution_row[9] or datetime.utcnow()
        })
        
        alerts = []
        for row in alerts_result.fetchall():
            alerts.append({
                'alert_name': row[0],
                'severity': row[1],
                'status': row[2],
                'description': row[3],
                'triggered_at': row[4].isoformat() if row[4] else None,
                'resolved_at': row[5].isoformat() if row[5] else None
            })
        
        session.close()
        
        # Generate HTML report
        html_content = HTMLReportService.generate_execution_report(
            execution_data=execution_data,
            testbed_data=testbed_data,
            operation_metrics=operation_metrics,
            prometheus_metrics=execution_data.get('metrics', {}),
            alerts=alerts
        )
        
        # Return as downloadable HTML file
        response = make_response(html_content)
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename="NMT_Execution_Report_{execution_id}.html"'
        
        return response
        
    except Exception as e:
        logging.exception("Error generating HTML report")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/execution-status/<execution_id>', methods=['GET'])
def get_execution_status(execution_id):
    """
    Get execution status with progress.
    
    Response:
    {
        "success": true,
        "execution_id": "NMT-20260127-123456-abc123",
        "status": "RUNNING",
        "progress": 75.0,
        "stats": {
            "total_operations": 100,
            "completed_operations": 75,
            "successful_operations": 73,
            "failed_operations": 2
        },
        "duration_minutes": 5.2,
        "estimated_end": "2026-01-27T12:45:00"
    }
    """
    try:
        adapter = get_nmt_execution_adapter()
        result = adapter.get_execution_status(execution_id)
        
        if result and result.get('success'):
            return jsonify(result), 200
        else:
            return jsonify(result or {'success': False, 'error': 'Execution not found'}), 404
            
    except Exception as e:
        logging.exception(f"Error getting execution status: {execution_id}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/stop-execution/<execution_id>', methods=['POST'])
def stop_execution(execution_id):
    """
    Stop a running execution.
    
    Request body (optional):
    {
        "reason": "User requested stop"
    }
    
    Response:
    {
        "success": true,
        "execution_id": "NMT-20260127-123456-abc123",
        "message": "Execution stopped successfully"
    }
    """
    try:
        data = request.get_json() or {}
        reason = data.get('reason', 'User requested')
        
        adapter = get_nmt_execution_adapter()
        result = adapter.stop_execution(execution_id, reason)
        
        return jsonify(result), 200 if result['success'] else 400
            
    except Exception as e:
        logging.exception(f"Error stopping execution: {execution_id}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/pause-execution/<execution_id>', methods=['POST'])
def pause_execution(execution_id):
    """
    Pause a running execution.
    
    Response:
    {
        "success": true,
        "execution_id": "NMT-20260127-123456-abc123",
        "message": "Execution paused successfully"
    }
    """
    try:
        adapter = get_nmt_execution_adapter()
        result = adapter.pause_execution(execution_id)
        
        return jsonify(result), 200 if result['success'] else 400
            
    except Exception as e:
        logging.exception(f"Error pausing execution: {execution_id}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/resume-execution/<execution_id>', methods=['POST'])
def resume_execution(execution_id):
    """
    Resume a paused execution.
    
    Response:
    {
        "success": true,
        "execution_id": "NMT-20260127-123456-abc123",
        "message": "Execution resumed successfully"
    }
    """
    try:
        adapter = get_nmt_execution_adapter()
        result = adapter.resume_execution(execution_id)
        
        return jsonify(result), 200 if result['success'] else 400
            
    except Exception as e:
        logging.exception(f"Error resuming execution: {execution_id}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/executions', methods=['GET'])
def list_executions():
    """
    List all executions with pagination.
    
    Query parameters:
    - testbed_id: Filter by testbed ID (optional)
    - limit: Maximum number of results (default: 50)
    - offset: Offset for pagination (default: 0)
    
    Response:
    {
        "success": true,
        "executions": [...],
        "count": 10
    }
    """
    try:
        testbed_id = request.args.get('testbed_id')
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
        
        adapter = get_nmt_execution_adapter()
        result = adapter.list_executions(testbed_id, limit, offset)
        
        return jsonify(result), 200
            
    except Exception as e:
        logging.exception("Error listing executions")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/active-executions', methods=['GET'])
def list_active_executions():
    """
    List all active (non-terminal) executions.
    
    Response:
    {
        "success": true,
        "execution_ids": ["NMT-...", "NMT-..."],
        "count": 2
    }
    """
    try:
        adapter = get_nmt_execution_adapter()
        result = adapter.list_active_executions()
        
        return jsonify(result), 200
            
    except Exception as e:
        logging.exception("Error listing active executions")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/delete-execution/<execution_id>', methods=['DELETE'])
def delete_execution_endpoint(execution_id):
    """
    Delete an execution and all its associated data.
    
    This will delete:
    - Execution record from executions table
    - All operation metrics for this execution
    
    Response:
    {
        "success": true,
        "message": "Execution deleted successfully",
        "execution_id": "NMT-..."
    }
    """
    try:
        from sqlalchemy import text
        from database import SessionLocal
        
        session = SessionLocal()
        
        try:
            # First, check if execution exists
            check_query = text("SELECT execution_id FROM executions WHERE execution_id = :execution_id")
            result = session.execute(check_query, {'execution_id': execution_id})
            execution = result.fetchone()
            
            if not execution:
                return jsonify({
                    'success': False,
                    'error': 'Execution not found'
                }), 404
            
            # Delete operation metrics first (foreign key)
            delete_ops_query = text("DELETE FROM operation_metrics WHERE execution_id = :execution_id")
            ops_result = session.execute(delete_ops_query, {'execution_id': execution_id})
            ops_deleted = ops_result.rowcount
            
            # Delete execution
            delete_exec_query = text("DELETE FROM executions WHERE execution_id = :execution_id")
            exec_result = session.execute(delete_exec_query, {'execution_id': execution_id})
            
            session.commit()
            
            logging.info(f"Deleted execution {execution_id} with {ops_deleted} operation metrics")
            
            return jsonify({
                'success': True,
                'message': 'Execution deleted successfully',
                'execution_id': execution_id,
                'operations_deleted': ops_deleted
            }), 200
            
        except Exception as e:
            session.rollback()
            raise e
        finally:
            session.close()
            
    except Exception as e:
        logging.exception(f"Error deleting execution {execution_id}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================================
# TIMELINE AND METRICS ENDPOINTS
# ============================================================================

@app.route('/api/testbed-timeline/<testbed_id>', methods=['GET'])
def get_testbed_timeline(testbed_id):
    """
    Get comprehensive timeline of all activities for a testbed.
    Shows entity operations with timestamps and metrics.
    
    Query parameters:
    - limit: Maximum number of events (default: 100)
    - offset: Offset for pagination (default: 0)
    
    Response:
    {
        "success": true,
        "testbed_id": "tb-xxx",
        "timeline": [
            {
                "timestamp": "2026-01-28T10:30:00Z",
                "execution_id": "NMT-...",
                "entity_type": "vm",
                "operation_type": "create",
                "entity_name": "test-vm-001",
                "status": "COMPLETED",
                "duration_seconds": 45.3,
                "pod_cpu_percent": 25.5,
                "pod_memory_mb": 512.0,
                ...
            }
        ],
        "count": 50
    }
    """
    try:
        from database import get_testbed_timeline
        
        limit = int(request.args.get('limit', 100))
        offset = int(request.args.get('offset', 0))
        
        timeline = get_testbed_timeline(testbed_id, limit, offset)
        
        return jsonify({
            'success': True,
            'testbed_id': testbed_id,
            'timeline': timeline,
            'count': len(timeline)
        }), 200
            
    except Exception as e:
        logging.exception(f"Error getting testbed timeline: {testbed_id}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/execution-operations/<execution_id>', methods=['GET'])
def get_execution_operations(execution_id):
    """
    Get all individual operations for a specific execution.
    
    Response:
    {
        "success": true,
        "execution_id": "NMT-...",
        "operations": [
            {
                "entity_type": "vm",
                "operation_type": "create",
                "entity_name": "test-vm",
                "entity_uuid": "uuid-xxx",
                "started_at": "2026-01-28T10:30:00Z",
                "completed_at": "2026-01-28T10:30:45Z",
                "duration_seconds": 45.0,
                "status": "COMPLETED",
                "metrics_snapshot": {...},
                "pod_cpu_percent": 25.5,
                "pod_memory_mb": 512.0
            }
        ],
        "count": 10
    }
    """
    try:
        from database import get_execution_operations
        
        operations = get_execution_operations(execution_id)
        
        return jsonify({
            'success': True,
            'execution_id': execution_id,
            'operations': operations,
            'count': len(operations)
        }), 200
            
    except Exception as e:
        logging.exception(f"Error getting execution operations: {execution_id}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/testbed-metrics-history/<testbed_id>', methods=['GET'])
def get_testbed_metrics_history(testbed_id):
    """
    Get historical metrics for a testbed (continuous monitoring data).
    
    Query parameters:
    - start_time: Start time (ISO format, optional)
    - end_time: End time (ISO format, optional)
    - limit: Maximum number of records (default: 1000)
    
    Response:
    {
        "success": true,
        "testbed_id": "tb-xxx",
        "metrics": [
            {
                "collected_at": "2026-01-28T10:30:00Z",
                "cpu_percent": 45.5,
                "memory_percent": 62.3,
                "disk_percent": 35.2,
                "network_rx_mbps": 12.5,
                "network_tx_mbps": 8.3,
                "active_alerts": 2,
                "pod_metrics": {...}
            }
        ],
        "count": 500
    }
    """
    try:
        from database import get_metrics_history
        from datetime import datetime
        
        start_time = request.args.get('start_time')
        end_time = request.args.get('end_time')
        limit = int(request.args.get('limit', 1000))
        
        # Parse datetime strings if provided
        if start_time:
            start_time = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
        if end_time:
            end_time = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
        
        metrics = get_metrics_history(testbed_id, start_time, end_time, limit)
        
        return jsonify({
            'success': True,
            'testbed_id': testbed_id,
            'metrics': metrics,
            'count': len(metrics)
        }), 200
            
    except Exception as e:
        logging.exception(f"Error getting testbed metrics history: {testbed_id}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reset-execution-manager', methods=['POST'])
def reset_execution_manager_endpoint():
    """
    Reset the global execution manager instance (for testing/reload).
    
    Response:
    {
        "success": true,
        "message": "Execution manager reset successfully"
    }
    """
    try:
        from services.execution_manager import reset_execution_manager
        reset_execution_manager()
        
        return jsonify({
            'success': True,
            'message': 'Execution manager reset successfully'
        }), 200
            
    except Exception as e:
        logging.exception("Error resetting execution manager")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/execution-detailed-metrics/<execution_id>', methods=['GET'])
def get_execution_detailed_metrics(execution_id):
    """
    Get detailed metrics for an execution including:
    - Overall execution metrics (CPU, memory, network over time)
    - Individual operation metrics
    - Pod-level metrics
    - Alert correlations
    
    Response:
    {
        "success": true,
        "execution_id": "NMT-...",
        "execution_metrics": {
            "cpu": {...},
            "memory": {...},
            "network": {...},
            "alerts": [...]
        },
        "operations": [...],
        "summary": {
            "total_operations": 10,
            "avg_cpu": 35.2,
            "max_memory": 75.5,
            "total_alerts": 2
        }
    }
    """
    try:
        from database import get_execution_by_id, get_execution_operations
        import json
        
        # Get execution record with metrics
        execution = get_execution_by_id(execution_id)
        if not execution:
            return jsonify({
                'success': False,
                'error': 'Execution not found'
            }), 404
        
        # Parse metrics JSON
        execution_metrics = {}
        if execution.get('metrics'):
            try:
                if isinstance(execution['metrics'], str):
                    execution_metrics = json.loads(execution['metrics'])
                else:
                    execution_metrics = execution['metrics']
            except:
                pass
        
        # Get individual operations
        operations = get_execution_operations(execution_id)
        
        # Calculate summary statistics
        summary = {
            'total_operations': len(operations),
            'completed_operations': len([o for o in operations if o.get('status') == 'COMPLETED']),
            'failed_operations': len([o for o in operations if o.get('status') == 'FAILED']),
            'avg_cpu': 0,
            'max_cpu': 0,
            'avg_memory': 0,
            'max_memory': 0,
            'total_alerts': len(execution_metrics.get('alerts', []))
        }
        
        # Calculate averages from execution metrics
        if execution_metrics.get('cpu', {}).get('summary'):
            summary['avg_cpu'] = execution_metrics['cpu']['summary'].get('avg', 0)
            summary['max_cpu'] = execution_metrics['cpu']['summary'].get('max', 0)
        
        if execution_metrics.get('memory', {}).get('summary'):
            summary['avg_memory'] = execution_metrics['memory']['summary'].get('avg', 0)
            summary['max_memory'] = execution_metrics['memory']['summary'].get('max', 0)
        
        return jsonify({
            'success': True,
            'execution_id': execution_id,
            'execution': execution,
            'execution_metrics': execution_metrics,
            'operations': operations,
            'summary': summary
        }), 200
            
    except Exception as e:
        logging.exception(f"Error getting execution detailed metrics: {execution_id}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# LEGACY JITA ENDPOINT (Kept for backward compatibility)
# ============================================================================

@app.route('/api/run-jita-jobs', methods=['POST'])
def run_jita_jobs():
    """
    DEPRECATED: Use /api/start-execution instead.
    
    Legacy JITA endpoint - redirects to new execution system.
    """
    try:
        data = request.get_json()
        if not data or 'testbed_filepath' not in data:
            return jsonify({'success': False, 'error': 'No testbed filepath provided'}), 400

        logging.warning("DEPRECATED: /api/run-jita-jobs called. Use /api/start-execution instead.")
        
        # Redirect to new execution system
        adapter = get_nmt_execution_adapter()
        result = adapter.start_testbed_execution(data)
        
        # Return in old format for compatibility
        if result['success']:
            return jsonify({
                'success': True,
                'message': 'Execution started (via new system)',
                'execution_id': result['execution_id']
            })
        else:
            return jsonify(result), 400

    except Exception as e:
        logging.exception("Error starting execution via legacy endpoint")
        return jsonify({'success': False, 'error': str(e)}), 500

#def run_jita_jobs():
#    """
#    Execute the jita_main.py script from the nmt_backend directory
#    """
#    try:
#        logging.info("Starting Jita job execution...")
#        
#        # Path to the original jita_main.py script
#        jita_script_path = "/mnt/data/nmt_backend/jita_main.py"
#        
#        # Check if the script exists
#        if not os.path.exists(jita_script_path):
#            logging.error(f"Jita script not found at {jita_script_path}")
#            return jsonify({
#                'success': False, 
#                'error': 'Jita script not found'
#            }), 404
#        
#        # Execute the jita_main.py script
#        logging.info(f"Executing jita script: {jita_script_path}")
#        
#        result = subprocess.run(
#            ['python3', jita_script_path],
#            stdout=subprocess.PIPE,
#            stderr=subprocess.PIPE,
#            universal_newlines=True,
#            cwd='/mnt/data/nmt_backend'  # Set working directory to nmt_backend
#            # Parse output to get status to display in UI
#        )
#    
#        
#        # Parse the output to extract JITA job information
#        jita_status = parse_jita_output(result.stdout)
#        
#        # Save JITA status to file for later retrieval
#        try:
#            status_file = os.path.join(os.path.dirname(__file__), 'jita_status.json')
#            with open(status_file, 'w') as f:
#                json.dump(jita_status, f, indent=2)
#        except Exception as file_error:
#            logging.warning(f"Could not save JITA status to file: {file_error}")
#        
#        if result.returncode == 0:
#            logging.info("Jita job execution completed successfully")
#            return jsonify({
#                'success': True,
#                'message': 'Jita jobs executed successfully',
#                'jita_status': jita_status,
#                'output': result.stdout,
#                'stderr': result.stderr if result.stderr else None
#            })
#        else:
#            logging.error(f"Jita script failed with return code: {result.returncode}")
#            return jsonify({
#                'success': False,
#                'error': 'Jita script execution failed',
#                'jita_status': jita_status,
#                'output': result.stdout,
#                'stderr': result.stderr,
#                'return_code': result.returncode
#            }), 500
#            
#    except Exception as e:
#        logging.error(f"Error running Jita jobs: {str(e)}")
#        return jsonify({
#            'success': False,
#            'error': f'Failed to execute Jita jobs: {str(e)}'
#        }), 500

def parse_jita_output(output):
    """
    Parse the output from jita_main.py to extract job status information
    """
    import re
    
    jita_info = {
        'job_profile_name': None,
        'job_profile_id': None,
        'task_id': None,
        'status': 'unknown',
        'current_status': {},
        'final_status': {},
        'error': None,
        'timestamp': datetime.now().isoformat()
    }
    
    if not output:
        return jita_info
    
    lines = output.split('\n')
    
    for line in lines:
        # Extract job profile info
        if 'Fetching job profile info for:' in line:
            match = re.search(r'Fetching job profile info for: (.+)', line)
            if match:
                jita_info['job_profile_name'] = match.group(1).strip()
        
        # Extract triggered task ID
        if 'Triggered Task ID:' in line:
            match = re.search(r'Triggered Task ID: (.+)', line)
            if match:
                jita_info['task_id'] = match.group(1).strip()
                jita_info['status'] = 'triggered'
        
        # Extract current status check
        if 'Checking status of task:' in line:
            jita_info['status'] = 'checking_status'
        
        # Extract status information
        if 'Task' in line and 'status:' in line:
            match = re.search(r'Task .+ status: (.+)', line)
            if match:
                jita_info['status'] = match.group(1).strip()
        
        # Check for completion messages
        if 'completed successfully' in line.lower():
            jita_info['status'] = 'completed'
        elif 'failed' in line.lower() and 'Task' in line:
            jita_info['status'] = 'failed'
        elif 'timeout' in line.lower():
            jita_info['status'] = 'timeout'
        elif 'waiting for task' in line.lower():
            jita_info['status'] = 'waiting'
    
    # Try to extract JSON status if present
    try:
        json_blocks = re.findall(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', output)
        for json_block in json_blocks:
            try:
                parsed_json = json.loads(json_block)
                if 'task_id' in parsed_json:
                    if 'Final status:' in output and output.index(json_block) > output.index('Final status:'):
                        jita_info['final_status'] = parsed_json
                    else:
                        jita_info['current_status'] = parsed_json
            except:
                continue
    except:
        pass
    
    return jita_info

#@app.route('/api/jita-job-status/<task_id>', methods=['GET'])
#def get_jita_job_status():
#    """
#    Get the current status of a specific JITA job by task ID
#    """
#
#    try:
#        import sys
#        sys.path.append('/mnt/data/nmt_backend')
#        from jita import JitaClient
#        
#        client = JitaClient()
#        username = "svc.qa.teamjignesh"  # You might want to move these to env vars
#        password = "nN6XFS7pZ+iC?5ujB&v*"
#        
#        status = client.get_task_status(task_id, username, password)
#        
#        return jsonify({
#            'success': True,
#            'task_id': task_id,
#            'status': status
#        })
#        
#    except Exception as e:
#        logging.error(f"Error getting JITA job status: {str(e)}")
#        return jsonify({
#            'success': False,
#            'error': f'Failed to get job status: {str(e)}'
#        }), 500

#@app.route('/api/jita-job-status', methods=['GET'])
#def get_latest_jita_status():
#    """
#    Get the latest JITA job status from stored logs/files
#    """
#    
#    try:
#        #last_file = "/mnt/data/nmt_ui/prism-onboarding-ui/backend/last_filepath.json"
#        #if os.path.exists(last_file):
#        #    with open(last_file, "r") as f:
#        #        data = json.load(f)
#        #        create_uuid = data.get("create_uuid", "empty")
#        #logging.info(f"the create uuid is {create_uuid}")
#        create_uuid = "1235"
#        status_file = (f'/mnt/data/nmt_backend/jita_env_status_{create_uuid}.json')
#        logging.info(f"{status_file}")
#
#        #if os.path.exists(status_file):
#        if True:
#            logging.info(f"File exists: {status_file}")
#            with open(status_file, 'r') as f:
#                status_data = json.load(f)
#                logging.info(f"printing contents: {status_data}")
#            return jsonify({
#                'success': True,
#                'jita_status': status_data
#            })
#        else:
#            return jsonify({
#                'success': False,
#                'error': 'No JITA status data available'
#            }), 404
#            
#    except Exception as e:
#        logging.error(f"Error reading JITA status: {str(e)}")
#        return jsonify({
#            'success': False,
#            'error': f'Failed to read JITA status: {str(e)}'
#        }), 500

@app.route('/api/jita-job-status', methods=['GET'])
def get_latest_jita_status():
    """
    Get the latest JITA job status from stored logs/files
    """
    try:
        unique_testbed_id = request.args.get("unique_testbed_id")
        if not unique_testbed_id:
            return jsonify({
                'success': False,
                'error': 'Missing unique_testbed_id'
            }), 400

        status_file = f'/mnt/data/nmt_backend/jita_env_status_{unique_testbed_id}.json'
        logging.info(f"Looking for status file: {status_file}")

        if os.path.exists(status_file):
            with open(status_file, 'r') as f:
                status_data = json.load(f)
            return jsonify({'success': True, 'jita_status': status_data})
        else:
            return jsonify({'success': False, 'error': 'No JITA status data available'}), 404

    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to read JITA status: {str(e)}'}), 500
        

    

@app.route('/api/dynamic-load-job-status', methods=['GET'])
def get_latest_dynamic_load_status():
    """
    Get the latest JITA job status from stored logs/files
    """

    try:
        unique_workload_id = request.args.get("unique_workload_id")
        if not unique_workload_id:
            return jsonify({
                'success': False,
                'error': 'Missing unique_workload_id'
            }), 400

        status_file = f'/mnt/data/nmt_backend/dynamic_job_status_{unique_workload_id}.json'
        logging.info(f"Looking for status file: {status_file}")

        if os.path.exists(status_file):
            with open(status_file, 'r') as f:
                status_data = json.load(f)
            return jsonify({'success': True, 'jita_status': status_data})
        else:
            return jsonify({'success': False, 'error': 'No JITA status data available'}), 404

    except Exception as e:
        return jsonify({'success': False, 'error': f'Failed to read JITA status: {str(e)}'}), 500

# --- Dynamic jita Job Runner Endpoint ---
@app.route('/api/run-dynamic-workload', methods=['POST'])
def run_dynamic_workload():
    """ Start dynamic workload jobs in background (fire & forget). """

    data = request.get_json()
    if not data or 'unique_testbed_id' not in data:
        return jsonify({'success': False, 'error': 'Both unique_testbed_id are required'}), 400

    unique_rule_id = data['unique_rule_id']     # string
    unique_testbed_id = data['unique_workload_id']  # string
    unique_workload_id = data['unique_testbed_id']  # string

    logging.info(f"Using filepath for JITA: {unique_rule_id}")
    logging.info(f"Using unique_testbed_id: {unique_testbed_id}")
    logging.info(f"Using unique_workload_id: {unique_workload_id}")

    if not unique_testbed_id:
        logging.error("Missing filepath or unique_testbed_id in request")
        return jsonify({
            'success': False,
            'error': 'Both filepath and unique_testbed_id are required'
        }), 400
    
    #prom_status_file = f"prom_status_file_{unique_rule_id}.json"
    #prom_status_file = f"/mnt/data/nmt_backend/prometheus_{unique_rule_id}_status.json"
    prom_status_file = f"/mnt/data/nmt_backend/prometheus_status.json"
    try:
        jita_script_path = "/mnt/data/nmt_backend/dynamic_workload_generation.py"

        if not os.path.exists(jita_script_path):
            logging.error(f"Jita script not found at {jita_script_path}")
            return jsonify({
                'success': False,
                'error': 'Jita script not found'
            }), 404

        logging.info(f"Launching JITA script in background: {jita_script_path}")

        # Fire & forget (non-blocking)
        subprocess.Popen(
        [
            'python3', jita_script_path,
            '--prom_status_file', prom_status_file,
            '--unique_workload_id', unique_workload_id
        ],
        cwd='/mnt/data/nmt_backend',
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL
        )

        return jsonify({
            'success': True,
            'message': 'Dynamic workload jobs started successfully'
        })

    except Exception as e:
        logging.error(f"Error starting Dynamic workload jobs: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Failed to start Dynamic workload jobs: {str(e)}'
        }), 500

@app.route('/api/deployment-status', methods=['GET'])
def get_ncm_deployment_status():
    """
    Get the latest NCM deployment status from deployment_status_file.json
    """

    try:
        testbed_label = request.args.get("testbed_label")
        
        if not testbed_label:
            return jsonify({
                'success': False,
                'error': 'Missing testbed_label'
            }), 400

        # Use testbed_label for status file naming
        status_file = f'/mnt/data/nmt_backend/deployment_status_{testbed_label}.json'
        logging.info(f"Looking for deployment status file: {status_file}")

        if os.path.exists(status_file):
            with open(status_file, 'r') as f:
                status_data = json.load(f)
            
            # Parse the deployment status and return structured data
            # Check if this is the final comprehensive status file or intermediate status
            if 'pc_deployment' in status_data:
                # Final comprehensive status file
                pc_status = status_data.get('pc_deployment', 'not_started')
                ncm_status = status_data.get('ncm_deployment', 'not_started')
                remote_pc_status = status_data.get('remote_pc_deployment', 'not_started')
                
                # Calculate phase timing for final status (assuming 30 seconds each)
                phase_timing = {
                    'pc_duration': '30s' if pc_status == 'completed' else None,
                    'ncm_duration': '30s' if ncm_status == 'completed' else None,
                    'remote_pc_duration': '30s' if remote_pc_status == 'completed' else None
                }
                
                deployment_status = {
                    'pc_deployment': pc_status,
                    'ncm_deployment': ncm_status,
                    'remote_pc_deployment': remote_pc_status,
                    'timing': status_data.get('timing', {}),
                    'phase_timing': phase_timing,
                    'deployment_details': status_data.get('deployment_details', {}),
                    'overall_status': 'completed' if all([
                        pc_status == 'completed',
                        ncm_status == 'completed'
                    ]) else 'in_progress' if any([
                        pc_status == 'in_progress',
                        ncm_status == 'in_progress',
                        remote_pc_status == 'in_progress'
                    ]) else 'pending'
                }
            else:
                # Intermediate status file with status/message format
                current_status = status_data.get('status', 'pending')
                current_message = status_data.get('message', '')
                
                # Determine which component is being deployed based on message
                pc_status = 'not_started'
                ncm_status = 'not_started'
                remote_pc_status = 'not_started'
                
                if 'PC deployment' in current_message:
                    if current_status == 'in_progress':
                        pc_status = 'in_progress'
                    elif current_status == 'completed':
                        pc_status = 'completed'
                elif 'NCM deployment' in current_message:
                    pc_status = 'completed'  # PC must be completed to reach NCM
                    if current_status == 'in_progress':
                        ncm_status = 'in_progress'
                    elif current_status == 'completed':
                        ncm_status = 'completed'
                elif 'Remote PC deployment' in current_message:
                    pc_status = 'completed'  # PC must be completed
                    ncm_status = 'completed'  # NCM must be completed
                    if current_status == 'in_progress':
                        remote_pc_status = 'in_progress'
                    elif current_status == 'completed':
                        remote_pc_status = 'completed'
                
                # Calculate estimated timing for each phase (assuming 30 seconds each as per trigger_deployment.py)
                phase_timing = {
                    'pc_duration': '30s' if pc_status == 'completed' else None,
                    'ncm_duration': '30s' if ncm_status == 'completed' else None,
                    'remote_pc_duration': '30s' if remote_pc_status == 'completed' else None
                }
                
                deployment_status = {
                    'pc_deployment': pc_status,
                    'ncm_deployment': ncm_status,
                    'remote_pc_deployment': remote_pc_status,
                    'timing': status_data.get('timing', {}),
                    'phase_timing': phase_timing,
                    'deployment_details': status_data.get('deployment_details', {}),
                    'overall_status': current_status,
                    'current_message': current_message
                }
            
            return jsonify({
                'success': True, 
                'deployment_status': deployment_status,
                'raw_data': status_data
            })
        else:
            return jsonify({
                'success': True, 
                'deployment_status': {
                    'pc_deployment': 'not_started',
                    'ncm_deployment': 'not_started', 
                    'remote_pc_deployment': 'not_started',
                    'overall_status': 'pending'
                }
            })

    except Exception as e:
        logging.error(f"Failed to read deployment status: {str(e)}")
        return jsonify({'success': False, 'error': f'Failed to read deployment status: {str(e)}'}), 500


# =============================================================================
# ALERT API ENDPOINTS
# =============================================================================

@app.route('/api/alerts/<testbed_id>', methods=['GET'])
def get_alerts_for_testbed(testbed_id):
    """Get alerts for a specific testbed (with fake data for testing)"""
    try:
        import random
        from datetime import datetime, timedelta
        
        # Generate fake alerts for testing
        alert_names = [
            'High CPU Usage',
            'Memory Threshold Exceeded',
            'Disk Space Low',
            'Network Latency High',
            'Pod Restart Detected',
            'Service Unavailable'
        ]
        
        severities = ['Critical', 'Warning', 'Info']
        statuses = ['Active', 'Pending', 'Resolved']
        
        # Generate 0-5 random alerts
        num_alerts = random.randint(0, 5)
        fake_alerts = []
        
        for i in range(num_alerts):
            alert = {
                'id': f'alert-{testbed_id}-{i}',
                'testbed_id': testbed_id,
                'alert_name': random.choice(alert_names),
                'severity': random.choice(severities),
                'status': random.choice(statuses),
                'description': f'Test alert {i+1} for monitoring verification',
                'timestamp': (datetime.now() - timedelta(hours=random.randint(0, 24))).isoformat()
            }
            fake_alerts.append(alert)
        
        logging.info(f"Generated {num_alerts} fake alerts for testbed {testbed_id}")
        
        return jsonify({
            'success': True,
            'alerts': fake_alerts,
            'count': num_alerts,
            'note': 'These are fake alerts for testing purposes'
        })
        
    except Exception as e:
        logging.error(f"Error generating alerts for testbed {testbed_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/check-prometheus', methods=['GET'])
def check_prometheus_status():
    """Check if Prometheus is reachable"""
    try:
        import requests
        url = request.args.get('url', '')
        
        if not url:
            return jsonify({'status': 'offline', 'error': 'No URL provided'})
        
        # Try to reach Prometheus health endpoint
        try:
            # Remove trailing slash and add /-/healthy
            prometheus_url = url.rstrip('/') + '/-/healthy'
            response = requests.get(prometheus_url, timeout=5, verify=False)
            
            if response.status_code == 200:
                return jsonify({'status': 'online', 'url': url})
            else:
                return jsonify({'status': 'offline', 'error': f'HTTP {response.status_code}'})
        except requests.exceptions.RequestException as e:
            return jsonify({'status': 'offline', 'error': str(e)})
            
    except Exception as e:
        logging.error(f"Error checking Prometheus status: {e}")
        return jsonify({'status': 'offline', 'error': str(e)})


# =============================================================================
# SLACK ALERT API ENDPOINTS
# =============================================================================

@app.route('/api/test-slack-alert', methods=['POST'])
def test_slack_alert():
    """Send a test alert to Slack webhook"""
    try:
        import requests
        from datetime import datetime
        
        data = request.get_json()
        webhook_url = data.get('webhook_url', '')
        testbed_id = data.get('testbed_id', 'unknown')
        testbed_label = data.get('testbed_label', 'Test Testbed')
        
        if not webhook_url:
            return jsonify({'success': False, 'error': 'No webhook URL provided'}), 400
        
        # Prepare Slack message
        message = {
            "text": f"🧪 *NMT Test Alert*",
            "blocks": [
                {
                    "type": "header",
                    "text": {
                        "type": "plain_text",
                        "text": "🧪 NMT Test Alert",
                        "emoji": True
                    }
                },
                {
                    "type": "section",
                    "fields": [
                        {
                            "type": "mrkdwn",
                            "text": f"*Testbed:*\n{testbed_label}"
                        },
                        {
                            "type": "mrkdwn",
                            "text": f"*Testbed ID:*\n`{testbed_id[:20]}...`"
                        },
                        {
                            "type": "mrkdwn",
                            "text": f"*Status:*\n✅ Webhook Connected"
                        },
                        {
                            "type": "mrkdwn",
                            "text": f"*Time:*\n{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                        }
                    ]
                },
                {
                    "type": "section",
                    "text": {
                        "type": "mrkdwn",
                        "text": "This is a test alert from *NMT Monitoring Tool*. If you received this message, your Slack webhook is configured correctly! 🎉"
                    }
                },
                {
                    "type": "divider"
                },
                {
                    "type": "context",
                    "elements": [
                        {
                            "type": "mrkdwn",
                            "text": "📊 Powered by NMT Monitoring Tool"
                        }
                    ]
                }
            ]
        }
        
        # Send to Slack
        response = requests.post(
            webhook_url,
            json=message,
            timeout=10
        )
        
        if response.status_code == 200:
            logging.info(f"Test Slack alert sent successfully to testbed {testbed_label}")
            return jsonify({
                'success': True,
                'message': 'Test alert sent to Slack successfully!'
            })
        else:
            logging.error(f"Slack webhook returned status {response.status_code}: {response.text}")
            return jsonify({
                'success': False,
                'error': f'Slack returned status {response.status_code}'
            }), 400
            
    except requests.exceptions.Timeout:
        logging.error("Slack webhook request timed out")
        return jsonify({'success': False, 'error': 'Request timed out'}), 408
    except requests.exceptions.RequestException as e:
        logging.error(f"Error sending test Slack alert: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    except Exception as e:
        logging.error(f"Error sending test Slack alert: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/get-slack-webhook/<testbed_id>', methods=['GET'])
def get_slack_webhook(testbed_id):
    """
    Get Slack webhook URL for a testbed.
    Returns the webhook from any rule associated with this testbed.
    """
    try:
        session = SessionLocal()
        
        # Find any config for the testbed to get the webhook URL
        config = session.query(Config).filter_by(unique_testbed_id=testbed_id).first()
        
        webhook_url = None
        if config and config.config_json:
            webhook_url = config.config_json.get('alert_destination', {}).get('value')
        
        session.close()
        
        return jsonify({
            'success': True,
            'webhook_url': webhook_url
        })
        
    except Exception as e:
        logging.error(f"Error fetching Slack webhook for testbed {testbed_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/update-slack-webhook', methods=['POST'])
def update_slack_webhook():
    """Update Slack webhook URL for a testbed's rules"""
    try:
        data = request.get_json()
        testbed_id = data.get('testbed_id', '')
        webhook_url = data.get('webhook_url', '')
        
        if not testbed_id or not webhook_url:
            return jsonify({'success': False, 'error': 'Missing testbed_id or webhook_url'}), 400
        
        # Get all rules for this testbed
        session = SessionLocal()
        configs = session.query(Config).filter(
            Config.unique_testbed_id == testbed_id
        ).all()
        
        if not configs:
            session.close()
            return jsonify({'success': False, 'error': 'No rules found for this testbed'}), 404
        
        # Update webhook URL in each rule's config
        updated_count = 0
        for config in configs:
            try:
                config_data = config.config
                if not config_data:
                    config_data = {}
                
                # Update alert_destination
                config_data['alert_destination'] = {
                    'type': 'slack',
                    'value': webhook_url
                }
                
                config.config = config_data
                updated_count += 1
            except Exception as e:
                logging.error(f"Error updating config {config.unique_rule_id}: {e}")
        
        session.commit()
        session.close()
        
        logging.info(f"Updated Slack webhook for {updated_count} rule(s) in testbed {testbed_id}")
        
        return jsonify({
            'success': True,
            'message': f'Slack webhook updated for {updated_count} rule(s)',
            'updated_count': updated_count
        })
        
    except Exception as e:
        logging.error(f"Error updating Slack webhook: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# EXECUTION HISTORY API ENDPOINT
# =============================================================================

@app.route('/api/execution-history', methods=['GET'])
def get_execution_history():
    """Get execution history, optionally filtered by testbed_id"""
    try:
        testbed_id = request.args.get('testbed_id', None)
        
        # Import here to avoid circular dependencies
        from database import SessionLocal
        from models.execution import ExecutionRecord
        
        session = SessionLocal()
        
        try:
            query = session.query(ExecutionRecord)
            
            if testbed_id:
                query = query.filter(ExecutionRecord.testbed_id == testbed_id)
            
            # Order by most recent first
            query = query.order_by(ExecutionRecord.started_at.desc())
            
            executions = query.all()
            
            # Convert to dict
            history = []
            for exec_record in executions:
                history.append({
                    'execution_id': exec_record.execution_id,
                    'testbed_id': exec_record.testbed_id,
                    'status': exec_record.status,
                    'started_at': exec_record.started_at.isoformat() if exec_record.started_at else None,
                    'completed_at': exec_record.completed_at.isoformat() if exec_record.completed_at else None,
                    'duration_minutes': exec_record.duration_minutes,
                    'total_operations': exec_record.total_operations,
                    'successful_operations': exec_record.successful_operations,
                    'failed_operations': exec_record.failed_operations
                })
            
            return jsonify({
                'success': True,
                'history': history,
                'count': len(history)
            })
            
        finally:
            session.close()
        
    except Exception as e:
        logging.error(f"Error fetching execution history: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ===========================
# Smart Execution APIs
# ===========================

@app.route('/api/smart-execution/start', methods=['POST'])
def start_smart_execution_api():
    """
    Start a smart execution with AI/ML-powered threshold-based control
    
    This endpoint has been upgraded to use the AI/ML system (PID + ML).
    The old non-AI system has been deprecated.
    
    Request:
    {
        "testbed_id": "unique_testbed_id",
        "target_config": {
            "cpu_threshold": 80,
            "memory_threshold": 80,
            "stop_condition": "any"  // "any" or "all"
        },
        "entities_config": {
            "VM": {"create": 5, "delete": 5},
            "Project": {"create": 2}
        },
        "namespaces": ["ntnx-system", "default"],
        "pods": [],
        "ai_settings": {
            "enable_ai": true,
            "enable_ml": true
        }
    }
    
    Response:
    {
        "success": true,
        "execution_id": "SMART-20260204-...",
        "message": "Smart execution started",
        "ai_enabled": true,
        "ml_enabled": true
    }
    """
    try:
        # This endpoint now uses the AI/ML-powered system
        # Redirecting to the new AI execution handler
        from services.smart_execution_service import start_smart_execution
        from sqlalchemy import text
        
        data = request.get_json()
        testbed_id = data.get('testbed_id')
        target_config = data.get('target_config', {})
        entities_config = data.get('entities_config', {})
        rule_config = data.get('rule_config', {})
        
        # Pass through advanced execution settings into target_config
        if 'advanced' in data:
            target_config['advanced'] = data['advanced']
        
        if not testbed_id:
            return jsonify({'success': False, 'error': 'testbed_id is required'}), 400
        
        if not target_config:
            return jsonify({'success': False, 'error': 'target_config is required'}), 400
        
        if not entities_config:
            return jsonify({'success': False, 'error': 'entities_config is required'}), 400
        
        # Handle both old and new entity_config formats
        # Old: {"VM": ["create", "delete"]}
        # New: {"vm": {"create": 5, "delete": 5}}
        normalized_entities = {}
        for entity, ops in entities_config.items():
            entity_lower = entity.lower().replace(' ', '_')
            if isinstance(ops, list):
                # Old format - convert to new format with default counts
                normalized_entities[entity_lower] = {op: 5 for op in ops}
            elif isinstance(ops, dict):
                # New format - use as is
                normalized_entities[entity_lower] = ops
        
        # Get testbed details
        session = SessionLocal()
        query = text("""
            SELECT unique_testbed_id, testbed_label, pc_ip, ncm_ip, username, password, testbed_json
            FROM testbeds
            WHERE unique_testbed_id = :testbed_id
        """)
        
        result = session.execute(query, {'testbed_id': testbed_id})
        testbed_row = result.fetchone()
        session.close()
        
        if not testbed_row:
            return jsonify({'success': False, 'error': f'Testbed {testbed_id} not found'}), 404
        
        testbed_info = {
            'unique_testbed_id': testbed_row[0],
            'testbed_label': testbed_row[1],
            'pc_ip': testbed_row[2],
            'ncm_ip': testbed_row[3],
            'username': testbed_row[4],
            'password': testbed_row[5]
        }
        
        # Add testbed_json data
        if len(testbed_row) > 6 and testbed_row[6]:
            testbed_json = testbed_row[6]
            if isinstance(testbed_json, str):
                import json
                testbed_json = json.loads(testbed_json)
            if isinstance(testbed_json, dict):
                testbed_info.update(testbed_json)
        
        logging.info(f"🚀 Starting AI-powered smart execution for testbed {testbed_info.get('testbed_label')}")
        
        # Use the existing start_smart_execution which now has AI capabilities
        execution_id = start_smart_execution(testbed_info, target_config, normalized_entities, rule_config)
        
        logging.info(f"✅ Smart execution started: {execution_id}")
        
        return jsonify({
            'success': True,
            'execution_id': execution_id,
            'message': 'Smart execution started successfully',
            'ai_enabled': True
        }), 200
        
    except Exception as e:
        logging.exception("Error starting smart execution")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/status/<execution_id>', methods=['GET'])
def get_smart_execution_status(execution_id):
    """
    Get status of a smart execution
    
    Response:
    {
        "success": true,
        "execution_id": "SMART-...",
        "status": "RUNNING",
        "is_running": true,
        "total_operations": 25,
        "current_metrics": {"cpu_percent": 75.2, "memory_percent": 68.5},
        "baseline_metrics": {"cpu_percent": 20.1, "memory_percent": 30.2},
        "target_config": {...},
        "operations_history": [...],
        "metrics_history": [...]
    }
    """
    try:
        from services.smart_execution_service import get_smart_execution
        from services.smart_execution_db import load_smart_execution
        
        controller = get_smart_execution(execution_id)
        
        if controller:
            status = controller.get_status()
            status['success'] = True
            return jsonify(status), 200
        
        # Try loading from database
        db_data = load_smart_execution(execution_id)
        if db_data:
            # Convert to status format
            status = {
                'success': True,
                'execution_id': db_data.get('execution_id'),
                'status': db_data.get('status', 'UNKNOWN'),
                'is_running': db_data.get('is_running', False),
                'start_time': db_data.get('start_time'),
                'end_time': db_data.get('end_time'),
                'duration_minutes': db_data.get('duration_minutes', 0),
                'total_operations': db_data.get('total_operations', 0),
                'successful_operations': db_data.get('successful_operations', 0),
                'failed_operations': db_data.get('failed_operations', 0),
                'success_rate': db_data.get('success_rate', 0),
                'operations_per_minute': db_data.get('operations_per_minute', 0),
                'target_config': db_data.get('target_config', {}),
                'baseline_metrics': db_data.get('baseline_metrics', {}),
                'current_metrics': db_data.get('final_metrics', {}),
                'operations_history': db_data.get('operations_history', [])[-10:],
                'metrics_history': db_data.get('metrics_history', [])[-20:],
                'threshold_reached': db_data.get('threshold_reached', False),
                'entity_breakdown': db_data.get('entity_breakdown', {}),
                'testbed_info': {
                    'testbed_label': db_data.get('testbed_label', 'Unknown'),
                    'testbed_id': db_data.get('testbed_id')
                }
            }
            # Include full execution data if available
            full_data = db_data.get('full_execution_data', {})
            if full_data:
                status.update({
                    'predictions': full_data.get('predictions'),
                    'detected_anomalies': full_data.get('detected_anomalies', []),
                    'recommendations': full_data.get('recommendations', []),
                    'operation_effectiveness': full_data.get('operation_effectiveness', [])
                })
            return jsonify(status), 200
        
        return jsonify({'success': False, 'error': 'Execution not found'}), 404
        
    except Exception as e:
        logging.exception("Error getting smart execution status")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/pre-check', methods=['POST'])
def pre_execution_check_api():
    """Run pre-execution resource validation without starting an execution."""
    try:
        from services.smart_execution_service import SmartExecutionController
        from sqlalchemy import text
        import asyncio

        data = request.get_json()
        testbed_id = data.get('testbed_id')
        if not testbed_id:
            return jsonify({'success': False, 'error': 'testbed_id is required'}), 400

        session = SessionLocal()
        query = text("""
            SELECT unique_testbed_id, testbed_label, pc_ip, ncm_ip, username, password, testbed_json
            FROM testbeds WHERE unique_testbed_id = :testbed_id
        """)
        result = session.execute(query, {'testbed_id': testbed_id})
        testbed_row = result.fetchone()
        session.close()

        if not testbed_row:
            return jsonify({'success': False, 'error': 'Testbed not found'}), 404

        testbed_info = {
            'unique_testbed_id': testbed_row[0],
            'testbed_label': testbed_row[1],
            'pc_ip': testbed_row[2],
            'ncm_ip': testbed_row[3],
            'username': testbed_row[4],
            'password': testbed_row[5]
        }
        if len(testbed_row) > 6 and testbed_row[6]:
            tj = testbed_row[6]
            if isinstance(tj, str):
                import json
                tj = json.loads(tj)
            if isinstance(tj, dict):
                testbed_info.update(tj)

        target_config = data.get('target_config', {'cpu_threshold': 80, 'memory_threshold': 80, 'stop_condition': 'any'})
        entities_config = data.get('entities_config', {'vm': ['create']})

        controller = SmartExecutionController(testbed_info, target_config, entities_config)
        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

        checks = loop.run_until_complete(controller.pre_execution_check())
        return jsonify({'success': True, 'checks': checks}), 200

    except Exception as e:
        logging.exception("Error in pre-execution check")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/latency/<execution_id>', methods=['GET'])
def get_execution_latency(execution_id):
    """Get API latency statistics for a running/completed execution."""
    try:
        from services.smart_execution_service import get_smart_execution
        controller = get_smart_execution(execution_id)
        if not controller:
            return jsonify({'success': False, 'error': 'Execution not found or not in memory'}), 404

        return jsonify({
            'success': True,
            'execution_id': execution_id,
            'latency_summary': controller.get_latency_summary(),
            'recent_latencies': controller.api_latency_history[-20:]
        }), 200

    except Exception as e:
        logging.exception("Error getting latency data")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/stop/<execution_id>', methods=['POST'])
def stop_smart_execution_api(execution_id):
    """
    Stop a running smart execution
    
    Response:
    {
        "success": true,
        "message": "Execution stopped"
    }
    """
    try:
        from services.smart_execution_service import stop_smart_execution
        
        success = stop_smart_execution(execution_id)
        
        if success:
            logging.info(f"✅ Smart execution stopped: {execution_id}")
            return jsonify({'success': True, 'message': 'Execution stopped'}), 200
        else:
            return jsonify({'success': False, 'error': 'Execution not found'}), 404
        
    except Exception as e:
        logging.exception("Error stopping smart execution")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/pause/<execution_id>', methods=['POST'])
def pause_smart_execution_api(execution_id):
    """
    Pause a running smart execution
    
    Response:
    {
        "success": true,
        "message": "Execution paused"
    }
    """
    try:
        from services.smart_execution_service import pause_smart_execution
        
        result = pause_smart_execution(execution_id)
        
        if result['success']:
            logging.info(f"⏸️  Smart execution paused: {execution_id}")
            return jsonify(result), 200
        else:
            return jsonify(result), 400
        
    except Exception as e:
        logging.exception("Error pausing smart execution")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/resume/<execution_id>', methods=['POST'])
def resume_smart_execution_api(execution_id):
    """
    Resume a paused smart execution
    
    Response:
    {
        "success": true,
        "message": "Execution resumed"
    }
    """
    try:
        from services.smart_execution_service import resume_smart_execution
        
        result = resume_smart_execution(execution_id)
        
        if result['success']:
            logging.info(f"▶️  Smart execution resumed: {execution_id}")
            return jsonify(result), 200
        else:
            return jsonify(result), 400
        
    except Exception as e:
        logging.exception("Error resuming smart execution")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/logs/<execution_id>', methods=['GET'])
def get_smart_execution_logs_api(execution_id):
    """
    Get live logs for a smart execution
    
    Query Params:
    - since (optional): ISO timestamp to get logs after
    - limit (optional): Max number of logs to return (default 100)
    
    Response:
    {
        "success": true,
        "logs": [...]
    }
    """
    try:
        from services.smart_execution_service import get_smart_execution_logs
        
        since = request.args.get('since')
        limit = int(request.args.get('limit', 100))
        
        logs = get_smart_execution_logs(execution_id, since=since, limit=limit)
        
        return jsonify({'success': True, 'logs': logs}), 200
        
    except Exception as e:
        logging.exception("Error getting smart execution logs")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/cleanup/<execution_id>', methods=['POST'])
def cleanup_smart_execution_entities(execution_id):
    """
    Cleanup (delete) all entities created during a smart execution
    
    Optional body:
    {
        "entity_types": ["VM", "Project"]  # Optional: specific types to cleanup
    }
    
    Response:
    {
        "success": true,
        "cleanup_summary": {
            "total": 10,
            "success": 8,
            "failed": 1,
            "skipped": 1,
            "results": {...}
        }
    }
    """
    try:
        from services.smart_execution_service import get_smart_execution
        import asyncio
        
        controller = get_smart_execution(execution_id)
        if not controller:
            return jsonify({'success': False, 'error': 'Execution not found'}), 404
        
        # Run cleanup in asyncio event loop
        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
        
        cleanup_summary = loop.run_until_complete(controller.cleanup_entities())
        
        logging.info(f"🧹 Cleanup complete for {execution_id}: {cleanup_summary['success']}/{cleanup_summary['total']} deleted")
        
        return jsonify({
            'success': True,
            'cleanup_summary': cleanup_summary
        }), 200
        
    except Exception as e:
        logging.exception(f"Error cleaning up entities for execution {execution_id}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/purge-stale', methods=['POST'])
def purge_stale_executions():
    """Delete old/broken executions from the database.
    
    Body (all optional):
    {
        "before_date": "2026-03-03",       # delete executions older than this
        "max_success_rate": 10,             # delete executions with success rate <= this %
        "execution_ids": ["SMART-..."],     # delete specific execution IDs
        "dry_run": true                     # preview what would be deleted
    }
    """
    session = None
    try:
        data = request.get_json(silent=True) or {}
        dry_run = data.get('dry_run', False)
        
        from models.smart_execution import SmartExecution
        from database import SessionLocal
        
        session = SessionLocal()
        query = session.query(SmartExecution)
        
        filters_applied = []
        
        if data.get('execution_ids'):
            query = query.filter(SmartExecution.execution_id.in_(data['execution_ids']))
            filters_applied.append(f"specific IDs: {data['execution_ids']}")
        else:
            conditions = []
            if data.get('before_date'):
                from datetime import datetime as dt
                cutoff = dt.fromisoformat(data['before_date'])
                conditions.append(SmartExecution.start_time < cutoff)
                filters_applied.append(f"before {data['before_date']}")
            if data.get('max_success_rate') is not None:
                max_rate = float(data['max_success_rate'])
                conditions.append(SmartExecution.success_rate <= max_rate)
                filters_applied.append(f"success_rate <= {max_rate}%")
            if conditions:
                query = query.filter(*conditions)
            else:
                session.close()
                return jsonify({'success': False, 'error': 'No filter criteria provided'}), 400
        
        candidates = query.all()
        preview = [{
            'execution_id': e.execution_id,
            'status': e.status,
            'success_rate': e.success_rate,
            'start_time': e.start_time.isoformat() if e.start_time else None,
            'total_operations': e.total_operations,
        } for e in candidates]
        
        if dry_run:
            session.close()
            return jsonify({
                'success': True, 'dry_run': True,
                'count': len(preview), 'candidates': preview,
                'filters': filters_applied,
            }), 200
        
        deleted_count = 0
        for execution in candidates:
            session.delete(execution)
            deleted_count += 1
        session.commit()
        session.close()
        
        logging.info(f"Purged {deleted_count} stale executions (filters: {filters_applied})")
        
        return jsonify({
            'success': True, 'deleted_count': deleted_count,
            'deleted': preview, 'filters': filters_applied,
        }), 200
    except Exception as e:
        if session:
            try:
                session.rollback()
                session.close()
            except Exception:
                pass
        logging.exception("Error purging stale executions")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/history', methods=['GET'])
def get_smart_execution_history():
    """
    Get list of all smart executions with Phase 2: Advanced filtering
    
    Query Parameters:
    - status: Filter by status (RUNNING, COMPLETED, STOPPED, FAILED)
    - testbed_id: Filter by testbed ID
    - date_from: Filter from date (ISO format)
    - date_to: Filter to date (ISO format)
    - min_success_rate: Minimum success rate (0-100)
    - min_operations: Minimum operations count
    - threshold_reached: Filter by threshold reached (true/false)
    
    Response:
    {
        "success": true,
        "executions": [...],
        "total": 45,
        "filtered": 12,
        "summary": {
            "avg_success_rate": 94.2,
            "avg_operations_per_min": 8.5,
            "avg_duration_minutes": 12.3
        }
    }
    """
    try:
        from services.smart_execution_service import get_all_smart_executions
        from datetime import datetime
        
        # Get filter parameters
        status_filter = request.args.get('status', 'all')
        testbed_filter = request.args.get('testbed_id', 'all')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        min_success_rate = request.args.get('min_success_rate', type=float)
        min_operations = request.args.get('min_operations', type=int)
        threshold_reached_filter = request.args.get('threshold_reached')
        
        executions_list = get_all_smart_executions()
        
        # Enrich executions with target and final metrics for frontend
        enriched_executions = []
        for exec_data in executions_list:
            enriched = exec_data.copy()
            
            # Extract target config
            target_config = exec_data.get('target_config', {})
            if isinstance(target_config, dict):
                enriched['target_cpu'] = target_config.get('cpu_threshold', 0)
                enriched['target_memory'] = target_config.get('memory_threshold', 0)
            else:
                enriched['target_cpu'] = 0
                enriched['target_memory'] = 0
            
            # Extract final metrics
            final_metrics = exec_data.get('final_metrics', {}) or exec_data.get('current_metrics', {})
            if isinstance(final_metrics, dict):
                enriched['final_cpu'] = final_metrics.get('cpu_percent', 0)
                enriched['final_memory'] = final_metrics.get('memory_percent', 0)
            else:
                enriched['final_cpu'] = 0
                enriched['final_memory'] = 0
            
            # Include new fields in history listing
            enriched['anomaly_count'] = exec_data.get('anomaly_count', 0)
            enriched['anomaly_high_count'] = exec_data.get('anomaly_high_count', 0)
            enriched['tags'] = exec_data.get('tags', [])
            enriched['learning_summary'] = exec_data.get('learning_summary')
            enriched['latency_avg'] = (exec_data.get('latency_summary') or {}).get('overall', {}).get('avg')
            
            if enriched.get('duration_minutes') is None or (isinstance(enriched.get('duration_minutes'), float) and (enriched.get('duration_minutes') != enriched.get('duration_minutes'))):  # NaN check
                if enriched.get('start_time') and enriched.get('end_time'):
                    try:
                        start = datetime.fromisoformat(enriched['start_time'].replace('Z', '+00:00'))
                        end = datetime.fromisoformat(enriched['end_time'].replace('Z', '+00:00'))
                        enriched['duration_minutes'] = (end - start).total_seconds() / 60
                    except:
                        enriched['duration_minutes'] = 0
                else:
                    enriched['duration_minutes'] = 0
            
            enriched_executions.append(enriched)
        
        # Apply filters
        filtered_executions = enriched_executions
        
        if status_filter != 'all':
            filtered_executions = [e for e in filtered_executions if e.get('status') == status_filter]
        
        if testbed_filter != 'all':
            filtered_executions = [e for e in filtered_executions if e.get('testbed_id') == testbed_filter]
        
        if date_from:
            try:
                date_from_dt = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
                filtered_executions = [
                    e for e in filtered_executions 
                    if e.get('start_time') and datetime.fromisoformat(e['start_time'].replace('Z', '+00:00')) >= date_from_dt
                ]
            except:
                pass
        
        if date_to:
            try:
                date_to_dt = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
                filtered_executions = [
                    e for e in filtered_executions 
                    if e.get('start_time') and datetime.fromisoformat(e['start_time'].replace('Z', '+00:00')) <= date_to_dt
                ]
            except:
                pass
        
        if min_success_rate is not None:
            filtered_executions = [
                e for e in filtered_executions 
                if e.get('success_rate', 0) >= min_success_rate
            ]
        
        if min_operations is not None:
            filtered_executions = [
                e for e in filtered_executions 
                if e.get('total_operations', 0) >= min_operations
            ]
        
        if threshold_reached_filter:
            threshold_bool = threshold_reached_filter.lower() == 'true'
            filtered_executions = [
                e for e in filtered_executions 
                if e.get('threshold_reached', False) == threshold_bool
            ]
        
        # Calculate summary statistics
        if filtered_executions:
            success_rates = [e.get('success_rate', 0) for e in filtered_executions if e.get('success_rate')]
            ops_per_min = [e.get('operations_per_minute', 0) for e in filtered_executions if e.get('operations_per_minute')]
            durations = [e.get('duration_minutes', 0) for e in filtered_executions if e.get('duration_minutes')]
            
            summary = {
                'avg_success_rate': sum(success_rates) / len(success_rates) if success_rates else 0,
                'avg_operations_per_min': sum(ops_per_min) / len(ops_per_min) if ops_per_min else 0,
                'avg_duration_minutes': sum(durations) / len(durations) if durations else 0,
                'total_executions': len(filtered_executions),
                'completed_count': sum(1 for e in filtered_executions if e.get('status') == 'COMPLETED'),
                'threshold_reached_count': sum(1 for e in filtered_executions if e.get('threshold_reached', False))
            }
        else:
            summary = {
                'avg_success_rate': 0,
                'avg_operations_per_min': 0,
                'avg_duration_minutes': 0,
                'total_executions': 0,
                'completed_count': 0,
                'threshold_reached_count': 0
            }
        
        return jsonify({
            'success': True,
            'executions': filtered_executions,
            'total': len(executions_list),
            'filtered': len(filtered_executions),
            'summary': summary
        }), 200
            
    except Exception as e:
        logging.exception("Error fetching smart execution history")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/smart-execution/compare', methods=['POST'])
def compare_smart_executions():
    """
    Phase 2: Compare multiple smart executions side-by-side
    
    Request:
    {
        "execution_ids": ["SMART-1", "SMART-2", "SMART-3"]
    }
    
    Response:
    {
        "success": true,
        "comparisons": [
            {
                "execution_id": "SMART-1",
                "testbed_label": "...",
                "duration_minutes": 12.5,
                "success_rate": 95.5,
                "operations_per_minute": 8.2,
                "cpu_change": 35.2,
                "memory_change": 28.5,
                ...
            }
        ],
        "summary": {
            "fastest": "SMART-1",
            "most_efficient": "SMART-2",
            "highest_success": "SMART-3"
        }
    }
    """
    try:
        from services.smart_execution_service import get_smart_execution
        
        data = request.get_json()
        execution_ids = data.get('execution_ids', [])
        
        if not execution_ids or len(execution_ids) < 2:
            return jsonify({'success': False, 'error': 'At least 2 execution IDs required'}), 400
        
        comparisons = []
        for exec_id in execution_ids:
            controller = get_smart_execution(exec_id)
            entry = None
            if controller:
                status = controller.get_status()
                entry = {
                    'execution_id': exec_id,
                    'testbed_label': status.get('execution_context', {}).get('testbed_label', 'Unknown'),
                    'status': status.get('status', 'UNKNOWN'),
                    'duration_minutes': status.get('duration_minutes', 0),
                    'success_rate': status.get('success_rate', 0),
                    'operations_per_minute': status.get('operations_per_minute', 0),
                    'total_operations': status.get('total_operations', 0),
                    'baseline_cpu': status.get('baseline_metrics', {}).get('cpu_percent', 0),
                    'final_cpu': status.get('current_metrics', {}).get('cpu_percent', 0),
                    'baseline_memory': status.get('baseline_metrics', {}).get('memory_percent', 0),
                    'final_memory': status.get('current_metrics', {}).get('memory_percent', 0),
                    'threshold_reached': status.get('threshold_reached', False),
                    'start_time': status.get('start_time'),
                    'end_time': status.get('end_time'),
                    'anomaly_count': len(status.get('detected_anomalies', [])),
                    'latency_avg': (status.get('latency_summary') or {}).get('overall', {}).get('avg'),
                    'learning_summary': status.get('learning_summary'),
                    'tags': status.get('tags', []),
                }
            else:
                from services.smart_execution_db import load_smart_execution
                db_data = load_smart_execution(exec_id)
                if db_data:
                    bl = db_data.get('baseline_metrics') or {}
                    fm = db_data.get('final_metrics') or {}
                    entry = {
                        'execution_id': exec_id,
                        'testbed_label': db_data.get('testbed_label', 'Unknown'),
                        'status': db_data.get('status', 'UNKNOWN'),
                        'duration_minutes': db_data.get('duration_minutes', 0),
                        'success_rate': db_data.get('success_rate', 0),
                        'operations_per_minute': db_data.get('operations_per_minute', 0),
                        'total_operations': db_data.get('total_operations', 0),
                        'baseline_cpu': bl.get('cpu_percent', 0),
                        'final_cpu': fm.get('cpu_percent', 0),
                        'baseline_memory': bl.get('memory_percent', 0),
                        'final_memory': fm.get('memory_percent', 0),
                        'threshold_reached': db_data.get('threshold_reached', False),
                        'start_time': db_data.get('start_time'),
                        'end_time': db_data.get('end_time'),
                        'anomaly_count': db_data.get('anomaly_count', 0),
                        'latency_avg': (db_data.get('latency_summary') or {}).get('overall', {}).get('avg'),
                        'learning_summary': db_data.get('learning_summary'),
                        'tags': db_data.get('tags', []),
                    }
            if entry:
                entry['cpu_change'] = entry['final_cpu'] - entry['baseline_cpu']
                entry['memory_change'] = entry['final_memory'] - entry['baseline_memory']
                comparisons.append(entry)
        
        # Calculate summary
        if comparisons:
            fastest = min(comparisons, key=lambda x: x['duration_minutes'] if x['duration_minutes'] > 0 else float('inf'))
            most_efficient = max(comparisons, key=lambda x: x['operations_per_minute'])
            highest_success = max(comparisons, key=lambda x: x['success_rate'])
            
            summary = {
                'fastest': fastest['execution_id'],
                'most_efficient': most_efficient['execution_id'],
                'highest_success': highest_success['execution_id']
            }
        else:
            summary = {}
        
        return jsonify({
            'success': True,
            'comparisons': comparisons,
            'summary': summary
        }), 200
        
    except Exception as e:
        logging.exception("Error comparing smart executions")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/csv/<execution_id>', methods=['GET'])
def export_execution_csv(execution_id):
    """Export operation-level CSV for an execution."""
    try:
        from services.smart_execution_service import get_smart_execution
        from services.smart_execution_db import load_smart_execution
        import csv, io

        ops_history = None
        controller = get_smart_execution(execution_id)
        if controller:
            ops_history = controller.operations_history
        else:
            db_data = load_smart_execution(execution_id)
            if db_data:
                ops_history = db_data.get('operations_history', [])

        if ops_history is None:
            return jsonify({'success': False, 'error': 'Execution not found'}), 404

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['timestamp', 'entity_type', 'operation', 'entity_name', 'status',
                         'duration_seconds', 'error'])
        for op in ops_history:
            writer.writerow([
                op.get('start_time', ''),
                op.get('entity_type', ''),
                op.get('operation', ''),
                op.get('entity_name', ''),
                op.get('status', ''),
                op.get('duration_seconds', ''),
                op.get('error', ''),
            ])

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=operations-{execution_id[:20]}.csv'}
        )
    except Exception as e:
        logging.exception("Error exporting CSV")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/tags/<execution_id>', methods=['PUT'])
def update_execution_tags(execution_id):
    """Update tags/labels for an execution."""
    try:
        from sqlalchemy import text
        data = request.get_json()
        tags = data.get('tags', [])
        if not isinstance(tags, list):
            return jsonify({'success': False, 'error': 'tags must be a list'}), 400

        session = SessionLocal()
        try:
            query = text("UPDATE smart_executions SET tags = :tags WHERE execution_id = :eid")
            import json
            result = session.execute(query, {'tags': json.dumps(tags), 'eid': execution_id})
            session.commit()
            updated = result.rowcount > 0
        finally:
            session.close()

        if updated:
            return jsonify({'success': True, 'tags': tags}), 200
        return jsonify({'success': False, 'error': 'Execution not found'}), 404
    except Exception as e:
        logging.exception("Error updating tags")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/rerun-config/<execution_id>', methods=['GET'])
def get_rerun_config(execution_id):
    """Get the configuration used by a past execution for re-running."""
    try:
        from services.smart_execution_db import load_smart_execution
        db_data = load_smart_execution(execution_id)
        if not db_data:
            return jsonify({'success': False, 'error': 'Execution not found'}), 404

        return jsonify({
            'success': True,
            'config': {
                'testbed_id': db_data.get('testbed_id'),
                'testbed_label': db_data.get('testbed_label'),
                'target_config': db_data.get('target_config', {}),
                'entities_config': db_data.get('entities_config', {}),
                'rule_config': db_data.get('rule_config', {}),
                'tags': db_data.get('tags', []),
                'alert_thresholds': db_data.get('alert_thresholds', {}),
            }
        }), 200
    except Exception as e:
        logging.exception("Error getting rerun config")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/report/<execution_id>/download', methods=['GET'])
def download_smart_execution_report(execution_id):
    """
    Download detailed HTML report of a smart execution
    """
    try:
        from services.smart_execution_service import get_smart_execution
        from services.smart_execution_db import load_smart_execution
        
        # Initialize status and report_data to avoid None errors
        status = None
        report_data = None
        
        # Try to get from memory first
        controller = get_smart_execution(execution_id)
        if controller:
            status = controller.get_status()
            report_data = controller.get_report()
        else:
            # Load from database
            db_data = None
            try:
                db_data = load_smart_execution(execution_id)
            except Exception as e:
                return jsonify({'success': False, 'error': f'Error loading execution: {str(e)}'}), 500
            
            # Validate db_data
            if db_data is None:
                return jsonify({'success': False, 'error': 'Execution not found'}), 404
            
            if not isinstance(db_data, dict):
                return jsonify({'success': False, 'error': 'Invalid execution data format'}), 500
            
            # Now safely access db_data - ensure it's a dict first
            if not isinstance(db_data, dict):
                return jsonify({'success': False, 'error': 'Execution data format invalid'}), 500
            
            # Extract full_execution_data safely
            full_execution_data = db_data.get('full_execution_data')
            if full_execution_data is None or not isinstance(full_execution_data, dict):
                full_execution_data = {}
            
            # Create status dict from db_data
            try:
                status = {
                    'execution_id': db_data.get('execution_id') or execution_id,
                    'status': db_data.get('status') or 'UNKNOWN',
                    'start_time': db_data.get('start_time'),
                    'end_time': db_data.get('end_time'),
                    'duration_minutes': db_data.get('duration_minutes') or 0,
                    'total_operations': db_data.get('total_operations') or 0,
                    'successful_operations': db_data.get('successful_operations') or 0,
                    'failed_operations': db_data.get('failed_operations') or 0,
                    'success_rate': db_data.get('success_rate') or 0,
                    'operations_per_minute': db_data.get('operations_per_minute') or 0,
                    'target_config': db_data.get('target_config') or {},
                    'baseline_metrics': db_data.get('baseline_metrics') or {},
                    'current_metrics': db_data.get('final_metrics') or {},
                    'operations_history': db_data.get('operations_history') or [],
                    'metrics_history': db_data.get('metrics_history') or [],
                    'threshold_reached': db_data.get('threshold_reached') or False,
                    'entity_breakdown': db_data.get('entity_breakdown') or {},
                    'testbed_info': {
                        'testbed_label': db_data.get('testbed_label') or 'Unknown',
                        'testbed_id': db_data.get('testbed_id') or 'unknown'
                    },
                    'predictions': full_execution_data.get('predictions') if isinstance(full_execution_data, dict) else None,
                    'detected_anomalies': full_execution_data.get('detected_anomalies', []) if isinstance(full_execution_data, dict) else [],
                    'recommendations': full_execution_data.get('recommendations', []) if isinstance(full_execution_data, dict) else [],
                    'operation_effectiveness': full_execution_data.get('operation_effectiveness', []) if isinstance(full_execution_data, dict) else [],
                    'resource_summary': full_execution_data.get('resource_summary', {}) if isinstance(full_execution_data, dict) else {},
                    'pod_metrics': full_execution_data.get('pod_metrics', []) if isinstance(full_execution_data, dict) else [],
                    'anomaly_summary': full_execution_data.get('anomaly_summary', {}) if isinstance(full_execution_data, dict) else {}
                }
                report_data = status
            except Exception as e:
                import traceback
                return jsonify({'success': False, 'error': f'Error constructing status: {str(e)}. Traceback: {traceback.format_exc()}'}), 500
        
        # Generate HTML report
        from jinja2 import Template
        import os
        
        template_path = os.path.join(os.path.dirname(__file__), 'templates', 'smart_execution_report.html')
        
        if not os.path.exists(template_path):
            # Create template on the fly
            html_template = '''<!DOCTYPE html>
<html>
<head>
    <title>Smart Execution Report - {{ execution_id }}</title>
    <meta charset="UTF-8">
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        h1 { color: #667eea; border-bottom: 3px solid #667eea; padding-bottom: 10px; }
        h2 { color: #764ba2; margin-top: 30px; }
        .metric-card { display: inline-block; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; margin: 10px; border-radius: 8px; min-width: 200px; }
        .metric-label { font-size: 14px; opacity: 0.9; }
        .metric-value { font-size: 32px; font-weight: bold; margin-top: 5px; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th { background: #667eea; color: white; padding: 12px; text-align: left; }
        td { padding: 10px; border-bottom: 1px solid #ddd; }
        tr:hover { background: #f5f5f5; }
        .status-badge { padding: 5px 10px; border-radius: 4px; color: white; font-weight: bold; }
        .status-success { background: #28a745; }
        .status-failed { background: #dc3545; }
        .status-running { background: #17a2b8; }
        .chart { margin: 20px 0; }
    </style>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
</head>
<body>
    <div class="container">
        <h1>🚀 Smart Execution Report</h1>
        
        <div style="color: #666; margin-bottom: 30px;">
            <strong>Execution ID:</strong> {{ execution_id }}<br>
            <strong>Testbed:</strong> {{ testbed_label }}<br>
            <strong>Status:</strong> <span class="status-badge status-{{ status|lower }}">{{ status }}</span><br>
            <strong>Generated:</strong> {{ timestamp }}
        </div>
        
        <h2>📊 Executive Summary</h2>
        <div>
            <div class="metric-card">
                <div class="metric-label">Total Operations</div>
                <div class="metric-value">{{ total_operations }}</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">Duration</div>
                <div class="metric-value">{{ duration_minutes|round(1) }}m</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">Operations/Min</div>
                <div class="metric-value">{{ ops_per_minute|round(1) }}</div>
            </div>
            <div class="metric-card">
                <div class="metric-label">Success Rate</div>
                <div class="metric-value">{{ success_rate|round(1) }}%</div>
            </div>
        </div>
        
        <h2>🎯 Target Configuration</h2>
        <table>
            <tr>
                <td><strong>Target CPU Threshold:</strong></td>
                <td>{{ target_config.cpu_threshold }}%</td>
            </tr>
            <tr>
                <td><strong>Target Memory Threshold:</strong></td>
                <td>{{ target_config.memory_threshold }}%</td>
            </tr>
            <tr>
                <td><strong>Stop Condition:</strong></td>
                <td>{{ target_config.stop_condition|upper }}</td>
            </tr>
        </table>
        
        <h2>📈 Metrics Summary</h2>
        <table>
            <tr>
                <th>Metric</th>
                <th>Baseline</th>
                <th>Final</th>
                <th>Change</th>
            </tr>
            <tr>
                <td><strong>CPU Usage</strong></td>
                <td>{{ baseline_metrics.get('cpu_percent', 0)|round(1) }}%</td>
                <td>{{ final_metrics.get('cpu_percent', 0)|round(1) }}%</td>
                <td>{{ (final_metrics.get('cpu_percent', 0) - baseline_metrics.get('cpu_percent', 0))|round(1) }}%</td>
            </tr>
            <tr>
                <td><strong>Memory Usage</strong></td>
                <td>{{ baseline_metrics.get('memory_percent', 0)|round(1) }}%</td>
                <td>{{ final_metrics.get('memory_percent', 0)|round(1) }}%</td>
                <td>{{ (final_metrics.get('memory_percent', 0) - baseline_metrics.get('memory_percent', 0))|round(1) }}%</td>
            </tr>
            {% if network_metrics %}
            <tr>
                <td><strong>Network RX (MB/s)</strong></td>
                <td>{{ baseline_metrics.get('network', {}).get('rx_mbps', 0)|round(2) }}</td>
                <td>{{ network_metrics.get('rx_mbps', 0)|round(2) }}</td>
                <td>{{ (network_metrics.get('rx_mbps', 0) - baseline_metrics.get('network', {}).get('rx_mbps', 0))|round(2) }}</td>
            </tr>
            <tr>
                <td><strong>Network TX (MB/s)</strong></td>
                <td>{{ baseline_metrics.get('network', {}).get('tx_mbps', 0)|round(2) }}</td>
                <td>{{ network_metrics.get('tx_mbps', 0)|round(2) }}</td>
                <td>{{ (network_metrics.get('tx_mbps', 0) - baseline_metrics.get('network', {}).get('tx_mbps', 0))|round(2) }}</td>
            </tr>
            {% endif %}
            {% if disk_metrics %}
            <tr>
                <td><strong>Disk Usage (%)</strong></td>
                <td>{{ baseline_metrics.get('disk', {}).get('usage_percent', 0)|round(1) }}%</td>
                <td>{{ disk_metrics.get('usage_percent', 0)|round(1) }}%</td>
                <td>{{ (disk_metrics.get('usage_percent', 0) - baseline_metrics.get('disk', {}).get('usage_percent', 0))|round(1) }}%</td>
            </tr>
            {% endif %}
            {% if latency_metrics %}
            <tr>
                <td><strong>P95 Latency (ms)</strong></td>
                <td>{{ baseline_metrics.get('latency', {}).get('p95_ms', 0)|round(2) }}</td>
                <td>{{ latency_metrics.get('p95_ms', 0)|round(2) }}</td>
                <td>{{ (latency_metrics.get('p95_ms', 0) - baseline_metrics.get('latency', {}).get('p95_ms', 0))|round(2) }}</td>
            </tr>
            {% endif %}
        </table>
        
        {% if threshold_reached %}
        <div style="background: #d4edda; color: #155724; padding: 15px; border-radius: 5px; margin: 20px 0;">
            <strong>✅ Threshold Reached!</strong> Target metrics have been achieved.
        </div>
        {% endif %}
        
        {% if predictions %}
        <h2>🔮 Predictive Insights</h2>
        <table>
            <tr>
                <td><strong>Estimated Operations Remaining:</strong></td>
                <td>{{ predictions.get('estimated_operations_remaining', 'N/A') }}</td>
            </tr>
            <tr>
                <td><strong>Estimated Time to Completion:</strong></td>
                <td>{{ predictions.get('estimated_time_minutes', 'N/A') }} minutes</td>
            </tr>
            <tr>
                <td><strong>Current Trend:</strong></td>
                <td>{{ predictions.get('current_trend', 'unknown')|upper }}</td>
            </tr>
            <tr>
                <td><strong>Efficiency Score:</strong></td>
                <td>{{ predictions.get('efficiency_score', 0)|round(1) }}/10</td>
            </tr>
            <tr>
                <td><strong>Bottleneck:</strong></td>
                <td>{{ predictions.get('bottleneck', 'unknown')|upper }}</td>
            </tr>
        </table>
        {% endif %}
        
        {% if anomalies %}
        <h2>⚠️ Detected Anomalies</h2>
        <table>
            <thead>
                <tr>
                    <th>Type</th>
                    <th>Severity</th>
                    <th>Description</th>
                    <th>Timestamp</th>
                </tr>
            </thead>
            <tbody>
                {% for anomaly in anomalies[-10:] %}
                <tr>
                    <td>{{ anomaly.get('type', 'Unknown') }}</td>
                    <td><span class="status-badge status-{{ anomaly.get('severity', 'low') }}">{{ anomaly.get('severity', 'low')|upper }}</span></td>
                    <td>{{ anomaly.get('description', 'N/A') }}</td>
                    <td>{{ anomaly.get('timestamp', 'N/A') }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        {% endif %}
        
        {% if recommendations %}
        <h2>💡 Automated Recommendations</h2>
        <ul>
            {% for rec in recommendations[-5:] %}
            <li style="margin: 10px 0;">
                <strong>{{ rec.get('type', 'Recommendation') }}:</strong> {{ rec.get('action', 'N/A') }}
                <br><small style="color: #666;">{{ rec.get('reason', '') }}</small>
            </li>
            {% endfor %}
        </ul>
        {% endif %}
        
        {% if operation_effectiveness %}
        <h2>📊 Most Effective Operations</h2>
        <table>
            <thead>
                <tr>
                    <th>Entity</th>
                    <th>Operation</th>
                    <th>Impact Score</th>
                    <th>Avg CPU Change</th>
                    <th>Avg Memory Change</th>
                </tr>
            </thead>
            <tbody>
                {% for eff in operation_effectiveness %}
                <tr>
                    <td>{{ eff.get('entity_type', 'N/A') }}</td>
                    <td>{{ eff.get('operation', 'N/A') }}</td>
                    <td>{{ eff.get('impact_score', 0)|round(2) }}</td>
                    <td>{{ eff.get('avg_cpu_change', 0)|round(2) }}%</td>
                    <td>{{ eff.get('avg_memory_change', 0)|round(2) }}%</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        {% endif %}
        
        {% if entity_breakdown %}
        <h2>📋 Entity Breakdown</h2>
        <table>
            <thead>
                <tr>
                    <th>Entity Type</th>
                    <th>Total</th>
                    <th>Success</th>
                    <th>Failed</th>
                    <th>Success Rate</th>
                </tr>
            </thead>
            <tbody>
                {% for entity_type, stats in entity_breakdown.items() %}
                <tr>
                    <td><strong>{{ entity_type }}</strong></td>
                    <td>{{ stats.get('total', 0) }}</td>
                    <td>{{ stats.get('success', 0) }}</td>
                    <td>{{ stats.get('failed', 0) }}</td>
                    <td>{{ ((stats.get('success', 0) / stats.get('total', 1)) * 100)|round(1) }}%</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        {% endif %}
        
        <h2>🔨 Operation Details (All {{ operations_history|length }} Operations)</h2>
        <table>
            <thead>
                <tr>
                    <th>#</th>
                    <th>Entity</th>
                    <th>Operation</th>
                    <th>Name</th>
                    <th>Duration</th>
                    <th>Status</th>
                </tr>
            </thead>
            <tbody>
                {% for op in operations_history %}
                <tr>
                    <td>{{ loop.index }}</td>
                    <td>{{ op.get('entity_type', 'N/A') }}</td>
                    <td>{{ op.get('operation', 'N/A') }}</td>
                    <td><code>{{ op.get('entity_name', 'N/A') }}</code></td>
                    <td>{{ op.get('duration_seconds', 0)|round(2) }}s</td>
                    <td>
                        <span class="status-badge status-{{ op.get('status', 'UNKNOWN')|lower }}">
                            {{ op.get('status', 'UNKNOWN') }}
                        </span>
                    </td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        
        <div style="margin-top: 50px; padding-top: 20px; border-top: 1px solid #ddd; color: #999; font-size: 12px;">
            Generated by NMT Smart Execution Tool • {{ timestamp }}
        
        {% if pod_operation_correlation and pod_operation_correlation.get('operations') %}
        <h2>🔗 Pod-Operation Correlation</h2>
        <p style="color: #666; margin-bottom: 20px;">This section shows which pods were affected by each operation and their resource changes.</p>
        {% for op in pod_operation_correlation.get('operations', []) %}
        <h3 style="color: #667eea; margin-top: 30px;">{{ op.get('entity_type') }}.{{ op.get('operation_type') }} - {{ op.get('entity_name') }}</h3>
        <table>
            <thead>
                <tr>
                    <th>Pod Name</th>
                    <th>Namespace</th>
                    <th>Node</th>
                    <th>CPU Before</th>
                    <th>CPU After</th>
                    <th>CPU Δ</th>
                    <th>Memory Before</th>
                    <th>Memory After</th>
                    <th>Memory Δ</th>
                    <th>Impact Score</th>
                </tr>
            </thead>
            <tbody>
                {% for pod in op.get('pods', []) %}
                <tr>
                    <td><code>{{ pod.get('pod_name') }}</code></td>
                    <td>{{ pod.get('namespace') }}</td>
                    <td>{{ pod.get('node_name', 'N/A') }}</td>
                    <td>{{ pod.get('cpu_before', 0)|round(2) }}%</td>
                    <td>{{ pod.get('cpu_after', 0)|round(2) }}%</td>
                    <td style="color: {% if pod.get('cpu_delta', 0) > 0 %}#dc3545{% else %}#28a745{% endif %}; font-weight: bold;">
                        {{ pod.get('cpu_delta', 0)|round(2) }}%
                    </td>
                    <td>{{ pod.get('memory_before', 0)|round(2) }}MB</td>
                    <td>{{ pod.get('memory_after', 0)|round(2) }}MB</td>
                    <td style="color: {% if pod.get('memory_delta', 0) > 0 %}#dc3545{% else %}#28a745{% endif %}; font-weight: bold;">
                        {{ pod.get('memory_delta', 0)|round(2) }}MB
                    </td>
                    <td>{{ pod.get('impact_score', 0)|round(2) }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        {% endfor %}
        <div style="background: #e7f3ff; padding: 15px; border-radius: 5px; margin: 20px 0;">
            <strong>Summary:</strong> {{ pod_operation_correlation.get('summary', {}).get('total_pods_affected', 0) }} pods affected across {{ pod_operation_correlation.get('summary', {}).get('total_operations', 0) }} operations
        </div>
        {% endif %}
        
        </div>
    </div>
</body>
</html>'''
            template = Template(html_template)
        else:
            with open(template_path, 'r') as f:
                template = Template(f.read())
        
        # Prepare data (safely handle None values)
        # Ensure status is a dict - CRITICAL CHECK
        if status is None:
            logging.error(f"CRITICAL: status is None before template rendering")
            return jsonify({'success': False, 'error': 'Execution status is None'}), 500
        
        if not isinstance(status, dict):
            logging.error(f"status is not a dict: {type(status)}")
            status = {}
        
        if report_data is None:
            logging.warning(f"report_data is None, using status as report_data")
            report_data = status
        
        if not isinstance(report_data, dict):
            logging.error(f"report_data is not a dict: {type(report_data)}")
            report_data = {}
        
        total_ops = status.get('total_operations', 0) or 0
        duration_mins = status.get('duration_minutes', 0) or 0.1
        
        ops_per_minute = status.get('operations_per_minute', 0) or (total_ops / max(duration_mins, 0.1))
        
        operations_history_list = status.get('operations_history') or []
        if not isinstance(operations_history_list, list):
            operations_history_list = []
        
        successful_ops = status.get('successful_operations', 0) or sum(1 for op in operations_history_list if isinstance(op, dict) and op.get('status') == 'SUCCESS')
        success_rate = status.get('success_rate', 0) or (successful_ops / max(total_ops, 1)) * 100
        
        # Get all operations history (not just last 20)
        if isinstance(operations_history_list, list) and len(operations_history_list) > 0:
            operations_history = operations_history_list
        else:
            operations_history = []
        
        # Get enhanced metrics (safely handle None)
        final_metrics = status.get('current_metrics') or status.get('final_metrics') or {}
        if not isinstance(final_metrics, dict):
            final_metrics = {}
        
        baseline_metrics = status.get('baseline_metrics') or {}
        if not isinstance(baseline_metrics, dict):
            baseline_metrics = {}
        
        # Get Phase 1-3 data (safely handle None values)
        predictions = status.get('predictions') or report_data.get('predictions') if isinstance(report_data, dict) else None
        if predictions is None:
            predictions = {}
        elif not isinstance(predictions, dict):
            predictions = {}
        
        anomalies = status.get('detected_anomalies') or (report_data.get('detected_anomalies') if isinstance(report_data, dict) else [])
        if not isinstance(anomalies, list):
            anomalies = []
        
        recommendations = status.get('recommendations') or (report_data.get('recommendations') if isinstance(report_data, dict) else [])
        if not isinstance(recommendations, list):
            recommendations = []
        
        operation_effectiveness = status.get('operation_effectiveness') or (report_data.get('operation_effectiveness') if isinstance(report_data, dict) else [])
        if not isinstance(operation_effectiveness, list):
            operation_effectiveness = []
        
        entity_breakdown = status.get('entity_breakdown') or (report_data.get('entity_breakdown') if isinstance(report_data, dict) else {})
        if not isinstance(entity_breakdown, dict):
            entity_breakdown = {}
        
        # Extract pod correlation data
        pod_operation_correlation = status.get('pod_operation_correlation') or (report_data.get('pod_operation_correlation') if isinstance(report_data, dict) else {})
        if not isinstance(pod_operation_correlation, dict):
            pod_operation_correlation = {}
        
        # Safely get testbed_label - ensure status is valid first
        if status is None:
            logging.error("CRITICAL: status is None when trying to get testbed_info")
            return jsonify({'success': False, 'error': 'Execution status is None'}), 500
        
        if not isinstance(status, dict):
            logging.error("CRITICAL: status is not a dict when trying to get testbed_info")
            status = {}
        
        testbed_info = status.get('testbed_info') if isinstance(status, dict) else {}
        if not isinstance(testbed_info, dict):
            testbed_info = {}
        testbed_label = testbed_info.get('testbed_label') or (report_data.get('testbed') if isinstance(report_data, dict) else 'Unknown') or 'Unknown'
        
        # Ensure final_metrics is a dict before accessing nested keys
        if not isinstance(final_metrics, dict):
            final_metrics = {}
        
        # Safely extract nested metrics
        network_metrics = final_metrics.get('network', {}) if isinstance(final_metrics, dict) else {}
        if not isinstance(network_metrics, dict):
            network_metrics = {}
        
        disk_metrics = final_metrics.get('disk', {}) if isinstance(final_metrics, dict) else {}
        if not isinstance(disk_metrics, dict):
            disk_metrics = {}
        
        latency_metrics = final_metrics.get('latency', {}) if isinstance(final_metrics, dict) else {}
        if not isinstance(latency_metrics, dict):
            latency_metrics = {}
        
        # Ensure baseline_metrics is a dict
        if not isinstance(baseline_metrics, dict):
            baseline_metrics = {}
        
        # Ensure operations_history is a list
        if not isinstance(operations_history, list):
            operations_history = []
        
        # Ensure target_config is a dict
        target_config = status.get('target_config') or {}
        if not isinstance(target_config, dict):
            target_config = {}
        
        logging.info(f"DEBUG: Template variables prepared - final_metrics type: {type(final_metrics)}, baseline_metrics type: {type(baseline_metrics)}")
        
        html_content = template.render(
            execution_id=execution_id,
            testbed_label=testbed_label,
            status=status.get('status', 'UNKNOWN') if isinstance(status, dict) else 'UNKNOWN',
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            total_operations=total_ops,
            duration_minutes=duration_mins,
            ops_per_minute=ops_per_minute,
            success_rate=success_rate,
            target_config=target_config,
            baseline_metrics=baseline_metrics,
            final_metrics=final_metrics,
            operations_history=operations_history,
            predictions=predictions,
            anomalies=anomalies,
            recommendations=recommendations,
            operation_effectiveness=operation_effectiveness,
            entity_breakdown=entity_breakdown,
            threshold_reached=status.get('threshold_reached', False) if isinstance(status, dict) else False,
            network_metrics=network_metrics,
            disk_metrics=disk_metrics,
            latency_metrics=latency_metrics,
            pod_operation_correlation=pod_operation_correlation
        )
        
        from flask import make_response
        response = make_response(html_content)
        response.headers['Content-Type'] = 'text/html'
        response.headers['Content-Disposition'] = f'attachment; filename=smart-execution-{execution_id[:10]}.html'
        
        return response
        
    except Exception as e:
        logging.exception("Error generating HTML report")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/report/<execution_id>/enhanced', methods=['GET'])
def get_enhanced_smart_execution_report(execution_id):
    """
    Generate AI-enhanced HTML report with spike analysis, cluster health,
    failure grouping, capacity planning, and historical comparison.
    """
    try:
        from services.smart_execution_service import get_smart_execution
        from services.smart_execution_db import load_smart_execution

        status = None
        report_data = None
        prometheus_url = None

        controller = get_smart_execution(execution_id)
        if controller:
            status = controller.get_status()
            report_data = controller.get_report()
            prometheus_url = getattr(controller, 'prometheus_url', None)
        else:
            db_data = None
            try:
                db_data = load_smart_execution(execution_id)
            except Exception as e:
                return jsonify({'success': False, 'error': f'Error loading execution: {str(e)}'}), 500

            if db_data is None:
                return jsonify({'success': False, 'error': 'Execution not found'}), 404

            if not isinstance(db_data, dict):
                return jsonify({'success': False, 'error': 'Invalid execution data format'}), 500

            full_execution_data = db_data.get('full_execution_data')
            if full_execution_data is None or not isinstance(full_execution_data, dict):
                full_execution_data = {}

            status = {
                'execution_id': db_data.get('execution_id') or execution_id,
                'status': db_data.get('status') or 'UNKNOWN',
                'start_time': db_data.get('start_time'),
                'end_time': db_data.get('end_time'),
                'duration_minutes': db_data.get('duration_minutes') or 0,
                'total_operations': db_data.get('total_operations') or 0,
                'successful_operations': db_data.get('successful_operations') or 0,
                'failed_operations': db_data.get('failed_operations') or 0,
                'success_rate': db_data.get('success_rate') or 0,
                'operations_per_minute': db_data.get('operations_per_minute') or 0,
                'target_config': db_data.get('target_config') or {},
                'baseline_metrics': db_data.get('baseline_metrics') or {},
                'current_metrics': db_data.get('final_metrics') or {},
                'operations_history': db_data.get('operations_history') or [],
                'metrics_history': db_data.get('metrics_history') or [],
                'threshold_reached': db_data.get('threshold_reached') or False,
                'entity_breakdown': db_data.get('entity_breakdown') or {},
                'testbed_info': {
                    'testbed_label': db_data.get('testbed_label') or 'Unknown',
                    'testbed_id': db_data.get('testbed_id') or 'unknown'
                },
                'detected_anomalies': full_execution_data.get('detected_anomalies', []),
                'operation_effectiveness': full_execution_data.get('operation_effectiveness', []),
                'pod_operation_correlation': full_execution_data.get('pod_operation_correlation', {}),
            }
            report_data = status

            testbed_ip = db_data.get('testbed_id', '')
            if testbed_ip:
                prometheus_url = f'http://{testbed_ip}:9090'

        if report_data is None:
            report_data = status or {}
        if status is None:
            status = report_data

        if not isinstance(report_data, dict):
            report_data = {}
        if not isinstance(status, dict):
            status = {}

        # Generate enhanced report data
        from services.enhanced_report_service import EnhancedReportService
        enhanced_svc = EnhancedReportService(prometheus_url=prometheus_url)

        testbed_id = None
        testbed_info = status.get('testbed_info') or {}
        if isinstance(testbed_info, dict):
            testbed_id = testbed_info.get('testbed_id')

        enhanced_data = enhanced_svc.generate_enhanced_report(
            report_data=report_data,
            status_data=status,
            execution_id=execution_id,
            testbed_id=testbed_id
        )

        # Prepare template variables
        import json as json_mod
        from jinja2 import Template

        total_ops = status.get('total_operations', 0) or 0
        duration_mins = status.get('duration_minutes', 0) or 0.1
        ops_per_minute = status.get('operations_per_minute', 0) or (total_ops / max(duration_mins, 0.1))
        successful_ops = status.get('successful_operations', 0)
        success_rate = status.get('success_rate', 0) or ((successful_ops / max(total_ops, 1)) * 100)

        operations_history = status.get('operations_history') or []
        if not isinstance(operations_history, list):
            operations_history = []
        metrics_history = status.get('metrics_history') or []
        if not isinstance(metrics_history, list):
            metrics_history = []

        final_metrics = status.get('current_metrics') or status.get('final_metrics') or {}
        baseline_metrics = status.get('baseline_metrics') or {}
        target_config = status.get('target_config') or {}
        entity_breakdown = status.get('entity_breakdown') or report_data.get('entity_breakdown') or {}
        operation_effectiveness = status.get('operation_effectiveness') or report_data.get('operation_effectiveness') or []

        testbed_label = 'Unknown'
        if isinstance(testbed_info, dict):
            testbed_label = testbed_info.get('testbed_label', 'Unknown')
        if testbed_label == 'Unknown':
            testbed_label = report_data.get('testbed', 'Unknown') if isinstance(report_data, dict) else 'Unknown'

        template_path = os.path.join(os.path.dirname(__file__), 'templates', 'enhanced_report.html')
        with open(template_path, 'r') as f:
            template = Template(f.read())

        html_content = template.render(
            execution_id=execution_id,
            testbed_label=testbed_label,
            status=status.get('status', 'UNKNOWN'),
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            total_operations=total_ops,
            duration_minutes=duration_mins,
            ops_per_minute=ops_per_minute,
            success_rate=success_rate,
            target_config=target_config,
            baseline_metrics=baseline_metrics,
            final_metrics=final_metrics,
            operations_history=operations_history,
            entity_breakdown=entity_breakdown,
            operation_effectiveness=operation_effectiveness if isinstance(operation_effectiveness, list) else [],
            threshold_reached=status.get('threshold_reached', False),
            # Enhanced data
            verdict=enhanced_data['verdict'],
            spike_analysis=enhanced_data['spike_analysis'],
            failure_analysis=enhanced_data['failure_analysis'],
            operation_heatmap=enhanced_data['operation_heatmap'],
            pod_stability=enhanced_data['pod_stability'],
            cluster_health=enhanced_data['cluster_health'],
            capacity_planning=enhanced_data['capacity_planning'],
            historical_comparison=enhanced_data['historical_comparison'],
            # JSON for charts
            metrics_history_json=json_mod.dumps(metrics_history),
            operations_history_json=json_mod.dumps(operations_history),
        )

        fmt = request.args.get('format', 'download')
        if fmt == 'json':
            return jsonify({
                'success': True,
                'enhanced_report': enhanced_data,
                'execution_id': execution_id,
                'testbed_label': testbed_label,
                'status': status.get('status', 'UNKNOWN'),
            })

        from flask import make_response
        response = make_response(html_content)
        response.headers['Content-Type'] = 'text/html'
        response.headers['Content-Disposition'] = f'attachment; filename=smart-execution-enhanced-{execution_id[:10]}.html'

        return response

    except Exception as e:
        logging.exception("Error generating enhanced report")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/<execution_id>', methods=['DELETE'])
def delete_smart_execution(execution_id):
    """
    Delete a smart execution record
    """
    try:
        from services.smart_execution_service import delete_smart_execution as delete_exec
        
        success = delete_exec(execution_id)
        if success:
            return jsonify({'success': True, 'message': 'Execution deleted'}), 200
        else:
            return jsonify({'success': False, 'error': 'Execution not found'}), 404
            
    except Exception as e:
        logging.exception("Error deleting execution")
        return jsonify({'success': False, 'error': str(e)}), 500


# Export route moved before report route to avoid route conflicts
    """
    Phase 3: Export smart execution report in various formats
    
    Formats: csv, json, pdf
    """
    try:
        from services.smart_execution_service import get_smart_execution
        
        controller = get_smart_execution(execution_id)
        if not controller:
            return jsonify({'success': False, 'error': 'Execution not found'}), 404
        
        report = controller.get_report()
        status = controller.get_status()
        
        if format.lower() == 'csv':
            import csv
            import io
            from flask import make_response
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow(['Field', 'Value'])
            
            # Write basic info
            writer.writerow(['Execution ID', report.get('execution_id')])
            writer.writerow(['Testbed', report.get('testbed')])
            writer.writerow(['Status', report.get('status')])
            writer.writerow(['Duration (minutes)', report.get('duration_minutes')])
            writer.writerow(['Total Operations', report.get('total_operations')])
            writer.writerow(['Success Rate (%)', report.get('success_rate')])
            
            # Write operations
            writer.writerow([])
            writer.writerow(['Operations'])
            writer.writerow(['#', 'Entity', 'Operation', 'Status', 'Duration (s)', 'Timestamp'])
            for i, op in enumerate(report.get('operations_history', []), 1):
                writer.writerow([
                    i,
                    op.get('entity_type', ''),
                    op.get('operation', ''),
                    op.get('status', ''),
                    op.get('duration_seconds', 0),
                    op.get('start_time', '')
                ])
            
            # Write metrics
            writer.writerow([])
            writer.writerow(['Metrics Timeline'])
            writer.writerow(['Timestamp', 'CPU %', 'Memory %'])
            for metric in report.get('metrics_history', []):
                writer.writerow([
                    metric.get('timestamp', ''),
                    metric.get('cpu_percent', 0),
                    metric.get('memory_percent', 0)
                ])
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv'
            response.headers['Content-Disposition'] = f'attachment; filename=smart-execution-{execution_id[:10]}.csv'
            return response
        
        elif format.lower() == 'json':
            from flask import make_response
            import json as json_lib
            
            response = make_response(json_lib.dumps(report, indent=2, default=str))
            response.headers['Content-Type'] = 'application/json'
            response.headers['Content-Disposition'] = f'attachment; filename=smart-execution-{execution_id[:10]}.json'
            return response
        
        elif format.lower() == 'pdf':
            # PDF export would require reportlab or similar
            # For now, return JSON with note that PDF requires additional library
            return jsonify({
                'success': False,
                'error': 'PDF export requires reportlab library. Use CSV or JSON export instead.',
                'available_formats': ['csv', 'json']
            }), 501
        
        else:
            return jsonify({'success': False, 'error': f'Unsupported format: {format}'}), 400
        
    except Exception as e:
        logging.exception("Error exporting smart execution report")
        return jsonify({'success': False, 'error': str(e)}), 500

# ===========================
# Phase 3: Export APIs (MUST BE BEFORE REPORT ROUTE)
# ===========================

@app.route('/api/smart-execution/report/<execution_id>/export/<format>', methods=['GET'])
def export_smart_execution_report(execution_id, format):
    """
    Phase 3: Export smart execution report in various formats
    
    Formats: csv, json, excel, pdf
    """
    try:
        from services.smart_execution_service import get_smart_execution
        
        controller = get_smart_execution(execution_id)
        if not controller:
            return jsonify({'success': False, 'error': 'Execution not found'}), 404
        
        report = controller.get_report()
        status = controller.get_status()
        
        if format.lower() == 'csv':
            import csv
            import io
            from flask import make_response
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow(['Field', 'Value'])
            
            # Write basic info
            writer.writerow(['Execution ID', report.get('execution_id')])
            writer.writerow(['Testbed', report.get('testbed')])
            writer.writerow(['Status', report.get('status')])
            writer.writerow(['Duration (minutes)', report.get('duration_minutes')])
            writer.writerow(['Total Operations', report.get('total_operations')])
            writer.writerow(['Success Rate (%)', report.get('success_rate')])
            
            # Write operations
            writer.writerow([])
            writer.writerow(['Operations'])
            writer.writerow(['#', 'Entity', 'Operation', 'Status', 'Duration (s)', 'Timestamp'])
            for i, op in enumerate(report.get('operations_history', []), 1):
                writer.writerow([
                    i,
                    op.get('entity_type', ''),
                    op.get('operation', ''),
                    op.get('status', ''),
                    op.get('duration_seconds', 0),
                    op.get('start_time', '')
                ])
            
            # Write metrics
            writer.writerow([])
            writer.writerow(['Metrics Timeline'])
            writer.writerow(['Timestamp', 'CPU %', 'Memory %'])
            for metric in report.get('metrics_history', []):
                writer.writerow([
                    metric.get('timestamp', ''),
                    metric.get('cpu_percent', 0),
                    metric.get('memory_percent', 0)
                ])
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'text/csv'
            response.headers['Content-Disposition'] = f'attachment; filename=smart-execution-{execution_id[:10]}.csv'
            return response
        
        elif format.lower() == 'json':
            from flask import make_response
            import json as json_lib
            
            response = make_response(json_lib.dumps(report, indent=2, default=str))
            response.headers['Content-Type'] = 'application/json'
            response.headers['Content-Disposition'] = f'attachment; filename=smart-execution-{execution_id[:10]}.json'
            return response
        
        elif format.lower() == 'excel' or format.lower() == 'xlsx':
            # Excel export - query database for pod-level data
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                from database import SessionLocal
                from sqlalchemy import text
                import json as json_lib
                from flask import make_response
                import io
                
                # Create workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Pod Metrics"
                
                # Headers matching snapshot file format
                headers = [
                    "Execution ID", "Operation #", "Entity Type", "Operation", "Status",
                    "Pod Name", "Namespace", "Node", "CPU Before (%)", "CPU After (%)", 
                    "CPU Delta (%)", "Memory Before (MB)", "Memory After (MB)", 
                    "Memory Delta (MB)", "Network RX Before (Mbps)", "Network RX After (Mbps)",
                    "Network TX Before (Mbps)", "Network TX After (Mbps)", "Timestamp", "Duration (s)"
                ]
                
                # Style headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Query database for pod metrics
                session = SessionLocal()
                row = 2
                op_idx = 0
                
                try:
                    # Get pod_operation_correlation (most complete data)
                    pod_query = text("""
                        SELECT 
                            entity_type, operation_type, pod_name, namespace, node_name,
                            cpu_percent_before, cpu_percent_after, cpu_delta,
                            memory_mb_before, memory_mb_after, memory_delta,
                            network_rx_mbps_before, network_rx_mbps_after, network_rx_delta,
                            network_tx_mbps_before, network_tx_mbps_after, network_tx_delta,
                            measured_at
                        FROM pod_operation_correlation
                        WHERE execution_id = :execution_id
                        ORDER BY measured_at ASC
                    """)
                    
                    pod_result = session.execute(pod_query, {'execution_id': execution_id})
                    pod_correlations = pod_result.fetchall()
                    
                    # Process pod correlations
                    for pod_row in pod_correlations:
                        op_idx += 1
                        ws.cell(row=row, column=1, value=execution_id)
                        ws.cell(row=row, column=2, value=op_idx)
                        ws.cell(row=row, column=3, value=pod_row[0] or '')
                        ws.cell(row=row, column=4, value=pod_row[1] or '')
                        ws.cell(row=row, column=5, value='COMPLETED')
                        ws.cell(row=row, column=6, value=pod_row[2] or '')
                        ws.cell(row=row, column=7, value=pod_row[3] or '')
                        ws.cell(row=row, column=8, value=pod_row[4] or '')
                        ws.cell(row=row, column=9, value=float(pod_row[5]) if pod_row[5] is not None else 0)
                        ws.cell(row=row, column=10, value=float(pod_row[6]) if pod_row[6] is not None else 0)
                        ws.cell(row=row, column=11, value=float(pod_row[7]) if pod_row[7] is not None else 0)
                        ws.cell(row=row, column=12, value=float(pod_row[8]) if pod_row[8] is not None else 0)
                        ws.cell(row=row, column=13, value=float(pod_row[9]) if pod_row[9] is not None else 0)
                        ws.cell(row=row, column=14, value=float(pod_row[10]) if pod_row[10] is not None else 0)
                        ws.cell(row=row, column=15, value=float(pod_row[11]) if pod_row[11] is not None else 0)
                        ws.cell(row=row, column=16, value=float(pod_row[12]) if pod_row[12] is not None else 0)
                        ws.cell(row=row, column=17, value=float(pod_row[13]) if pod_row[13] is not None else 0)
                        ws.cell(row=row, column=18, value=float(pod_row[14]) if pod_row[14] is not None else 0)
                        ws.cell(row=row, column=19, value=pod_row[16].isoformat() if pod_row[16] else '')
                        ws.cell(row=row, column=20, value=0)
                        row += 1
                    
                    # Fallback: If no pod correlations, use operation_metrics
                    if row == 2:
                        query = text("""
                            SELECT 
                                entity_type, operation_type, entity_name,
                                started_at, completed_at, duration_seconds, status,
                                pod_metrics_before, pod_metrics_after
                            FROM operation_metrics
                            WHERE execution_id = :execution_id
                            ORDER BY started_at ASC
                        """)
                        
                        result = session.execute(query, {'execution_id': execution_id})
                        db_metrics = result.fetchall()
                        
                        for db_metric in db_metrics:
                            op_idx += 1
                            pods_before = {}
                            pods_after = {}
                            
                            if db_metric[7]:  # pod_metrics_before
                                try:
                                    if isinstance(db_metric[7], str):
                                        pods_before = json_lib.loads(db_metric[7])
                                    else:
                                        pods_before = db_metric[7]
                                except:
                                    pass
                            
                            if db_metric[8]:  # pod_metrics_after
                                try:
                                    if isinstance(db_metric[8], str):
                                        pods_after = json_lib.loads(db_metric[8])
                                    else:
                                        pods_after = db_metric[8]
                                except:
                                    pass
                            
                            if pods_before or pods_after:
                                all_pods = set(list(pods_before.keys()) + list(pods_after.keys()))
                                for pod_name in all_pods:
                                    pod_before = pods_before.get(pod_name, {})
                                    pod_after = pods_after.get(pod_name, {})
                                    
                                    ws.cell(row=row, column=1, value=execution_id)
                                    ws.cell(row=row, column=2, value=op_idx)
                                    ws.cell(row=row, column=3, value=db_metric[0] or '')
                                    ws.cell(row=row, column=4, value=db_metric[1] or '')
                                    ws.cell(row=row, column=5, value=db_metric[6] or '')
                                    ws.cell(row=row, column=6, value=pod_name)
                                    ws.cell(row=row, column=7, value=pod_before.get('namespace') or pod_after.get('namespace', ''))
                                    ws.cell(row=row, column=8, value=pod_before.get('node') or pod_after.get('node', ''))
                                    ws.cell(row=row, column=9, value=pod_before.get('cpu_usage', 0))
                                    ws.cell(row=row, column=10, value=pod_after.get('cpu_usage', 0))
                                    ws.cell(row=row, column=11, value=(pod_after.get('cpu_usage', 0) - pod_before.get('cpu_usage', 0)))
                                    ws.cell(row=row, column=12, value=pod_before.get('memory_mb', 0))
                                    ws.cell(row=row, column=13, value=pod_after.get('memory_mb', 0))
                                    ws.cell(row=row, column=14, value=(pod_after.get('memory_mb', 0) - pod_before.get('memory_mb', 0)))
                                    ws.cell(row=row, column=15, value=pod_before.get('network_rx_mbps', 0))
                                    ws.cell(row=row, column=16, value=pod_after.get('network_rx_mbps', 0))
                                    ws.cell(row=row, column=17, value=pod_before.get('network_tx_mbps', 0))
                                    ws.cell(row=row, column=18, value=pod_after.get('network_tx_mbps', 0))
                                    ws.cell(row=row, column=19, value=db_metric[3].isoformat() if db_metric[3] else '')
                                    ws.cell(row=row, column=20, value=float(db_metric[5]) if db_metric[5] else 0)
                                    row += 1
                
                finally:
                    session.close()
                
                # Auto-adjust column widths
                for col_idx in range(1, len(headers) + 1):
                    max_length = 0
                    column = get_column_letter(col_idx)
                    for cell in ws[column]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
                
                # Save to BytesIO
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                response = make_response(output.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = f'attachment; filename=snapshot_individual_pods_{execution_id[:20]}.xlsx'
                return response
                
            except ImportError:
                return jsonify({
                    'success': False,
                    'error': 'Excel export requires openpyxl library. Install with: pip install openpyxl',
                    'available_formats': ['csv', 'json', 'excel']
                }), 501
            except Exception as e:
                logging.exception("Error exporting Excel report")
                return jsonify({'success': False, 'error': str(e)}), 500
        
        elif format.lower() == 'pdf':
            # PDF export would require reportlab or similar
            # For now, return JSON with note that PDF requires additional library
            return jsonify({
                'success': False,
                'error': 'PDF export requires reportlab library. Use CSV or JSON export instead.',
                'available_formats': ['csv', 'json', 'excel']
            }), 501
        
        else:
            return jsonify({'success': False, 'error': f'Unsupported format: {format}. Supported: csv, json, excel'}), 400
        
    except Exception as e:
        logging.exception("Error exporting smart execution report")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/smart-execution/<execution_id>/children', methods=['GET'])
def get_smart_execution_children(execution_id):
    """Get child executions for a parent smart execution"""
    try:
        from database import SessionLocal
        from sqlalchemy import text
        import json as json_lib
        
        session = SessionLocal()
        
        query = text("""
            SELECT execution_id, status, start_time, end_time, total_operations,
                   successful_operations, failed_operations, success_rate,
                   duration_minutes, target_config, threshold_reached
            FROM smart_executions
            WHERE target_config::jsonb @> :filter
            ORDER BY start_time
        """)
        
        result = session.execute(query, {'filter': json_lib.dumps({'parent_execution_id': execution_id})})
        
        children = []
        for row in result:
            tc = row[9]
            children.append({
                'execution_id': row[0],
                'status': row[1],
                'start_time': row[2].isoformat() if row[2] else None,
                'end_time': row[3].isoformat() if row[3] else None,
                'total_operations': row[4],
                'successful_operations': row[5],
                'failed_operations': row[6],
                'success_rate': row[7],
                'duration_minutes': row[8],
                'target_config': tc,
                'threshold_reached': row[10]
            })
        
        session.close()
        
        return jsonify({
            'success': True,
            'parent_execution_id': execution_id,
            'child_count': len(children),
            'children': children
        }), 200
        
    except Exception as e:
        logging.error(f"Error fetching children: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/smart-execution/report/<execution_id>', methods=['GET'])
def get_smart_execution_report(execution_id):
    """
    Get detailed report of a smart execution
    
    Response:
    {
        "success": true,
        "execution_id": "SMART-...",
        "testbed": "Prod-NCM-Cluster",
        "status": "COMPLETED",
        "duration_minutes": 12.5,
        "total_operations": 45,
        "successful_operations": 43,
        "failed_operations": 2,
        "success_rate": 95.5,
        "baseline_metrics": {...},
        "final_metrics": {...},
        "target_config": {...},
        "entity_breakdown": {...},
        "operations_history": [...],
        "metrics_history": [...]
    }
    """
    try:
        from services.smart_execution_service import get_smart_execution
        from services.smart_execution_db import load_smart_execution
        
        controller = get_smart_execution(execution_id)
        
        if controller:
            report = controller.get_report()
            report['success'] = True
            # Ensure testbed_label is always set for frontend
            if not report.get('testbed_label'):
                report['testbed_label'] = report.get('testbed', report.get('execution_context', {}).get('testbed_label', 'Unknown'))
            # Ensure current_metrics is always set (fallback to final_metrics)
            if not report.get('current_metrics'):
                report['current_metrics'] = report.get('final_metrics', {'cpu_percent': 0, 'memory_percent': 0})
            return jsonify(report), 200
        
        # Try loading from database
        db_data = load_smart_execution(execution_id)
        if db_data:
            # Fetch child executions
            from database import SessionLocal
            from sqlalchemy import text
            
            child_executions = []
            try:
                session = SessionLocal()
                filter_json = json.dumps({'parent_execution_id': execution_id})
                logging.info(f"Searching for children of {execution_id}")
                
                query = text("""
                    SELECT execution_id, status, start_time, end_time, total_operations,
                           successful_operations, failed_operations, success_rate,
                           duration_minutes, target_config, entity_breakdown,
                           operations_history, baseline_metrics, final_metrics
                    FROM smart_executions
                    WHERE target_config::jsonb @> :filter
                    ORDER BY start_time
                """)
                result = session.execute(query, {'filter': filter_json})
                
                for row in result:
                    target_config = row[9]
                    if isinstance(target_config, dict):
                        child_executions.append({
                            'execution_id': row[0],
                            'status': row[1],
                            'start_time': row[2].isoformat() if row[2] else None,
                            'end_time': row[3].isoformat() if row[3] else None,
                            'total_operations': row[4],
                            'successful_operations': row[5],
                            'failed_operations': row[6],
                            'success_rate': row[7],
                            'duration_minutes': row[8],
                            'target_config': row[9],
                            'entity_breakdown': row[10],
                            'operations_history': row[11],
                            'baseline_metrics': row[12],
                            'final_metrics': row[13]
                        })
                
                session.close()
            except Exception as e:
                logging.error(f"Error fetching child executions: {e}")
            
            testbed_label = db_data.get('testbed_label', 'Unknown')
            final_metrics = db_data.get('final_metrics') or {}
            baseline_metrics = db_data.get('baseline_metrics') or {}
            
            # Convert to report format
            report = {
                'success': True,
                'execution_id': db_data.get('execution_id'),
                'testbed': testbed_label,
                'testbed_label': testbed_label,
                'status': db_data.get('status', 'UNKNOWN'),
                'start_time': db_data.get('start_time'),
                'end_time': db_data.get('end_time'),
                'duration_minutes': db_data.get('duration_minutes', 0) or 0,
                'total_operations': db_data.get('total_operations', 0),
                'successful_operations': db_data.get('successful_operations', 0),
                'failed_operations': db_data.get('failed_operations', 0),
                'success_rate': db_data.get('success_rate', 0) or 0,
                'operations_per_minute': db_data.get('operations_per_minute', 0) or 0,
                'baseline_metrics': baseline_metrics,
                'final_metrics': final_metrics,
                'current_metrics': final_metrics if final_metrics else {'cpu_percent': 0, 'memory_percent': 0},
                'target_config': db_data.get('target_config') or {},
                'entity_breakdown': db_data.get('entity_breakdown') or {},
                'operations_history': db_data.get('operations_history') or [],
                'metrics_history': db_data.get('metrics_history') or [],
                'threshold_reached': db_data.get('threshold_reached', False),
                'created_entities': db_data.get('created_entities') or [],
                'child_executions': child_executions,
                'testbed_info': {
                    'testbed_label': testbed_label,
                    'testbed_id': db_data.get('testbed_id')
                }
            }
            # Include AI/ML fields
            report['ai_enabled'] = db_data.get('ai_enabled', False)
            report['ai_settings'] = db_data.get('ai_settings')
            report['ml_stats'] = db_data.get('ml_stats')
            report['pid_stats'] = db_data.get('pid_stats')
            report['training_data_collected'] = db_data.get('training_data_collected', 0)
            
            # Generate AI insights
            if report['ai_enabled'] or report['pid_stats'] or report['ml_stats']:
                ai_insights = {}
                
                # PID Performance
                if report['pid_stats']:
                    ai_insights['pid_performance'] = report['pid_stats']
                
                # ML Performance
                if report['ml_stats']:
                    ai_insights['ml_performance'] = report['ml_stats']
                
                # Recommendations
                recommendations = []
                if report['ml_stats'] and report['ml_stats'].get('model_trained'):
                    recommendations.append("✅ ML model successfully trained and ready for future executions")
                if report['pid_stats'] and report['pid_stats'].get('phase') == 'maintain':
                    recommendations.append("✅ PID controller achieved stable target maintenance")
                if report['success_rate'] and report['success_rate'] > 90:
                    recommendations.append("✅ High success rate indicates stable execution")
                elif report['success_rate'] and report['success_rate'] < 70:
                    recommendations.append("⚠️ Low success rate - consider reviewing operation configurations")
                
                if report['training_data_collected'] and report['training_data_collected'] > 50:
                    recommendations.append(f"📊 Collected {report['training_data_collected']} training samples for ML improvement")
                
                if recommendations:
                    ai_insights['recommendations'] = recommendations
                
                report['ai_insights'] = ai_insights
            
            # Include full execution data if available
            full_data = db_data.get('full_execution_data', {})
            if full_data:
                report.update({
                    'predictions': full_data.get('predictions'),
                    'detected_anomalies': full_data.get('detected_anomalies', []),
                    'recommendations': full_data.get('recommendations', []),
                    'operation_effectiveness': full_data.get('operation_effectiveness', []),
                    'analysis': full_data.get('analysis', {}),
                    'network': full_data.get('current_metrics', {}).get('network'),
                    'disk': full_data.get('current_metrics', {}).get('disk'),
                    'latency': full_data.get('current_metrics', {}).get('latency')
                })
            return jsonify(report), 200
        
        return jsonify({'success': False, 'error': 'Execution not found'}), 404
        
    except Exception as e:
        logging.exception("Error getting smart execution report")
        return jsonify({'success': False, 'error': str(e)}), 500


# ===========================
# Phase 3: Execution Templates APIs
# ===========================

@app.route('/api/smart-execution/templates', methods=['GET'])
def get_execution_templates():
    """
    Phase 3: Get all execution templates
    """
    try:
        from models.execution_template import ExecutionTemplate
        from database import SessionLocal
        
        session = SessionLocal()
        try:
            templates = session.query(ExecutionTemplate).order_by(ExecutionTemplate.created_at.desc()).all()
            return jsonify({
                'success': True,
                'templates': [t.to_dict() for t in templates],
                'total': len(templates)
            }), 200
        finally:
            session.close()
    except Exception as e:
        logging.exception("Error fetching execution templates")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/smart-execution/templates', methods=['POST'])
def save_execution_template():
    """
    Phase 3: Save an execution template
    
    Request:
    {
        "template_name": "High CPU Load Test",
        "description": "Template for testing high CPU scenarios",
        "target_config": {...},
        "entities_config": {...},
        "advanced_settings": {...}
    }
    """
    try:
        from models.execution_template import ExecutionTemplate
        from database import SessionLocal
        import uuid
        
        data = request.get_json()
        template_name = data.get('template_name')
        target_config = data.get('target_config')
        entities_config = data.get('entities_config')
        
        if not template_name or not target_config or not entities_config:
            return jsonify({'success': False, 'error': 'template_name, target_config, and entities_config are required'}), 400
        
        session = SessionLocal()
        try:
            template_id = f"TEMPLATE-{uuid.uuid4().hex[:12]}"
            
            template = ExecutionTemplate(
                template_id=template_id,
                template_name=template_name,
                description=data.get('description'),
                target_config=target_config,
                entities_config=entities_config,
                advanced_settings=data.get('advanced_settings'),
                created_by=data.get('created_by', 'system')
            )
            
            session.add(template)
            session.commit()
            
            return jsonify({
                'success': True,
                'template_id': template_id,
                'message': 'Template saved successfully'
            }), 201
        finally:
            session.close()
    except Exception as e:
        logging.exception("Error saving execution template")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/smart-execution/templates/<template_id>', methods=['GET'])
def get_execution_template(template_id):
    """Get a specific execution template"""
    try:
        from models.execution_template import ExecutionTemplate
        from database import SessionLocal
        
        session = SessionLocal()
        try:
            template = session.query(ExecutionTemplate).filter_by(template_id=template_id).first()
            if not template:
                return jsonify({'success': False, 'error': 'Template not found'}), 404
            
            # Increment usage count
            template.usage_count += 1
            session.commit()
            
            return jsonify({
                'success': True,
                'template': template.to_dict()
            }), 200
        finally:
            session.close()
    except Exception as e:
        logging.exception("Error fetching execution template")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/smart-execution/templates/<template_id>', methods=['DELETE'])
def delete_execution_template(template_id):
    """Delete an execution template"""
    try:
        from models.execution_template import ExecutionTemplate
        from database import SessionLocal
        
        session = SessionLocal()
        try:
            template = session.query(ExecutionTemplate).filter_by(template_id=template_id).first()
            if not template:
                return jsonify({'success': False, 'error': 'Template not found'}), 404
            
            session.delete(template)
            session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Template deleted successfully'
            }), 200
        finally:
            session.close()
    except Exception as e:
        logging.exception("Error deleting execution template")
        return jsonify({'success': False, 'error': str(e)}), 500


_shutdown_in_progress = False  # Global flag to prevent recursive shutdown


def graceful_shutdown(signum=None, frame=None):
    """
    Graceful shutdown handler for SIGINT (Ctrl+C) and SIGTERM.
    
    Cleans up resources before exiting:
    - Marks active executions as stopped
    - Stops scheduler
    - Closes database connection pool
    - Logs shutdown event
    
    Args:
        signum: Signal number (if called by signal handler)
        frame: Current stack frame (if called by signal handler)
    """
    global _shutdown_in_progress
    
    # Prevent recursive shutdown calls
    if _shutdown_in_progress:
        return
    _shutdown_in_progress = True
    
    signal_name = signal.Signals(signum).name if signum else "unknown"
    logging.info(f"\n🛑 Received shutdown signal ({signal_name}) - initiating graceful shutdown...")
    
    try:
        # 1. Stop scheduler to prevent new jobs from starting
        try:
            logging.info("   ↳ Stopping scheduler...")
            if scheduler_service:
                scheduler_service.stop()
                logging.info("   ✅ Scheduler stopped")
            else:
                logging.info("   ℹ️  Scheduler was not initialized")
        except Exception as e:
            logging.error(f"   ⚠️  Error stopping scheduler: {e}")
        
        # 2. Mark active executions as stopped
        try:
            logging.info("   ↳ Marking active executions as stopped...")
            from services.execution_manager import get_execution_manager
            manager = get_execution_manager()
            
            active_count = len(manager.active_executions)
            if active_count > 0:
                logging.info(f"   Found {active_count} active execution(s) to stop")
                
                # Mark each active execution as stopped
                for exec_id in list(manager.active_executions.keys()):
                    try:
                        context = manager.active_executions[exec_id]
                        context.status = 'STOPPED'
                        context.stopped_at = datetime.utcnow()
                        
                        # Update in database
                        from database import update_execution_status
                        update_execution_status(
                            exec_id,
                            status='STOPPED',
                            last_error='Backend shutdown - execution interrupted',
                            end_time=datetime.utcnow()
                        )
                        logging.info(f"   ✅ Stopped execution {exec_id[:20]}...")
                    except Exception as e:
                        logging.error(f"   ⚠️  Error stopping execution {exec_id}: {e}")
                
                logging.info(f"   ✅ Marked {active_count} execution(s) as stopped")
            else:
                logging.info("   ℹ️  No active executions to stop")
        except Exception as e:
            logging.error(f"   ⚠️  Error marking executions as stopped: {e}")
        
        # 3. Close database connection pool
        try:
            logging.info("   ↳ Closing database connection pool...")
            from database import engine
            engine.dispose()
            logging.info("   ✅ Database connections closed")
        except Exception as e:
            logging.error(f"   ⚠️  Error closing database: {e}")
        
        # 4. Log final status
        logging.info("✅ Graceful shutdown complete")
        logging.info("👋 NMT Backend stopped\n")
        
    except Exception as e:
        logging.error(f"❌ Error during graceful shutdown: {e}", exc_info=True)
    finally:
        # Only exit if called by signal handler (not atexit)
        if signum is not None:
            sys.exit(0)


# ============================================================================
# WEBSOCKET EVENT HANDLERS (LIVE LOGS)
# ============================================================================

@socketio.on('connect')
def handle_connect():
    """Handle WebSocket connection"""
    logging.info(f"✅ WebSocket client connected: {request.sid}")
    emit('connected', {'message': 'Connected to Smart Execution live logs'})

@socketio.on('disconnect')
def handle_disconnect():
    """Handle WebSocket disconnection"""
    logging.info(f"❌ WebSocket client disconnected: {request.sid}")

@socketio.on('subscribe')
def handle_subscribe(data):
    """Subscribe to a specific execution's logs"""
    execution_id = data.get('execution_id')
    logging.info(f"📡 Client {request.sid} subscribed to execution {execution_id}")
    emit('subscribed', {'execution_id': execution_id})

# Global socketio reference for smart_execution_service
def get_socketio():
    """Get the SocketIO instance for broadcasting logs"""
    return socketio

# ============================================================================
# MAIN
# ============================================================================

if __name__ == '__main__':
    # Register shutdown handlers
    signal.signal(signal.SIGINT, graceful_shutdown)   # Ctrl+C
    signal.signal(signal.SIGTERM, graceful_shutdown)  # Kill signal
    atexit.register(lambda: graceful_shutdown())      # Normal exit
    
    logging.info("🔒 Graceful shutdown handlers registered (Ctrl+C will clean up properly)")

    # Initialize and start the scheduler
    scheduler_service = get_scheduler()
    if not scheduler_service:
        logging.warning("⚠️  Scheduler service failed to initialize - scheduled executions will not work")
    else:
        scheduler_service.start()
    
    # Add background jobs only if scheduler is available
    if scheduler_service and scheduler_service.scheduler:
        # Add background JITA job monitoring (runs every 30 seconds)
        scheduler_service.scheduler.add_job(
            monitor_jita_jobs_background,
            'interval',
            seconds=30,
            id='jita_background_monitor',
            name='Background JITA Job Monitor',
            replace_existing=True
        )
        
        # Add background Prometheus job monitoring (runs every 30 seconds)
        scheduler_service.scheduler.add_job(
            monitor_prometheus_jobs_background,
            'interval',
            seconds=30,
            id='prometheus_background_monitor',
            name='Background Prometheus Job Monitor',
            replace_existing=True
        )
        
        # Add continuous metrics collection (runs every 60 seconds)
        scheduler_service.scheduler.add_job(
            collect_continuous_testbed_metrics,
            'interval',
            seconds=60,
            id='continuous_metrics_collector',
            name='Continuous Testbed Metrics Collector',
            replace_existing=True
        )
    
    logging.info("✅ All background jobs scheduled:")
    logging.info("   - JITA Monitor (30s interval)")
    logging.info("   - Prometheus Monitor (30s interval)")
    logging.info("   - Continuous Metrics Collector (60s interval)")
    
    try:
        socketio.run(app, host='0.0.0.0', port=5000, allow_unsafe_werkzeug=True)  # Enable WebSocket support
    except KeyboardInterrupt:
        # Handled by graceful_shutdown signal handler
        pass
    except Exception as e:
        logging.error(f"❌ Unexpected error in main: {e}", exc_info=True)
        graceful_shutdown()

